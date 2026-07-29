"""
_cota_import.py — Mede no retorno do ADMS quantos sinais cada tipo de
dispositivo REALMENTE aceita, por Signal Type.

O validador recusa com "Same signal X ... was already mapped on device: D in
same mapping action" quando se manda sinais demais do mesmo papel para o mesmo
dispositivo. Nao ha documentacao do limite: a unica fonte confiavel e o proprio
retorno de um import. Este script le o CSV de erros junto com a TDT que o
gerou e conta, por dispositivo, quantos sinais de cada Signal Type PASSARAM.

Uso: python _cota_import.py "C:/Users/egnpo/Downloads/ERROS 98.csv" \
                            "C:/Users/egnpo/Downloads/TDT_CASCA_ATUAL_COMPLETA.xlsx"
Saida: data/cota_import.json   {"DJ|RelayTrip": 5, "PROT|Custom": 3, ...}
"""
from __future__ import annotations
import collections, csv, io, json, sys
from pathlib import Path
import openpyxl
import casca_devmap as devmap

DATA = Path(__file__).parent / "data"
OUT = DATA / "cota_import.json"
HR = 4
ABAS = ("DNP3_DiscreteSignals", "DNP3_AnalogSignals", "DNP3_DiscreteAnalog")


def ler_csv(p: Path) -> dict[str, bool]:
    """Signal Name -> passou?"""
    with io.open(p, encoding="utf-8-sig", newline="") as f:
        linhas = list(csv.reader(f, delimiter=";"))
    return {r[1]: not r[0].startswith("Falh")
            for r in linhas[1:] if len(r) > 3 and r[1]}


def main(csv_path: str, tdt_path: str):
    ok = ler_csv(Path(csv_path))
    wb = openpyxl.load_workbook(tdt_path, read_only=True, data_only=True)
    # (dispositivo, Signal Type) -> quantos PASSARAM
    passou = collections.Counter()
    tentou = collections.Counter()
    for sn in ABAS:
        if sn not in wb.sheetnames:
            continue
        linhas = list(wb[sn].iter_rows(values_only=True))
        hdr = list(linhas[HR - 1])
        ix = {n: i for i, n in enumerate(hdr) if n}
        for r in linhas[HR:]:
            nome = r[ix["Signal Name"]]
            if not nome:
                continue
            dm = r[ix["Device Mapping"]]
            st = r[ix.get("Signal Type", 0)] if "Signal Type" in ix else ""
            tentou[(dm, st)] += 1
            if ok.get(nome):
                passou[(dm, st)] += 1
    wb.close()

    # a COTA de um tipo de dispositivo e o MAIOR numero de sinais de um mesmo
    # Signal Type que um dispositivo daquele tipo aceitou. Onde ninguem forcou
    # o limite (tentou == passou), o numero e piso, nao teto — marcado com '+'
    # so no relatorio; o JSON guarda o piso mesmo.
    cota = {}
    forcado = set()
    for (dm, st), q in passou.items():
        if not dm or not st:
            continue
        k = f"{devmap.tipo_dispositivo(dm)}|{st}"
        cota[k] = max(cota.get(k, 0), q)
        if tentou[(dm, st)] > q:
            forcado.add(k)

    # SO VALE COMO LIMITE o par em que o ADMS REALMENTE recusou o excedente.
    # Onde ninguem forcou, o numero e PISO, nao teto — grava-lo como cota
    # bloquearia o dispositivo antes de ele ser testado. Foi exatamente o que
    # aconteceu com PROT|RelayTrip=1: veio de reles que so tinham 1 sinal, e
    # como teto impediu o roteamento para o _PROT generico.
    confirmada = {k: v for k, v in cota.items() if k in forcado}

    # ACUMULA com o que ja se sabia. Um limite confirmado num import continua
    # valendo no proximo, mesmo que la ninguem tenha chegado nele — foi o que
    # aconteceu com DJ|RelayTrip=5: depois que o gerador passou a respeita-lo,
    # o import seguinte nao forcou mais o disjuntor e o limite sumiria daqui.
    # Entre dois valores confirmados vale o MENOR (o teto mais apertado).
    antes = {}
    if OUT.exists():
        antes = json.loads(OUT.read_text(encoding="utf-8"))
    novos = [k for k in confirmada if k not in antes]
    for k, v in antes.items():
        confirmada[k] = min(confirmada[k], v) if k in confirmada else v

    DATA.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(dict(sorted(confirmada.items())),
                              ensure_ascii=False, indent=1), encoding="utf-8")
    if novos:
        print(f"limites NOVOS neste import: {', '.join(novos)}")
    print(f"limites conhecidos ao todo: {len(confirmada)} -> "
          f"{json.dumps(dict(sorted(confirmada.items())))}\n")
    print(f"{OUT.name}: {len(cota)} pares (tipo de dispositivo, Signal Type)\n")
    print(f"{'par':<28}{'cota':>6}   evidencia")
    for k, v in sorted(cota.items(), key=lambda x: -x[1]):
        marca = "LIMITE ATINGIDO (o ADMS recusou o excedente)" if k in forcado \
            else "piso — ninguem tentou mais que isso"
        print(f"  {k:<26}{v:>4}   {marca}")


if __name__ == "__main__":
    a = sys.argv[1:]
    main(a[0] if a else "C:/Users/egnpo/Downloads/ERROS 98.csv",
         a[1] if len(a) > 1 else
         "C:/Users/egnpo/Downloads/TDT_CASCA_ATUAL_COMPLETA.xlsx")
