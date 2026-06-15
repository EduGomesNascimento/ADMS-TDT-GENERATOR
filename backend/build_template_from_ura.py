"""
build_template_from_ura.py — Reconstrói reference_template.xlsx a partir de uma
TDT real de transformador (TDT_DNP3_UTR_URA_TR2.xlsx), que é um export válido do
ADMS com as 3 sheets de sinais corretas (incluindo DNP3_DiscreteAnalog de 48 col
com tabela) + DMSMatchingTemplateInfo.

Mantém: cabeçalhos (4 linhas), tabelas (ListObjects), DMSMatchingTemplateInfo.
Limpa: as linhas de dados (a partir da linha 6), preservando a linha 5 como
linha-modelo de estilo (conteúdo limpo, formatação mantida).

A formatação nativa final é aplicada no export pelo excel_native (re-save MS Excel).
Uso: python build_template_from_ura.py
"""
import io
import os
import shutil
import warnings

import openpyxl

warnings.filterwarnings("ignore")

ROOT = os.path.dirname(os.path.abspath(__file__))
SOURCE = r"C:\Users\egnpo\Downloads\TDT_DNP3_UTR_URA_TR2.xlsx"
OUT = os.path.join(ROOT, "data", "reference_template.xlsx")
HEADER_ROWS = 4
SIGNAL_SHEETS = ["DNP3_DiscreteSignals", "DNP3_AnalogSignals", "DNP3_DiscreteAnalog"]


def _clear_data(ws):
    """Limpa o conteúdo das linhas de dados, mantém linha 5 como modelo de estilo,
    e remove as demais (6+). Ajusta o ref da tabela para header + 1."""
    last = ws.max_row
    # limpa conteúdo da linha 5 (mantém formatação para servir de modelo)
    if last >= HEADER_ROWS + 1:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=HEADER_ROWS + 1, column=c).value = None
    # remove linhas 6+
    if last > HEADER_ROWS + 1:
        ws.delete_rows(HEADER_ROWS + 2, last - (HEADER_ROWS + 1))
    # ajusta ref das tabelas para A4:<col>5
    import re
    for name in list(ws.tables.keys()):
        tbl = ws.tables[name]
        m = re.match(r"^([A-Z]+)(\d+):([A-Z]+)\d+$", str(tbl.ref))
        if m:
            tbl.ref = f"{m.group(1)}{m.group(2)}:{m.group(3)}{HEADER_ROWS + 1}"


def main():
    if not os.path.exists(SOURCE):
        raise SystemExit(f"Fonte não encontrada: {SOURCE}")

    # backup
    if os.path.exists(OUT):
        shutil.copy(OUT, OUT.replace(".xlsx", ".prev.xlsx"))

    wb = openpyxl.load_workbook(SOURCE)
    # mantém só as sheets necessárias, na ordem canônica
    keep = ["DMSMatchingTemplateInfo"] + SIGNAL_SHEETS
    for sn in list(wb.sheetnames):
        if sn not in keep:
            del wb[sn]

    for sn in SIGNAL_SHEETS:
        if sn in wb.sheetnames:
            _clear_data(wb[sn])

    # ordena sheets
    wb._sheets.sort(key=lambda s: keep.index(s.title) if s.title in keep else 99)

    wb.save(OUT)

    # validação
    wb2 = openpyxl.load_workbook(OUT)
    print("Sheets:", wb2.sheetnames)
    for sn in SIGNAL_SHEETS:
        ws = wb2[sn]
        tabs = {n: ws.tables[n].ref for n in ws.tables.keys()}
        print(f"  {sn}: {ws.max_row}x{ws.max_column}  tabelas={tabs}")
    print(f"OK — template reconstruído em {OUT}")


if __name__ == "__main__":
    main()
