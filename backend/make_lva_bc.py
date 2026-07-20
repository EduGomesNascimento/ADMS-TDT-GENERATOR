"""
make_lva_bc.py — TDT do BANCO DE CAPACITOR da SE LVA (módulo BC1, disjuntor 52-26).

Fontes:
  - Lista de pontos LVA (abas IHM/PAS, topologia BANCO DE CAPACITOR): define o
    conjunto de pontos reais (trips 27/59/61N, 86, automatismo, aux do DJ, 14 medições)
  - TDT v3 (AL21) como esqueleto/estilo: analógicos e sinais comuns do disjuntor
    clonados do AL21 (identidade BC1/52-26, offset de bloco)
  - sigla_index (full base): templates dos sinais específicos de BC (27E1, 59E1,
    61N, 61, 27 incl/excl, AUTO, 86, TRIP)

Bloco de índices: 900 (próximo na progressão AL11=0..AL24=800).
Regras herdadas: escala P/Q/S=1000, RPCID={nome}_UTR_LVA_2, SCID limpo.

Uso: python make_lva_bc.py
"""
from __future__ import annotations
import io, json
from copy import copy
from pathlib import Path
import openpyxl
import excel_native

SRC = Path("C:/Users/egnpo/Downloads/TDT_LVA_20260720_v3.xlsx")
OUT = Path("C:/Users/egnpo/Downloads/TDT_LVA_BC.xlsx")
DATA = Path(__file__).parent / "data"
RU = "UTR_LVA_2"
HEADER_ROWS = 4
BASE = 900                      # bloco do BC
ALIAS, MODULE, DEVICE = "LVA", "BC1", "52-26"
PREFIX = f"{ALIAS}_{MODULE}_{DEVICE}"
SCALE_1000 = {"P", "Q", "S"}

# sinais do BC clonados do AL21 (mesma linha/estilo, identidade trocada)
FROM_AL21_DIG = ["DJF1", "MOLA", "43LR", "CCMO", "CCCO", "CAFL", "CAB",
                 "43TC", "FCOM", "2649", "62BF", "50F", "50N", "51F", "51N1"]
# sinais específicos de BC vindos do template da base (lista LVA → SIGLA):
#   Trip 27→27E1 · Trip 59→59E1 · Trip 61N→61N · Trip 61N Temporizado→61 ·
#   Comando 27_59 Incl/Excl→27 · Automatismo→AUTO · Relé 86BC→86 · 74TC Superv.→TRIP
FROM_BASE_DIG = ["27E1", "59E1", "61N", "61", "27", "AUTO", "86", "TRIP"]

# pontos da lista SEM sigla firme na base — ficam de fora e são reportados
PENDING = ["Secc 29-10 (DP + comando)", "Comando Rearme 86BC", "MCB TC Aberto",
           "Programador Horário", "Pickups 50N/61N/61NT", "Retrip 62BF",
           "Diagnósticos do relé (bateria/canal óptico/GPS/memória/troca ajuste)"]


def _subst(v, mapping):
    if not isinstance(v, str):
        return v
    for k, val in mapping.items():
        v = v.replace(k, val)
    return v


def _offset(val, delta):
    if val in (None, ""):
        return val
    parts = str(val).split(";")
    if not all(p.strip().lstrip("-").isdigit() for p in parts):
        return val
    out = [str(int(p) + delta) for p in parts]
    return ";".join(out) if len(out) > 1 else int(out[0])


def main():
    idx = json.loads((DATA / "sigla_index.json").read_text(encoding="utf-8"))
    dig_tpl = idx["DNP3_DiscreteSignals"]
    wb = openpyxl.load_workbook(SRC)
    mapping = {"<<PREFIX>>": PREFIX, "<<ALIAS>>": ALIAS,
               "<<MODULE>>": MODULE, "<<DEVICE>>": DEVICE, "<<N>>": "1"}

    for sheet in ("DNP3_DiscreteSignals", "DNP3_AnalogSignals"):
        ws = wb[sheet]
        lab = {ws.cell(HEADER_ROWS, c).value: c for c in range(1, ws.max_column + 1)
               if ws.cell(HEADER_ROWS, c).value}
        cN = lab["Signal Name"]; cIN = lab.get("Input Coordinates")
        cOUT = lab.get("Output Coordinates"); cRPC = lab.get("Remote Point Custom ID")
        cSC = lab.get("Signal Custom ID"); cESC = lab.get("Scaling Factor")
        cRU = lab.get("Remote Unit"); cRPN = lab.get("Remote Point Name")
        cAOR = lab.get("Signal AOR Group")
        ncol = ws.max_column
        delta = BASE - 500                      # AL21 = bloco 500

        # captura linhas do AL21 (fonte) + estilo + AOR real
        al21, styles, aor_val = {}, None, None
        for r in range(HEADER_ROWS + 1, ws.max_row + 1):
            nm = ws.cell(r, cN).value
            if nm and "_AL21_" in str(nm):
                suf = str(nm).split("_", 3)[-1]
                al21[suf] = [ws.cell(r, c).value for c in range(1, ncol + 1)]
                if styles is None:
                    styles = [copy(ws.cell(r, c)._style) for c in range(1, ncol + 1)]
                if aor_val is None and cAOR:
                    aor_val = ws.cell(r, cAOR).value

        rows = []
        if sheet == "DNP3_AnalogSignals":
            # medições do BC = as 14 do AL21 (mesmo conjunto na lista LVA)
            for suf, vals in al21.items():
                row = list(vals)
                for c in range(ncol):
                    if isinstance(row[c], str):
                        row[c] = row[c].replace("AL21", MODULE).replace("52-21", DEVICE)
                row[cIN - 1] = _offset(row[cIN - 1], delta)
                if cESC and suf in SCALE_1000:
                    row[cESC - 1] = 1000
                rows.append(row)
        else:
            seq_in, seq_out = BASE, BASE
            for suf in FROM_AL21_DIG + FROM_BASE_DIG:
                if suf in al21:
                    row = list(al21[suf])
                    for c in range(ncol):
                        if isinstance(row[c], str):
                            row[c] = row[c].replace("AL21", MODULE).replace("52-21", DEVICE)
                    # re-sequencia dentro do bloco do BC
                    old_in = str(al21[suf][cIN - 1] or "")
                    n_coords = len(old_in.split(";")) if old_in else 1
                    if old_in:
                        if n_coords == 2:
                            row[cIN - 1] = f"{seq_in};{seq_in + 1}"; seq_in += 2
                        else:
                            row[cIN - 1] = seq_in; seq_in += 1
                    if cOUT and al21[suf][cOUT - 1] not in (None, ""):
                        row[cOUT - 1] = f"{seq_out};{seq_out}"; seq_out += 1
                else:
                    tpl = dig_tpl[suf]
                    row = [_subst(v, mapping) for v in tpl]
                    while len(row) < ncol:
                        row.append(None)
                    row = row[:ncol]
                    if cRU and isinstance(row[cRU - 1], str):
                        row[cRU - 1] = RU
                    if cAOR and aor_val:
                        row[cAOR - 1] = aor_val
                    row[cIN - 1] = seq_in; seq_in += 1
                    # comando: só se o template é comandável (Control Codes)
                    if cOUT and row[cOUT - 1] not in (None, ""):
                        row[cOUT - 1] = f"{seq_out};{seq_out}"; seq_out += 1
                nome = row[cN - 1]
                if cRPN:
                    row[cRPN - 1] = nome
                rows.append(row)

        # RPCID nominal + Signal Custom ID limpo (objetos novos)
        for row in rows:
            if cRPC:
                row[cRPC - 1] = f"{row[cN - 1]}_{RU}"
            if cSC:
                row[cSC - 1] = None

        if ws.max_row > HEADER_ROWS:
            ws.delete_rows(HEADER_ROWS + 1, ws.max_row - HEADER_ROWS)
        for i, row in enumerate(rows):
            for c in range(ncol):
                cell = ws.cell(HEADER_ROWS + 1 + i, c + 1, value=row[c])
                if styles:
                    cell._style = copy(styles[c])

    buf = io.BytesIO()
    wb.save(buf)
    data = excel_native.resave_native(buf.getvalue())
    OUT.write_bytes(data)
    print(f"TDT BC salva: {OUT.name} ({len(data)} bytes) | bloco {BASE}+")
    print("PENDENTES (sem sigla firme, decidir depois):")
    for p in PENDING:
        print(f"   - {p}")


if __name__ == "__main__":
    main()
