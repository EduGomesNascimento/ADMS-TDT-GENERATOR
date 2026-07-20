"""
make_lva_completa.py — TDT ÚNICA da LVA: união dos 10 módulos já gerados
(AL11..AL15, AL21..AL24, BC2), concatenados no esqueleto da v3.

Os arquivos separados continuam valendo; esta é a consolidada.
Uso: python make_lva_completa.py
"""
from __future__ import annotations
import io
from copy import copy
from pathlib import Path
import openpyxl
import excel_native

SRC = Path("C:/Users/egnpo/Downloads/TDT_LVA_20260720_v3.xlsx")
OUT = Path("C:/Users/egnpo/Downloads/TDT_LVA_COMPLETA.xlsx")
FILES = [f"TDT_LVA_{m}.xlsx" for m in
         ["AL11", "AL12", "AL13", "AL14", "AL15",
          "AL21", "AL22", "AL23", "AL24", "BC2"]]
HEADER_ROWS = 4
SHEETS = ("DNP3_DiscreteSignals", "DNP3_AnalogSignals")


def main():
    wb = openpyxl.load_workbook(SRC)
    # coleta linhas de todos os módulos
    all_rows = {sh: [] for sh in SHEETS}
    for f in FILES:
        src = openpyxl.load_workbook(Path("C:/Users/egnpo/Downloads") / f,
                                     read_only=True, data_only=True)
        for sh in SHEETS:
            ws = src[sh]
            ncol = ws.max_column
            for r in ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True):
                if r and r[0]:
                    all_rows[sh].append(list(r[:ncol]))
        src.close()

    for sh in SHEETS:
        ws = wb[sh]
        ncol = ws.max_column
        styles = [copy(ws.cell(HEADER_ROWS + 1, c)._style) for c in range(1, ncol + 1)]
        if ws.max_row > HEADER_ROWS:
            ws.delete_rows(HEADER_ROWS + 1, ws.max_row - HEADER_ROWS)
        for i, row in enumerate(all_rows[sh]):
            while len(row) < ncol:
                row.append(None)
            for c in range(ncol):
                cell = ws.cell(HEADER_ROWS + 1 + i, c + 1, value=row[c])
                cell._style = copy(styles[c])
        print(f"{sh}: {len(all_rows[sh])} linhas")

    buf = io.BytesIO()
    wb.save(buf)
    data = excel_native.resave_native(buf.getvalue())
    OUT.write_bytes(data)
    print(f"salva: {OUT.name} ({len(data)} bytes)")


if __name__ == "__main__":
    main()
