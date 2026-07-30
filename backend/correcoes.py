"""
correcoes.py — A TABELA DO QUE FAZER NO ADMS.

Le o CSV de retorno do import e produz UMA LINHA POR CORRECAO, ordenada por
quantos sinais cada uma destrava. Cada linha diz: o que fazer, onde, qual o ID
atual, qual o ID correto e quais sinais dependem daquilo.

O CSV do ADMS vem ora em portugues com ';' (Severidade/Falhou), ora em ingles
com ',' (Severity/Failed) — o formato depende do idioma da sessao. Le os dois.

Uso: python correcoes.py [caminho do csv]
Saida: C:/Users/egnpo/Downloads/CASCA_CORRECOES.xlsx
"""
from __future__ import annotations
import collections, csv, io, re, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import excel_native
import casca_devmap as devmap

DOWN = Path("C:/Users/egnpo/Downloads")
OUT = DOWN / "CASCA_CORRECOES.xlsx"
TDT = DOWN / "CAS_TDT_COMPLETA.xlsx"
HR = 4


def ler_csv(p: Path):
    """(severidade, elemento, atributo, descricao) — aceita os dois formatos."""
    txt = io.open(p, encoding="utf-8-sig", newline="").read()
    cab = txt.split("\n", 1)[0]
    delim = ";" if cab.count(";") > cab.count(",") else ","
    linhas = list(csv.reader(io.StringIO(txt), delimiter=delim))
    return [r for r in linhas[1:] if len(r) > 3 and r[1]]


def falhou(sev: str) -> bool:
    return sev.strip().lower().startswith(("falh", "fail"))


def tipo_dispositivo(dm: str) -> str:
    d = dm.replace("_NEW", "")
    if "_P_PROT_" in d:
        return "estagio do rele PRINCIPAL"
    if "_A_PROT_" in d:
        return "estagio do rele ALTERNADO"
    if "_PROT_" in d:
        return "estagio do rele"
    for suf, nome in (("_P_PROT", "rele PRINCIPAL"), ("_A_PROT", "rele ALTERNADO"),
                      ("_PROT", "rele"), ("_SEC", "seccionadora"),
                      ("_DJ", "disjuntor"), ("_TP", "transformador de potencial"),
                      ("_TC", "transformador de corrente"), ("_BP", "barra"),
                      ("_BT", "barra"), ("_TR", "transformador")):
        if d.endswith(suf):
            return nome
    return "?"


def main(csv_path: str):
    dados = ler_csv(Path(csv_path))
    falhas = [r for r in dados if falhou(r[0])]
    ok = len(dados) - len(falhas)

    # sinal -> Device Mapping que a TDT pediu
    dm_do_sinal = {}
    if TDT.exists():
        wb = openpyxl.load_workbook(TDT, read_only=True, data_only=True)
        for sn in ("DNP3_DiscreteSignals", "DNP3_AnalogSignals",
                   "DNP3_DiscreteAnalog"):
            if sn not in wb.sheetnames:
                continue
            ls = list(wb[sn].iter_rows(values_only=True))
            ix = {n: i for i, n in enumerate(ls[HR - 1]) if n}
            for r in ls[HR:]:
                if r[ix["Signal Name"]]:
                    dm_do_sinal[r[ix["Signal Name"]]] = r[ix["Device Mapping"]]
        wb.close()

    # ── agrupa as falhas por CORRECAO ────────────────────────────────────────
    # chave = (acao, alvo). O alvo e o dispositivo que precisa ser criado ou
    # corrigido — e o que a pessoa vai procurar no ADMS.
    corr = collections.defaultdict(lambda: {"sinais": [], "erro": "", "id": ""})
    for r in falhas:
        sinal, desc = r[1], r[3]
        if "Could not find any" in desc:
            m = re.search(r"Device Mapping: (\S+?)\.", desc)
            alvo = m.group(1) if m else dm_do_sinal.get(sinal, "?")
            k = ("CRIAR o dispositivo", alvo)
            corr[k]["erro"] = "Could not find any device"
        elif "Found multiple" in desc:
            m = re.search(r"Device Mapping: (\S+?)\.", desc)
            alvo = m.group(1) if m else "?"
            k = ("DAR ID PROPRIO ao 2o dispositivo", alvo)
            corr[k]["erro"] = "Found multiple devices"
        elif "already mapped" in desc:
            m = re.search(r"on device:? (\S+?) in same", desc)
            alvo = m.group(1) if m else dm_do_sinal.get(sinal, "?")
            # o que resolve nao e mexer no rele: e CRIAR o estagio da funcao
            sig = "_".join(sinal.split("_")[3:])
            k = ("CRIAR o estagio da funcao", f"{alvo}_{sig}")
            corr[k]["erro"] = "Same signal already mapped on device"
        else:
            k = ("VERIFICAR", desc[:60])
            corr[k]["erro"] = desc[:80]
        corr[k]["sinais"].append(sinal)

    # ── planilha ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E78")
    cor_acao = {"CRIAR o dispositivo": "FFC7CE",
                "CRIAR o estagio da funcao": "FFEB9C",
                "DAR ID PROPRIO ao 2o dispositivo": "BDD7EE",
                "VERIFICAR": "D9D9D9"}

    ws = wb.active
    ws.title = "CORRECOES"
    ws.append(["#", "O QUE FAZER", "ONDE (dispositivo no ADMS)",
               "Tipo de dispositivo", "Vao", "Sinais que isso destrava",
               "Sinais afetados", "Erro que o ADMS da hoje"])
    for c in ws[1]:
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    ordenado = sorted(corr.items(), key=lambda x: (-len(x[1]["sinais"]), x[0][1]))
    for i, ((acao, alvo), d) in enumerate(ordenado, 1):
        pp = str(alvo).split("_")
        ws.append([i, acao, alvo, tipo_dispositivo(str(alvo)),
                   pp[1] if len(pp) > 1 else "", len(d["sinais"]),
                   ", ".join(sorted(d["sinais"])[:6])
                   + (" ..." if len(d["sinais"]) > 6 else ""),
                   d["erro"]])
        ws.cell(ws.max_row, 2).fill = PatternFill(
            "solid", fgColor=cor_acao.get(acao, "FFFFFF"))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    for col, w in zip("ABCDEFGH", (5, 32, 40, 26, 10, 10, 60, 38)):
        ws.column_dimensions[col].width = w

    # ── resumo ───────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("0-RESUMO", 0)
    por_acao = collections.Counter()
    sin_acao = collections.Counter()
    for (acao, _a), d in corr.items():
        por_acao[acao] += 1
        sin_acao[acao] += len(d["sinais"])
    linhas = [
        ["CORRECOES NO MODELO — SE CASCA / UTR_CAS_3"], [],
        [f"Import da CAS_TDT_COMPLETA: {ok} sinais mapearam, "
         f"{len(falhas)} falharam."], [],
        ["Cada linha da aba CORRECOES e UMA acao no ADMS, ordenada por quantos"],
        ["sinais ela destrava. Nenhuma delas se resolve na TDT — o Device"],
        ["Mapping ja aponta para o nome certo; o que falta e o dispositivo."], [],
        ["ACAO", "Quantas", "Sinais que destrava"],
    ]
    for a, q in por_acao.most_common():
        linhas.append([a, q, sin_acao[a]])
    linhas += [[], ["COMO LER"], [],
               ["CRIAR o dispositivo — o ID nao existe no Cas_Obra. E o nome"],
               ["   que o dispositivo PRECISA ter para o sinal encontrar."], [],
               ["CRIAR o estagio da funcao — o rele existe, mas a funcao"],
               ["   daquele sinal nao tem estagio proprio nele. Varios sinais"],
               ["   disputando o mesmo rele e o que gera 'already mapped'."], [],
               ["DAR ID PROPRIO — dois dispositivos com o MESMO ID de"],
               ["   Mapeamento SCADA. O ADMS nao sabe qual escolher."]]
    for r in linhas:
        ws2.append(r)
    ws2["A1"].font = Font(bold=True, size=13)
    for row in ws2.iter_rows():
        if row[0].value == "ACAO":
            for c in row:
                c.font = hf; c.fill = hfill
    for col, w in zip("ABC", (46, 12, 22)):
        ws2.column_dimensions[col].width = w

    alvo_f = OUT
    try:
        wb.save(alvo_f)
    except PermissionError:
        alvo_f = OUT.with_name(OUT.stem + "_NOVA.xlsx")
        wb.save(alvo_f)
        print(f"ATENCAO: {OUT.name} aberto no Excel — gravado como {alvo_f.name}")
    alvo_f.write_bytes(excel_native.resave_native(alvo_f.read_bytes()))
    print(f"{alvo_f.name}: {len(corr)} correcoes para {len(falhas)} sinais")
    for a, q in por_acao.most_common():
        print(f"   {q:>3} x {a:<34} destrava {sin_acao[a]:>3} sinais")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else str(DOWN / "erros.csv"))
