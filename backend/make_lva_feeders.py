"""
make_lva_feeders.py — SE LVA: separa a TDT v3 em um arquivo por alimentador,
corrige as escalas de P/Q/S (=1000) e GERA os alimentadores que faltam.

Padrão descoberto na TDT v3 (validado: delta +100 exato em 39/39 siglas):
  cada alimentador ocupa um bloco de 100 índices (entrada E comando):
  AL21=500-5xx, AL22=600-6xx  →  AL11=0, AL12=100, AL13=200, AL14=300,
  AL15=400, AL21=500, AL22=600, AL23=700, AL24=800
  (a lista de pontos LVA não traz índice numérico — blocos extrapolados do padrão)

Novos alimentadores são clonados do AL21 (linhas limpas, RPCID nominal):
  - AL21→ALxx e 52-21→52-xx nos textos
  - coordenadas com offset do bloco
  - Remote Point Custom ID = {nome}_UTR_LVA_2 ; Signal Custom ID limpo
  - escala de P/Q/S = 1000 (também corrigida nos AL21/AL22 existentes)

Uso: python make_lva_feeders.py            (gera os 9)
     python make_lva_feeders.py AL13 AL24  (só os pedidos)
"""
from __future__ import annotations
import io, re, sys
from copy import copy
from pathlib import Path
import openpyxl
import excel_native

SRC = Path("C:/Users/egnpo/Downloads/TDT_LVA_20260720_v3.xlsx")
OUT_DIR = Path("C:/Users/egnpo/Downloads")
RU = "UTR_LVA_2"
HEADER_ROWS = 4
SHEETS = ("DNP3_DiscreteSignals", "DNP3_AnalogSignals")

BASES = {"AL11": 0, "AL12": 100, "AL13": 200, "AL14": 300, "AL15": 400,
         "AL21": 500, "AL22": 600, "AL23": 700, "AL24": 800}
EXISTING = ("AL21", "AL22")
SCALE_1000 = {"P", "Q", "S"}
# AJG2: Message Mapping padrão (v3 trazia G1@G2___GRUPO 1@GRUPO 2, errado)
AJG2_MM = "DESATIVAR@ATIVAR___DESATIVADO@ATIVADO___Custom_S_TC_SV"


def _labels(ws):
    return {ws.cell(HEADER_ROWS, c).value: c for c in range(1, ws.max_column + 1)
            if ws.cell(HEADER_ROWS, c).value}


def _offset_coord(val, delta: int):
    """Aplica offset preservando o formato ('n' ou 'a;b')."""
    if val in (None, ""):
        return val
    parts = str(val).split(";")
    if not all(p.strip().lstrip("-").isdigit() for p in parts):
        return val
    out = [str(int(p) + delta) for p in parts]
    return ";".join(out) if len(out) > 1 else int(out[0])


def _suffix(nome: str) -> str:
    p = str(nome).split("_")
    return "_".join(p[3:]) if len(p) > 3 else p[-1]


def make_feeder(target: str) -> bytes:
    delta = BASES[target] - BASES["AL21"]
    wb = openpyxl.load_workbook(SRC)
    for sheet in SHEETS:
        ws = wb[sheet]
        lab = _labels(ws)
        cN = lab["Signal Name"]
        cIN = lab.get("Input Coordinates")
        cOUT = lab.get("Output Coordinates")
        cRPC = lab.get("Remote Point Custom ID")
        cSC = lab.get("Signal Custom ID")
        cESC = lab.get("Scaling Factor")
        cMM = lab.get("Message Mapping")
        ncol = ws.max_column

        # linhas-fonte: as do próprio alvo (se já existe) ou as do AL21 (clone)
        src_tag = f"_{target}_" if target in EXISTING else "_AL21_"
        src_rows, styles = [], None
        for r in range(HEADER_ROWS + 1, ws.max_row + 1):
            nm = ws.cell(r, cN).value
            if nm and src_tag in str(nm):
                src_rows.append([ws.cell(r, c).value for c in range(1, ncol + 1)])
                if styles is None:
                    styles = [copy(ws.cell(r, c)._style) for c in range(1, ncol + 1)]

        # limpa TODAS as linhas de dados e reescreve só o alvo
        if ws.max_row > HEADER_ROWS:
            ws.delete_rows(HEADER_ROWS + 1, ws.max_row - HEADER_ROWS)
        for i, vals in enumerate(src_rows):
            row = list(vals)
            if target not in EXISTING:
                # troca de identidade + offset de bloco + IDs novos
                n_new = target[2:]
                for c in range(ncol):
                    if isinstance(row[c], str):
                        row[c] = row[c].replace("AL21", target).replace(
                            "52-21", f"52-{n_new}")
                if cIN:
                    row[cIN - 1] = _offset_coord(row[cIN - 1], delta)
                if cOUT and sheet == "DNP3_DiscreteSignals":
                    row[cOUT - 1] = _offset_coord(row[cOUT - 1], delta)
                if cRPC:
                    row[cRPC - 1] = f"{row[cN - 1]}_{RU}"
                if cSC:
                    row[cSC - 1] = None          # objeto novo: ADMS gera o GUID
            # escala de P/Q/S = 1000 (novos E existentes)
            if cESC and sheet == "DNP3_AnalogSignals" and _suffix(row[cN - 1]) in SCALE_1000:
                row[cESC - 1] = 1000
            # AJG2: Message Mapping padrão (novos E existentes)
            if cMM and _suffix(row[cN - 1]) == "AJG2":
                row[cMM - 1] = AJG2_MM
            excel_row = HEADER_ROWS + 1 + i
            for c in range(ncol):
                cell = ws.cell(excel_row, c + 1, value=row[c])
                if styles:
                    cell._style = copy(styles[c])

    buf = io.BytesIO()
    wb.save(buf)
    return excel_native.resave_native(buf.getvalue())


if __name__ == "__main__":
    targets = sys.argv[1:] or list(BASES)
    for t in targets:
        data = make_feeder(t)
        out = OUT_DIR / f"TDT_LVA_{t}.xlsx"
        out.write_bytes(data)
        novo = "" if t in EXISTING else "  [NOVO]"
        print(f"{t}: {out.name} ({len(data)} bytes) bloco={BASES[t]}+{novo}")
