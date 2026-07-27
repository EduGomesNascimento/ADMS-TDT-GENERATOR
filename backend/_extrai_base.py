"""Extrai da base completa a convencao de Device Mapping de TODAS as SEs."""
import collections, json, re
from pathlib import Path
import openpyxl

BASE = Path("C:/Users/egnpo/Downloads/Export_base_Full__27_fev_2026 - Copia.xlsx")
OUT = Path(__file__).parent / "data" / "convencao_dm.json"
HR = 4

sig_suf = collections.defaultdict(collections.Counter)   # sigla -> sufixo
quota = collections.Counter()                            # (tipo, SignalType)->max

# Configuracao de SAIDA por sigla (como as outras SEs comandam cada funcao).
# Serve p/ preencher Output Data Type / Control Codes quando a linha-molde da
# sigla e de status — sem isso o ADMS recusa: "Output Data Type cannot be
# left empty".
CMD_COLS = ["Output Data Type", "Control Codes", "Commanding Mode",
            "Command Times [s]", "Direction", "Command Timeout [s]",
            "Scan After Command"]
cmd_cfg = collections.defaultdict(collections.Counter)   # sigla -> config

def tipo_disp(dm):
    s = dm.split("_", 3)[-1] if dm.count("_") > 2 else ""
    return "PROT" if s.startswith(("PROT", "P_PROT", "A_PROT")) else (s or "?")

wb = openpyxl.load_workbook(BASE, read_only=True, data_only=True)
por_disp = collections.Counter()
for sn in wb.sheetnames:
    if "Signals" not in sn and "DiscreteAnalog" not in sn:
        continue
    ws = wb[sn]
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == HR:
            hdr = [str(c or "").strip() for c in row]
            ix = {n: j for j, n in enumerate(hdr) if n}
            cN = ix.get("Signal Name"); cD = ix.get("Device Mapping")
            cT = ix.get("Signal Type")
            cO = ix.get("Output Coordinates")
            cCmd = [ix.get(c) for c in CMD_COLS]
            if cN is None or cD is None:
                break
            continue
        if i <= HR or not row or not row[cN]:
            continue
        nome = str(row[cN]).strip()
        dm = str(row[cD] or "").strip()
        # A configuracao de saida so vale entre protocolos IGUAIS: as abas
        # IEC104 usam IECDoubleBit e nao tem Control Codes, o que nao existe
        # em DNP3. Sem este filtro, SECC/DJF1/86 vinham com tipo de IEC.
        if sn.startswith("DNP3_") and cO is not None and row[cO] not in (None, ""):
            _s = nome.split("_")[-1]
            _cfg = tuple(row[j] if j is not None and j < len(row) else None
                         for j in cCmd)
            if _cfg[0] and _cfg[1]:       # exige Output Data Type e Control Codes
                cmd_cfg[_s][_cfg] += 1
        st = str(row[cT] or "").strip() if cT is not None else ""
        pn, pd = nome.split("_"), dm.split("_")
        if not dm or len(pn) < 4 or len(pd) < 3 or pn[0] != pd[0]:
            continue
        sigla = "_".join(pn[3:])
        pref = f"{pd[0]}_{pd[1]}_{pd[2]}_"
        if dm.startswith(pref):
            sig_suf[sigla][dm[len(pref):]] += 1
        por_disp[(dm, st)] += 1
    print(f"  {sn}: ok")
wb.close()

for (dm, st), v in por_disp.items():
    k = f"{tipo_disp(dm)}|{st}"
    quota[k] = max(quota[k], v)

OUT.parent.mkdir(exist_ok=True)
OUT.write_text(json.dumps({
    "sigla_sufixo": {k: c.most_common(1)[0][0] for k, c in sig_suf.items()},
    "quota": dict(quota),
    "cmd_cols": CMD_COLS,
    "cmd_cfg": {k: list(c.most_common(1)[0][0]) for k, c in cmd_cfg.items()},
}, ensure_ascii=False, indent=1), encoding="utf-8")
print(f"\n{OUT.name}: {len(sig_suf)} siglas, {len(quota)} pares (tipo, Signal Type), "
      f"{len(cmd_cfg)} siglas com configuracao de comando")
