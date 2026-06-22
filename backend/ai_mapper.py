"""
ai_mapper.py — Mapeador inteligente de listas de pontos não-padrão.

Fluxo:
  1. parse_raw_excel()   — extrai sinais de qualquer Excel de UTR
  2. map_signals()       — heurística (tokens) + LLM para mapear a SIGLA ADMS
  3. to_lista_resumida() — converte para o formato de parse_points_list (para gerar TDT)
"""
from __future__ import annotations

import io
import json
import logging
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

log = logging.getLogger(__name__)

# ─── Modelos de dados ────────────────────────────────────────────────────────

@dataclass
class RawSignal:
    utr_id: str
    description: str
    dnp3_addr: Optional[int]
    signal_type: str       # analog | discrete | command | unknown
    module: str
    source_sheet: str

@dataclass
class MappedSignal:
    utr_id: str
    description: str
    dnp3_addr: Optional[int]
    signal_type: str
    module: str
    # resultado do mapeamento
    sigla: Optional[str] = None
    sigla_desc: Optional[str] = None
    confidence: int = 0
    confidence_label: str = "SEM"   # ALTA | MÉDIA | BAIXA | SEM
    alternative: Optional[str] = None
    match_method: str = "none"
    candidates: list = field(default_factory=list)   # top-N p/ o usuário escolher
    source_sheet: str = ""

# ─── Parsing de Excel ────────────────────────────────────────────────────────

_MODULE_RE   = re.compile(r'mód|módulo|modulo|module|bay|vão', re.I)
_UTR_ID_RE   = re.compile(r'identif|ponto.?na.?utr|utr.?id|ident\.', re.I)
_DNP3_RE     = re.compile(r'dnp3?\.?0?$|endereço.?dnp|index.?dnp', re.I)
_DESC_RE     = re.compile(r'descriç|descric|description|nome|designaç', re.I)
_SKIP_SHEETS = re.compile(r'^(capa|calculados?|slot|saca|sumário|config|setup|cron)', re.I)
_PREFER_SHEETS = re.compile(r'^(digitais|analogicos|analógicos|discretos)', re.I)
_ANALOG_SHEETS = re.compile(r'^(analóg|analog)', re.I)

_UTR_ID_PATTERN = re.compile(r'^[A-Z][A-Z0-9]{1,10}_[A-Z0-9_]{2,40}$')


def _is_utr_id(val) -> bool:
    if not val or not isinstance(val, str):
        return False
    return bool(_UTR_ID_PATTERN.match(str(val).strip()))


def _map_columns(header_vals: list) -> dict:
    cols: dict = {}
    for idx, cell in enumerate(header_vals):
        if cell is None:
            continue
        c = str(cell).strip()
        if not c:
            continue
        if _MODULE_RE.search(c) and 'module' not in cols:
            cols['module'] = idx
        if _UTR_ID_RE.search(c) and 'utr_id' not in cols:
            cols['utr_id'] = idx
        if _DNP3_RE.search(c) and 'dnp3' not in cols:
            cols['dnp3'] = idx
        if _DESC_RE.search(c) and 'desc' not in cols:
            cols['desc'] = idx
    return cols


def _parse_dnp3(raw) -> Optional[int]:
    if raw is None:
        return None
    try:
        v = int(float(str(raw)))
        return v if 1 <= v <= 99_999 else None
    except (ValueError, TypeError):
        return None


def _infer_type(desc: str, sheet_name: str) -> str:
    dl = desc.lower()
    if any(k in dl for k in ['corrente', 'tensão', 'tensao', 'potência', 'potencia', 'temp', 'tap atual', 'medida tap', 'frequência', 'frequencia']):
        return 'analog'
    if any(k in dl for k in ['abrir', 'fechar', 'comando']):
        return 'command'
    if _ANALOG_SHEETS.search(sheet_name):
        return 'analog'
    return 'discrete'


def _scan_grid(ws, max_rows: int = 3000, max_cols: int = 30):
    """Lê a grade da sheet (read_only-friendly) numa matriz de valores."""
    grid = []
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r >= max_rows:
            break
        grid.append(list(row[:max_cols]))
    return grid


def _find_utr_column(grid) -> Optional[tuple[int, int]]:
    """Acha a coluna com mais valores no padrão UTR_ID. Dirige a extração
    independentemente do cabeçalho. Retorna (col_idx, primeira_linha_de_dados)."""
    counts: dict[int, int] = {}
    first: dict[int, int] = {}
    for r, row in enumerate(grid):
        for c, cell in enumerate(row):
            if _is_utr_id(cell):
                counts[c] = counts.get(c, 0) + 1
                first.setdefault(c, r)
    if not counts:
        return None
    col = max(counts, key=counts.get)
    if counts[col] < 2:
        return None
    return col, first[col]


def _find_header_above(grid, data_start: int, n_cols: int) -> dict:
    """Procura a linha de cabeçalho logo acima do 1º dado (mais rótulos de texto)
    e mapeia as colunas module/desc/dnp3."""
    best_idx, best_score = None, 0
    for r in range(max(0, data_start - 1), -1, -1):
        if data_start - r > 6:
            break
        labels = grid[r]
        score = sum(1 for v in labels if isinstance(v, str) and len(v.strip()) > 2)
        if score > best_score:
            best_score, best_idx = score, r
    if best_idx is None:
        return {}
    return _map_columns(grid[best_idx][:n_cols])


_TIPO_MAP = {'A': 'analog', 'D': 'discrete', 'C': 'command',
             'A/D': 'analog', 'AD': 'analog', 'D/A': 'analog'}


def _extract_structured(grid, sheet_name: str) -> list[RawSignal]:
    """Formato por COLUNAS (Módulo / Tipo A,D,C / Descrição / Index) — listas que
    descrevem o ponto por texto, sem coluna de UTR-ID (ex.: padrão de modelagem)."""
    hdr = None
    cols: dict = {}
    for r, row in enumerate(grid[:8]):
        cells = [_norm(c) for c in row]
        txt = ' '.join(cells)
        if 'MODULO' in txt and 'TIPO' in txt and 'DESCRI' in txt:
            hdr = r
            for i, c in enumerate(cells):
                if c == 'MODULO' and 'module' not in cols:
                    cols['module'] = i
                elif c == 'TIPO' and 'tipo' not in cols:
                    cols['tipo'] = i
                elif 'DESCRI' in c and 'desc' not in cols:
                    cols['desc'] = i
                elif re.search(r'\bINDEX\b', c) and 'dnp3' not in cols:
                    cols['dnp3'] = i
            break
    if hdr is None or 'desc' not in cols or 'tipo' not in cols:
        return []

    out: list[RawSignal] = []
    for r in range(hdr + 1, len(grid)):
        row = grid[r]
        def cell(key):
            i = cols.get(key)
            return row[i] if (i is not None and i < len(row)) else None
        desc = cell('desc')
        if not desc or not str(desc).strip():
            continue
        tipo = _norm(cell('tipo'))
        sig_type = _TIPO_MAP.get(tipo, 'discrete')
        module = str(cell('module') or '').strip()
        dnp3 = _parse_dnp3(cell('dnp3'))
        out.append(RawSignal(utr_id=f"{module}_{module}" if module else "",
                             description=str(desc).strip(), dnp3_addr=dnp3,
                             signal_type=sig_type, module=module, source_sheet=sheet_name))
    return out


def _extract_sheet(ws, sheet_name: str, default_type: Optional[str] = None) -> list[RawSignal]:
    grid = _scan_grid(ws)
    if not grid:
        return []
    n_cols = max((len(r) for r in grid), default=0)

    found = _find_utr_column(grid)
    if not found:
        # sem coluna de UTR-ID → tenta formato por colunas (Módulo/Tipo/Descrição)
        return _extract_structured(grid, sheet_name)
    utr_col, data_start = found

    cols = _find_header_above(grid, data_start, n_cols)
    cols['utr_id'] = utr_col

    signals: list[RawSignal] = []
    for r in range(data_start, len(grid)):
        row = grid[r]
        raw_id = row[utr_col] if utr_col < len(row) else None
        utr_id = str(raw_id).strip() if raw_id else ''
        if not utr_id or not _is_utr_id(utr_id):
            continue

        def cell(key):
            i = cols.get(key)
            if i is not None and i < len(row) and row[i] not in (None, ''):
                return row[i]
            return None

        description = str(cell('desc')).strip() if cell('desc') else ''
        module = str(cell('module')).strip() if cell('module') else ''
        dnp3 = _parse_dnp3(cell('dnp3'))

        sig_type = default_type or _infer_type(description, sheet_name)
        signals.append(RawSignal(utr_id=utr_id, description=description,
                                 dnp3_addr=dnp3, signal_type=sig_type,
                                 module=module, source_sheet=sheet_name))
    return signals


def _detect_alias(wb) -> str:
    """Tenta detectar o alias da subestação na aba CAPA."""
    for name in wb.sheetnames:
        if name.upper() == 'CAPA':
            ws = wb[name]
            for row in ws.iter_rows(max_row=30, values_only=True):
                for cell in row:
                    if cell and re.search(r'SIGLA\s*:\s*([A-Z]{2,6})', str(cell), re.I):
                        m = re.search(r'SIGLA\s*:\s*([A-Z]{2,6})', str(cell), re.I)
                        if m:
                            return m.group(1).upper()
    return ''


def parse_raw_excel(data: bytes) -> tuple[list[RawSignal], str]:
    """
    Parseia qualquer Excel de lista de pontos.
    Retorna (sinais, alias_detectado).
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    alias = _detect_alias(wb)

    preferred = [n for n in wb.sheetnames if _PREFER_SHEETS.search(n)]
    others    = [n for n in wb.sheetnames if n not in preferred and not _SKIP_SHEETS.search(n)]

    signals: list[RawSignal] = []
    seen: set[str] = set()

    def process(sheet_names):
        for sn in sheet_names:
            # só força o tipo se a aba for explicitamente analógica/discreta;
            # senão deixa _infer_type decidir por linha (sheets mistas).
            if _ANALOG_SHEETS.search(sn):
                default = 'analog'
            elif re.search(r'discret|digita', sn, re.I):
                default = 'discrete'
            else:
                default = None
            for sig in _extract_sheet(wb[sn], sn, default):
                if sig.utr_id not in seen:
                    seen.add(sig.utr_id)
                    signals.append(sig)

    process(preferred if preferred else others)
    if not signals:
        process(others if preferred else [])

    wb.close()
    log.info("parse_raw_excel: %d sinais de %d abas; alias='%s'",
             len(signals), len(preferred or others), alias)
    return signals, alias

# ─── Carrega sigla_index ─────────────────────────────────────────────────────

_SHEET_TYPE = {
    'DNP3_DiscreteSignals': 'discrete',
    'DNP3_AnalogSignals': 'analog',
    'DNP3_DiscreteAnalog': 'analog',
    'IEC104_DiscreteSignals': 'discrete',
    'IEC104_AnalogSignals': 'analog',
}


def _load_sigla_flat(protocol: str = 'dnp3') -> dict:
    """
    Retorna {SIGLA: {desc, type, sheet_name}} — dicionário plano de todos os SIGLAs.
    """
    fname = 'sigla_index.json' if protocol == 'dnp3' else 'sigla_index_iec104.json'
    path = Path(__file__).parent / 'data' / fname
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)

    flat: dict = {}
    for sheet_name, entries in raw.items():
        sig_type = _SHEET_TYPE.get(sheet_name, 'discrete')
        for sigla, row in entries.items():
            desc = row[2] if len(row) > 2 else sigla
            flat[sigla] = {'desc': str(desc or sigla), 'type': sig_type,
                           'sheet_name': sheet_name}
    return flat

# ─── Mapeamento heurístico ───────────────────────────────────────────────────

_ACCENTS = [('Ã','A'),('Â','A'),('Á','A'),('À','A'),('É','E'),('Ê','E'),
            ('Í','I'),('Ó','O'),('Ô','O'),('Õ','O'),('Ú','U'),('Ç','C')]


# Abreviações comuns em listas de campo → forma plena (alinha lista x base ADMS).
# Aplicadas por palavra inteira (word-boundary) para não estragar palavras maiores.
_ABBREV = {
    'DISJ': 'DISJUNTOR', 'SEC': 'SECCIONADORA', 'SECC': 'SECCIONADORA',
    'TEMP': 'TEMPERATURA', 'POT': 'POTENCIA', 'CORR': 'CORRENTE',
    'TENS': 'TENSAO', 'DESL': 'DESLIGAMENTO', 'DESLIG': 'DESLIGAMENTO',
    'ALM': 'ALARME', 'ALARM': 'ALARME', 'RELIG': 'RELIGAMENTO',
    'SUBTENS': 'SUBTENSAO', 'SOBRETENS': 'SOBRETENSAO', 'SOBRECORR': 'SOBRECORRENTE',
    'FREQ': 'FREQUENCIA', 'ENROL': 'ENROLAMENTO', 'PRES': 'PRESSAO',
    'SINAL': 'SINALIZACAO', 'COMUT': 'COMUTADOR', 'PROT': 'PROTECAO',
}
_ABBREV_RE = re.compile(r'\b(' + '|'.join(sorted(_ABBREV, key=len, reverse=True)) + r')\b')


def _norm(s: str) -> str:
    s = re.sub(r'\s+', ' ', str(s or '').upper().strip())
    for old, new in _ACCENTS:
        s = s.replace(old, new)
    s = _ABBREV_RE.sub(lambda m: _ABBREV[m.group(1)], s)
    return s


# Dicionário semântico das medições analógicas mais comuns (descrição → SIGLA).
# Ordenado do mais específico ao mais genérico (o 1º regex que casar vence).
_SEMANTIC_RULES: list[tuple[str, str]] = [
    (r'TENSAO.*\bAB\b|\bVAB\b',            'VAB'),
    (r'TENSAO.*\bBC\b|\bVBC\b',            'VBC'),
    (r'TENSAO.*\bCA\b|\bVCA\b',            'VCA'),
    (r'TENSAO.*FASE\s*A\b|TENSAO\s*A\b|\bVA\b',  'VA'),
    (r'TENSAO.*FASE\s*B\b|TENSAO\s*B\b|\bVB\b',  'VB'),
    (r'TENSAO.*FASE\s*C\b|TENSAO\s*C\b|\bVC\b',  'VC'),
    (r'CORRENTE.*NEUTRO|CORRENTE.*RESIDUAL|\bIN\b', 'IN'),
    (r'CORRENTE.*FASE\s*A\b|CORRENTE\s*A\b|\bIA\b', 'IA'),
    (r'CORRENTE.*FASE\s*B\b|CORRENTE\s*B\b|\bIB\b', 'IB'),
    (r'CORRENTE.*FASE\s*C\b|CORRENTE\s*C\b|\bIC\b', 'IC'),
    (r'POT.*REATIVA',                      'Q'),   # antes de ATIVA (REATIVA contém ATIVA)
    (r'POT.*ATIVA',                        'P'),
    (r'POT.*APARENTE',                     'S'),
    (r'TEMPERATURA.*OLEO|OLEO.*TEMPERATURA',  'TOLE'),
    (r'TEMPERATURA.*ENROLAMENTO',          'TENR'),
    (r'TEMPERATURA.*AMBIENTE',             'TPAM'),
    (r'FATOR.*POTENCIA|\bFP\b',            'FP'),
    (r'ANGULO',                            'ANG'),
    (r'\bUMIDADE\b',                       'UMID'),
    (r'CORRENTE.*FUGA|\bFUGA\b',           'FUGA'),
    (r'FREQUENCIA',                        'F'),
    (r'TAP|COMUTADOR',                     'TAP'),
]
_SEMANTIC_COMPILED = [(re.compile(p), s) for p, s in _SEMANTIC_RULES]


def _semantic_match(description: str, sig_type: str, sigla_flat: dict) -> Optional[tuple[str, int]]:
    """Mapeia medições analógicas comuns por semântica da descrição."""
    if sig_type != 'analog' or not description:
        return None
    d = _norm(description)
    for rx, sigla in _SEMANTIC_COMPILED:
        if rx.search(d) and sigla in sigla_flat and sigla_flat[sigla]['type'] == 'analog':
            return sigla, 92
    return None


# Regras semânticas de PROTEÇÃO (descrição → SIGLA de função, discretos).
# Family-match: a função é certa mas o vão pode variar → confiança MÉDIA (82).
_PROT_RULES: list[tuple[str, str]] = [
    (r'SUBTENSAO',                       '27_T'),
    (r'SOBRETENSAO',                     '59'),
    (r'RELIGAMENTO',                     '79'),
    (r'\bBLOQUEIO\b|LOCKOUT',            '86'),
    (r'SEQUENCIA NEGATIVA',              '46'),
    (r'\b86\s*BF\b|FALHA.*DISJUNTOR|BREAKER FAILURE', '86BF'),
    (r'DIFERENCIAL',                     '87'),
    (r'\bBUCHHOLZ\b',                    '63TD'),
    (r'\b20\b.*VALVULA|VALVULA.*ALIVIO|ALIVIO.*PRESSAO', '20A'),
    (r'\b46\b.*SEQUENCIA',               '46'),
    (r'BAIXA PRESSAO SF6|SF6.*ALARME',   'SF6A'),
    (r'\bMOLA\b.*(CARREG|DESCARR)|MOLA DO DISJUNTOR', 'MOLA'),
    (r'LOCAL.*REMOTO|REMOTO.*LOCAL|\b43LR\b', '43LR'),
    (r'TELECOMANDO|\b43TC\b',            '43TC'),
    (r'FALHA.*COMUNICACAO|COMUNICACAO.*FALHA|\bFCOM\b', 'FCOM'),
    (r'\b86\b.*ATUAD|\b86\b.*BLOQUEIO|\b86FD\b', '86'),
    (r'\b50BF\b|\b62BF\b|FALHA.*DISJUNTOR', '50BF'),
    (r'\bMOLA\b', 'MOLA'),
    (r'\bSF6\b', 'SF6A'),
]
_PROT_COMPILED = [(re.compile(p), s) for p, s in _PROT_RULES]


def _prot_match(description: str, sig_type: str, sigla_flat: dict) -> Optional[tuple[str, int]]:
    """Mapeia funções de proteção/controle comuns por palavra-chave da descrição.
    Só discretos/comandos; confiança MÉDIA (família certa, vão pode variar)."""
    st = 'discrete' if sig_type == 'command' else sig_type
    if st != 'discrete' or not description:
        return None
    d = _norm(description)
    for rx, sigla in _PROT_COMPILED:
        if rx.search(d) and sigla in sigla_flat and sigla_flat[sigla]['type'] != 'analog':
            return sigla, 82
    return None


# Base oficial ADMS (Pontos Padrão v1): descrição normalizada → SIGLA. Quando a
# lista usa a descrição padronizada, é um match exato e confiável (ALTA).
_OFFICIAL: Optional[dict] = None


def _load_official() -> dict:
    global _OFFICIAL
    if _OFFICIAL is not None:
        return _OFFICIAL
    _OFFICIAL = {'discrete': {}, 'analog': {}}
    path = Path(__file__).parent / 'data' / 'padrao_adms.json'
    try:
        raw = json.loads(path.read_text(encoding='utf-8'))
        for kind in ('discrete', 'analog'):
            for sig, desc in raw.get(kind, {}).items():
                d = _norm(desc)
                if d and d != _norm(sig):
                    _OFFICIAL[kind].setdefault(d, sig)
    except Exception:
        pass
    return _OFFICIAL


def _official_match(description: str, sig_type: str, sigla_flat: dict) -> Optional[tuple[str, int]]:
    """Match exato contra as descrições da base oficial ADMS (ALTA)."""
    if not description:
        return None
    off = _load_official()
    st = 'discrete' if sig_type == 'command' else sig_type
    d = _norm(description)
    for kind in (('analog',) if st == 'analog' else ('discrete', 'analog')):
        sig = off[kind].get(d)
        if sig and sig in sigla_flat:
            return sig, 94
    return None


def _token_match(utr_id: str, sigla_flat: dict, sig_type: Optional[str] = None) -> Optional[tuple[str, int]]:
    """
    Encontra um SIGLA como sub-sequência de tokens do UTR ID, exigindo que o tipo
    do SIGLA case com o tipo do sinal (evita falsos positivos: TENSAO_FA→FA discreto).
    Prefere a janela mais LONGA (mais específica). Retorna (sigla, confidence) ou None.
    """
    tokens = utr_id.upper().split('_')
    n = len(tokens)
    best = None  # (window_len, sigla)
    for start in range(1, n):          # pula o 1º token (alias/módulo)
        for window in range(1, min(5, n - start + 1)):
            candidate = '_'.join(tokens[start:start + window])
            if len(candidate) <= 1:        # rejeita SIGLA de 1 char (A/B/D/F) → falso positivo
                continue
            info = sigla_flat.get(candidate)
            if not info:
                continue
            # tipo deve ser compatível (command conta como discrete)
            st = 'discrete' if sig_type == 'command' else sig_type
            if st and info['type'] != st:
                continue
            if best is None or window > best[0]:
                best = (window, candidate)
    if best:
        return best[1], 100
    return None


def _desc_match(description: str, sigla_flat: dict, sig_type: Optional[str] = None) -> Optional[tuple[str, int]]:
    """Match por descrição exata (normalizada), respeitando o tipo do sinal."""
    if not description:
        return None
    desc_norm = _norm(description)
    st = 'discrete' if sig_type == 'command' else sig_type
    for sigla, info in sigla_flat.items():
        if st and info['type'] != st:
            continue
        if _norm(info['desc']) == desc_norm:
            return sigla, 96
    return None


# ─── Matcher determinístico por código ANSI + similaridade ───────────────────
# Códigos ANSI/IEEE de função (50=sobrecorrente, 51=temporizada, 67=direcional,
# 87=diferencial, 27=subtensão, 59=sobretensão, 79=religamento, 86=bloqueio,
# 52=disjuntor, 89=seccionadora, etc.) — embutidos nas descrições das SIGLAs.
_ANSI_RE = re.compile(r'(?<![0-9])([2-9][0-9])(?![0-9])')


def _ansi_codes(text: str) -> set:
    return set(_ANSI_RE.findall(_norm(text)))


def _deterministic_index(sigla_flat: dict) -> dict:
    """Pré-computa, por SIGLA, seus tokens e códigos ANSI (descrição + código)."""
    out = {}
    for sigla, info in sigla_flat.items():
        toks = _desc_tokens(info['desc']) | _desc_tokens(sigla)
        codes = _ansi_codes(info['desc']) | _ansi_codes(sigla)
        out[sigla] = (toks, codes, info['type'])
    return out


def _fuzzy_match(sig, sigla_flat: dict, inv: dict, det: dict,
                 sig_type: Optional[str], min_conf: int = 48) -> Optional[tuple[str, int]]:
    """Match determinístico (reproduzível) por sobreposição de tokens da descrição
    + concordância de código ANSI. Confiança limitada a 78 (MÉDIA/BAIXA): nunca
    marca ALTA — esses ficam para token/semântico/descrição-exata."""
    st = 'discrete' if sig_type == 'command' else sig_type
    raw_tokens = _desc_tokens(sig.description) | _desc_tokens(sig.utr_id)
    raw_codes = _ansi_codes(sig.description) | _ansi_codes(sig.utr_id)
    if not raw_tokens:
        return None
    # candidatos: SIGLAs que compartilham ao menos 1 token
    cand: dict = {}
    for t in raw_tokens:
        for sg in inv.get(t, ()):
            if st and sigla_flat[sg]['type'] != st:
                continue
            cand[sg] = cand.get(sg, 0) + 1
    if not cand:
        return None
    best, best_conf = None, 0
    for sg, overlap in cand.items():
        if len(sg) <= 1:               # rejeita SIGLA de 1 char (falso positivo)
            continue
        sg_tokens, sg_codes, _ = det[sg]
        union = len(raw_tokens | sg_tokens) or 1
        jac = overlap / union                      # Jaccard de tokens
        code_match = len(raw_codes & sg_codes)
        conf = jac * 58 + code_match * 20
        if raw_codes and code_match == 0:
            conf *= 0.45                            # raw tem código mas SIGLA não → penaliza
        conf = int(min(78, conf))
        if conf > best_conf or (conf == best_conf and best and len(sg) < len(best)):
            best_conf, best = conf, sg
    if best and best_conf >= min_conf:
        return best, best_conf
    return None


def _rank_candidates(sig, sigla_flat: dict, inv: dict, det: dict, top: int = 6) -> list[dict]:
    """Lista os top-N SIGLAs candidatos para o sinal (score determinístico).
    Serve para o usuário ESCOLHER na revisão, mesmo quando não houve match firme."""
    st = 'discrete' if sig.signal_type == 'command' else sig.signal_type
    raw_tokens = _desc_tokens(sig.description) | _desc_tokens(sig.utr_id)
    raw_codes = _ansi_codes(sig.description) | _ansi_codes(sig.utr_id)
    if not raw_tokens:
        return []
    cand: dict = {}
    for t in raw_tokens:
        for sg in inv.get(t, ()):
            if len(sg) <= 1:
                continue
            if st and sigla_flat[sg]['type'] != st:
                continue
            cand[sg] = cand.get(sg, 0) + 1
    ranked = []
    for sg, overlap in cand.items():
        sg_tokens, sg_codes, _ = det[sg]
        union = len(raw_tokens | sg_tokens) or 1
        jac = overlap / union
        cm = len(raw_codes & sg_codes)
        score = jac * 58 + cm * 22
        if raw_codes and cm == 0:
            score *= 0.5
        ranked.append({'sigla': sg, 'score': int(min(96, score)),
                       'desc': sigla_flat[sg]['desc']})
    ranked.sort(key=lambda x: -x['score'])
    return ranked[:top]

# ─── Backend LLM ─────────────────────────────────────────────────────────────

_BATCH_SIZE = 25
_CAND_CAP = 160          # nº máx. de SIGLAs candidatos enviados por lote
_STOP = {'do','da','de','no','na','o','a','e','-','(',')','at','bt','mt','fase'}


def _desc_tokens(text: str) -> set:
    return {t for t in re.split(r'[^A-Z0-9]+', _norm(text)) if len(t) >= 2 and t not in _STOP}


def _build_inverted(sigla_flat: dict) -> dict:
    """token (da descrição/código do SIGLA) -> set de SIGLAs."""
    inv: dict = {}
    for sigla, info in sigla_flat.items():
        toks = _desc_tokens(info['desc']) | _desc_tokens(sigla)
        for t in toks:
            inv.setdefault(t, set()).add(sigla)
    return inv


def _candidates(batch, sigla_flat: dict, inv: dict, sig_type: Optional[str]) -> list[str]:
    """Recupera os SIGLAs mais relevantes para o lote (por sobreposição de tokens),
    respeitando o tipo. Mantém o prompt pequeno (cabe em tiers grátis)."""
    st = 'discrete' if sig_type == 'command' else sig_type
    score: dict = {}
    for s in batch:
        for t in (_desc_tokens(s.description) | _desc_tokens(s.utr_id)):
            for sig in inv.get(t, ()):
                if st and sigla_flat[sig]['type'] != st:
                    continue
                score[sig] = score.get(sig, 0) + 1
    ranked = sorted(score, key=score.get, reverse=True)[:_CAND_CAP]
    return ranked

_PROMPT = """\
Você é especialista em sistemas SCADA de subestações elétricas brasileiras.

Mapeie cada sinal da UTR ao código SIGLA ADMS correto do dicionário abaixo.

=== DICIONÁRIO SIGLA ADMS ===
{sigla_list}

=== SINAIS PARA MAPEAR (idx | utr_id | descrição | tipo) ===
{signal_list}

REGRAS:
- SIGLAs são códigos curtos: "IA", "50F1", "43TC", "2649", "TAP" etc.
- Combine pelo código de função embutido no utr_id OU pela descrição semântica
- Analógicos: IA=Corrente FA, IB=FB, IC=FC, P=Pot.Ativa, Q=Reativa, VA/VB/VC=Tensão, TAP=comutador
- Se não encontrar correspondência use null
- confidence: 100=certeza, 90=muito provável, 70=provável, 50=possível, 0=sem match

Retorne APENAS um array JSON (sem markdown, sem explicação):
[{{"idx":0,"sigla":"IA","confidence":100,"alt":null}}, ...]
"""


def _build_sigla_text(sigla_flat: dict, siglas: list[str]) -> str:
    """Monta o texto do dicionário só com os SIGLAs candidatos informados."""
    lines = []
    for sigla in siglas:
        info = sigla_flat.get(sigla)
        if info:
            lines.append(f"{sigla}: {info['desc']}")
    return '\n'.join(lines)


def _parse_llm_json(text: str, n: int) -> list[dict]:
    """Aceita array puro OU objeto que embrulha o array (ex.: {"mappings":[...]}),
    como pode vir do modo JSON do Ollama."""
    empty = [{'idx': i, 'sigla': None, 'confidence': 0, 'alt': None} for i in range(n)]
    # tenta o texto inteiro como JSON
    parsed = None
    try:
        parsed = json.loads(text)
    except (json.JSONDecodeError, TypeError):
        m = re.search(r'\[.*\]', text, re.DOTALL)
        if m:
            try:
                parsed = json.loads(m.group(0))
            except json.JSONDecodeError:
                parsed = None
    if isinstance(parsed, list):
        return parsed
    if isinstance(parsed, dict):
        # pega a 1ª lista de dicts dentro do objeto
        for v in parsed.values():
            if isinstance(v, list):
                return v
        # objeto único {idx,sigla,...}
        if 'sigla' in parsed or 'idx' in parsed:
            return [parsed]
    return empty


def _call_llm(prompt: str, cfg: dict) -> str:
    provider = cfg.get('provider', 'groq')
    model    = cfg.get('model', 'gemini-2.5-flash')
    api_key  = cfg.get('api_key', '')

    if provider == 'gemini':
        # SDK novo (google-genai). Aceita chaves "AQ." e modelos do free tier
        # (gemini-2.5-flash / -flash-lite / 3.1-flash-lite). NÃO usar gemini-2.0-flash
        # (cota zero em muitos projetos → 429).
        from google import genai as google_genai  # type: ignore
        client = google_genai.Client(api_key=api_key)
        resp = client.models.generate_content(model=model, contents=prompt)
        return resp.text

    if provider == 'groq':
        from groq import Groq  # type: ignore
        client = Groq(api_key=api_key)
        resp = client.chat.completions.create(
            model=model,
            messages=[{'role': 'user', 'content': prompt}],
            temperature=0.05,
        )
        return resp.choices[0].message.content

    if provider == 'ollama':
        import requests
        base = cfg.get('base_url', 'http://localhost:11434')
        resp = requests.post(f'{base}/api/chat', json={
            'model': model,
            'messages': [{'role': 'user', 'content': prompt}],
            'stream': False,
            'format': 'json',          # força saída JSON válida
            'options': {
                'temperature': 0,      # determinístico
                'num_ctx': 8192,       # contexto p/ o dicionário + lote
            },
        }, timeout=600)               # CPU é lento; lotes podem demorar
        return resp.json()['message']['content']

    raise ValueError(f"Provider desconhecido: {provider}")


def _map_llm(signals: list[RawSignal], sigla_flat: dict, llm_cfg: dict,
             stats: Optional[dict] = None) -> list[Optional[tuple[str, int]]]:
    results: list[Optional[tuple[str, int]]] = [None] * len(signals)
    inv = _build_inverted(sigla_flat)
    st = stats if stats is not None else {}
    st.setdefault('batches', 0); st.setdefault('failed_batches', 0)
    st.setdefault('skipped', 0); st.setdefault('last_error', '')
    st.setdefault('quota_hit', False)
    for batch_start in range(0, len(signals), _BATCH_SIZE):
        batch = signals[batch_start:batch_start + _BATCH_SIZE]
        types = [s.signal_type for s in batch]
        dominant = max(set(types), key=types.count) if types else None
        cands = _candidates(batch, sigla_flat, inv, dominant)
        if not cands:
            continue
        sigla_text = _build_sigla_text(sigla_flat, cands)
        signal_text = '\n'.join(
            f"{i} | {s.utr_id} | {s.description} | {s.signal_type}"
            for i, s in enumerate(batch)
        )
        prompt = _PROMPT.format(sigla_list=sigla_text, signal_list=signal_text)
        st['batches'] += 1
        # retry com backoff p/ limites de taxa dos tiers grátis (429/413/rate)
        for attempt in range(4):
            try:
                raw = _call_llm(prompt, llm_cfg)
                parsed = _parse_llm_json(raw, len(batch))
                for item in parsed:
                    idx = item.get('idx', 0)
                    sigla = item.get('sigla')
                    conf = int(item.get('confidence') or 0)
                    if sigla and sigla in sigla_flat and 0 <= idx < len(batch):
                        results[batch_start + idx] = (sigla, conf)
                break
            except Exception as e:
                msg = str(e).lower()
                transient = any(k in msg for k in ('429', '413', 'rate', 'quota', 'exhaust', 'timeout'))
                if transient and attempt < 3:
                    time.sleep(8 * (attempt + 1))   # 8s, 16s, 24s
                    continue
                # exceção: lote não processado pela IA — registra para avisar o usuário
                st['failed_batches'] += 1
                st['skipped'] += len(batch)
                st['last_error'] = str(e)[:200]
                if any(k in msg for k in ('quota', 'tokens per day', 'tpd', 'exhaust', 'resource_exhausted')):
                    st['quota_hit'] = True
                log.warning("LLM batch %d falhou: %s", batch_start, e)
                break
    return results

# ─── Mapeamento principal ─────────────────────────────────────────────────────

def _label(conf: int) -> str:
    if conf >= 90:
        return 'ALTA'
    if conf >= 70:
        return 'MÉDIA'
    if conf > 0:
        return 'BAIXA'
    return 'SEM'


def map_signals(
    raw_signals: list[RawSignal],
    protocol: str = 'dnp3',
    llm_cfg: Optional[dict] = None,
    stats: Optional[dict] = None,
) -> list[MappedSignal]:
    """
    Mapeia sinais brutos para SIGLAs ADMS.
    llm_cfg = {'provider': 'gemini'|'groq'|'ollama', 'model': ..., 'api_key': ...}
    stats   = dict opcional preenchido com exceções da IA (lotes falhos, cota, etc.)
    """
    sigla_flat = _load_sigla_flat(protocol)
    inv = _build_inverted(sigla_flat)        # token -> SIGLAs (reutilizado p/ fuzzy e LLM)
    det = _deterministic_index(sigla_flat)   # SIGLA -> (tokens, códigos ANSI, tipo)
    heuristic: list[Optional[tuple[str, int, str]]] = []
    unmatched_idx: list[int] = []

    # Camadas DETERMINÍSTICAS (reproduzíveis, sem LLM), da mais forte à mais fraca:
    # 1) token (código ADMS embutido no nome) → 2) semântico (medições) →
    # 3) descrição exata → 4) fuzzy (código ANSI + sobreposição de descrição).
    for i, sig in enumerate(raw_signals):
        st = sig.signal_type
        match = _token_match(sig.utr_id, sigla_flat, st)
        method = 'token'
        if not match:
            match = _semantic_match(sig.description, st, sigla_flat)
            method = 'semantic'
        if not match:
            match = _official_match(sig.description, st, sigla_flat)
            method = 'oficial'
        if not match:
            match = _desc_match(sig.description, sigla_flat, st)
            method = 'desc'
        if not match:
            match = _prot_match(sig.description, st, sigla_flat)
            method = 'protecao'
        if not match:
            match = _fuzzy_match(sig, sigla_flat, inv, det, st)
            method = 'fuzzy'
        if match:
            heuristic.append((match[0], match[1], method))
        else:
            heuristic.append(None)
            unmatched_idx.append(i)

    if llm_cfg and unmatched_idx:
        unmatched = [raw_signals[i] for i in unmatched_idx]
        llm_res = _map_llm(unmatched, sigla_flat, llm_cfg, stats=stats)
        for batch_i, orig_i in enumerate(unmatched_idx):
            if llm_res[batch_i]:
                sigla, conf = llm_res[batch_i]
                # CONFIABILIDADE: LLM nunca marca ALTA — ele acerta a função mas
                # chuta o vão (52-10 vs 52-2 real). Teto 85 (MÉDIA) → sempre revisar.
                heuristic[orig_i] = (sigla, min(conf, 85), 'llm')

    mapped: list[MappedSignal] = []
    for sig, match in zip(raw_signals, heuristic):
        ms = MappedSignal(
            utr_id=sig.utr_id, description=sig.description,
            dnp3_addr=sig.dnp3_addr, signal_type=sig.signal_type,
            module=sig.module, source_sheet=getattr(sig, 'source_sheet', ''),
        )
        if match:
            sigla, conf, method = match
            ms.sigla = sigla
            ms.sigla_desc = sigla_flat.get(sigla, {}).get('desc', '')
            ms.confidence = conf
            ms.confidence_label = _label(conf)
            ms.match_method = method
        # candidatos para a tela de revisão (sempre, p/ o usuário escolher)
        ms.candidates = _rank_candidates(sig, sigla_flat, inv, det)
        mapped.append(ms)

    matched = sum(1 for m in mapped if m.sigla)
    log.info("map_signals: %d/%d mapeados (%d heurística, %d LLM)",
             matched, len(mapped),
             sum(1 for _, m in zip(raw_signals, heuristic) if m and m[2] != 'llm'),
             sum(1 for _, m in zip(raw_signals, heuristic) if m and m[2] == 'llm'))
    return mapped

# ─── Converte para lista resumida (entrada do generate_tdt_from_list) ─────────

def _add_to_lista(discrete, analog, da, sigla, nome, stype, addr, aor):
    """Adiciona um sinal à lista no formato que generate_tdt_from_list consome
    (chaves inCoord/outCoord/escala — NÃO indexDnp3*, senão o endereço some)."""
    addr = str(addr) if addr not in (None, '') else ''
    if stype == 'analog':
        analog.append({'sigla': sigla, 'nome': nome, 'escala': '',
                       'inCoord': addr, 'aor': aor})
    elif stype == 'discrete_analog':
        da.append({'sigla': sigla, 'nome': nome, 'escala': '',
                   'inCoord': addr, 'aor': aor})
    elif stype == 'command':
        discrete.append({'sigla': sigla, 'nome': nome,
                         'inCoord': '', 'outCoord': addr, 'aor': aor})
    else:
        discrete.append({'sigla': sigla, 'nome': nome,
                         'inCoord': addr, 'outCoord': '', 'aor': aor})


def to_lista_resumida(mapped: list[MappedSignal], alias: str,
                      min_confidence: int = 60) -> dict:
    """Converte sinais mapeados → formato de generate_tdt_from_list (filtra por confiança).
    Marca como INCERTO (destaque na TDT) tudo que não for ALTA."""
    discrete, analog, da = [], [], []
    uncertain: set = set()
    for m in mapped:
        if not m.sigla or m.confidence < min_confidence:
            continue
        module = m.module or 'XX'
        nome = f"{alias}_{module}_{module}_{m.sigla}"
        _add_to_lista(discrete, analog, da, m.sigla, nome, m.signal_type,
                      m.dnp3_addr, f"{alias} Distr")
        if m.confidence_label != 'ALTA':
            uncertain.add(nome)
    return {'discrete': discrete, 'analog': analog,
            'discrete_analog': da, 'inputErrors': [], 'uncertain': uncertain}


def reviewed_to_lista(signals: list[dict], alias: str) -> dict:
    """Converte os sinais REVISADOS pelo usuário (cada um já com a SIGLA confirmada)
    → formato de generate_tdt_from_list. Sem filtro de confiança (já confirmado)."""
    discrete, analog, da = [], [], []
    for s in signals:
        sig = (s.get('sigla') or '').strip()
        if not sig:
            continue
        module = (s.get('module') or 'XX').strip() or 'XX'
        nome = f"{alias}_{module}_{module}_{sig}"
        _add_to_lista(discrete, analog, da, sig, nome,
                      s.get('signalType', 'discrete'), s.get('dnp3Addr'),
                      f"{alias} Distr")
    return {'discrete': discrete, 'analog': analog,
            'discrete_analog': da, 'inputErrors': []}
