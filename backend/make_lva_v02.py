"""
make_lva_v02.py — TDTs da LVA geradas da LISTA V02 (specs do chefe, autoritativas).

A V02 traz abas por módulo com SIGLA e INDEX DNP3 definidos:
  - AL21/AL22: spec completa (layout canônico de alimentador; 79LO=518 no index
    reserva do "DJ indefinido") — chefe já aplicou no ADMS, geramos p/ referência
  - AL11: spec própria completa (base 0; secc 29-8x e extras como RESERVA)
  - AL12..AL15, AL23, AL24: marcador de bloco no "26_49 Excluido"
    → AL12=100 (único bloco livre; marcador ausente), AL13=200, AL14=300,
      AL15=400, AL24=700, AL23=800 (INVERTIDO: 23=800, 24=700!)
  - BC1: spec própria completa (900-929; sem secc 29-10; módulo BC1)

Comandos são LINHAS "C" próprias no spec: out = índice da linha C da mesma SIGLA
(formato n;n). Sinais sem linha C têm os campos de comando LIMPOS (senão o
validador exige Output Coordinates). Linhas RESERVA são puladas e reportadas.

Uso: python make_lva_v02.py            (todos)
     python make_lva_v02.py AL23 BC1   (só os pedidos)
"""
from __future__ import annotations
import io, json, sys
from copy import copy
from pathlib import Path
import openpyxl
import excel_native

LISTA = Path("C:/Users/egnpo/Downloads/Lista de pontos LVA_V02 (1).xlsx")
SKEL = Path("C:/Users/egnpo/Downloads/TDT_LVA_20260720_v3.xlsx")
OUT_DIR = Path("C:/Users/egnpo/Downloads")
DATA = Path(__file__).parent / "data"
RU = "UTR_LVA_2"
AOR = "LVA Distr"
HEADER_ROWS = 4
SCALE_1000 = {"P", "Q", "S"}
AJG2_MM = "DESATIVAR@ATIVAR___DESATIVADO@ATIVADO___Custom_S_TC_SV"
FALLBACK = {"4RLR": "43LR"}          # sigla do chefe sem template exato na base

# (módulo, aba-spec, delta de índice, device)
AL21_BASE = 500
MODULES = [
    ("AL11", "AL11", 0,               "52-11"),
    ("AL12", "AL21", 100 - AL21_BASE, "52-12"),
    ("AL13", "AL21", 200 - AL21_BASE, "52-13"),
    ("AL14", "AL21", 300 - AL21_BASE, "52-14"),
    ("AL15", "AL21", 400 - AL21_BASE, "52-15"),
    ("AL21", "AL21", 0,               "52-21"),
    ("AL22", "AL22", 0,               "52-22"),
    ("AL23", "AL21", 800 - AL21_BASE, "52-23"),   # marcador do chefe: 800
    ("AL24", "AL21", 700 - AL21_BASE, "52-24"),   # marcador do chefe: 700
    ("BC1",  "BC1",  0,               "52-26"),
]


def read_spec(sn: str):
    """Lê a aba do chefe → {'A': [(idx,sigla)], 'D': [(idx,sigla)], 'C': {sigla: out}}
    + reservas (linhas RESERVA puladas)."""
    wb = openpyxl.load_workbook(LISTA, read_only=True, data_only=True)
    ws = wb[sn]
    hdr = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    off = 1 if hdr[0] == "LINHA" else 0
    iT, iS, iI = 2 + off, 5 + off, 6 + off
    spec = {"A": [], "D": [], "C": {}}
    reservas = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or len(r) <= iI or r[iI] in (None, ""):
            continue
        tipo = str(r[iT] or "").strip()
        sigla = str(r[iS] or "").strip()
        idxv = str(r[iI]).strip()
        if idxv.endswith(".0"):
            idxv = idxv[:-2]
        if sigla in ("", "RESERVA"):
            reservas.append((tipo, idxv))
            continue
        if tipo == "A":
            spec["A"].append((idxv, sigla))
        elif tipo == "C":
            spec["C"][sigla] = idxv
        elif tipo == "D":
            spec["D"].append((idxv, sigla))
    wb.close()
    return spec, reservas


def _shift(idxv: str, delta: int):
    parts = [str(int(p) + delta) for p in idxv.split(";")]
    return ";".join(parts) if len(parts) > 1 else int(parts[0])


def _subst(v, mapping):
    if not isinstance(v, str):
        return v
    for k, val in mapping.items():
        v = v.replace(k, val)
    return v


def main(targets):
    idx = json.loads((DATA / "sigla_index.json").read_text(encoding="utf-8"))
    DIG, ANL = idx["DNP3_DiscreteSignals"], idx["DNP3_AnalogSignals"]

    for mod, spec_sheet, delta, device in MODULES:
        if targets and mod not in targets:
            continue
        spec, reservas = read_spec(spec_sheet)
        wb = openpyxl.load_workbook(SKEL)
        prefix = f"LVA_{mod}_{device}"
        mapping = {"<<PREFIX>>": prefix, "<<ALIAS>>": "LVA",
                   "<<MODULE>>": mod, "<<DEVICE>>": device, "<<N>>": "1"}
        seen = {}

        for sheet, klass in (("DNP3_DiscreteSignals", "D"), ("DNP3_AnalogSignals", "A")):
            ws = wb[sheet]
            lab = {ws.cell(HEADER_ROWS, c).value: c
                   for c in range(1, ws.max_column + 1) if ws.cell(HEADER_ROWS, c).value}
            ncol = ws.max_column
            cN = lab["Signal Name"]; cIN = lab.get("Input Coordinates")
            cOUT = lab.get("Output Coordinates"); cRPC = lab.get("Remote Point Custom ID")
            cSC = lab.get("Signal Custom ID"); cESC = lab.get("Scaling Factor")
            cRU = lab.get("Remote Unit"); cRPN = lab.get("Remote Point Name")
            cAOR = lab.get("Signal AOR Group"); cMM = lab.get("Message Mapping")
            cIDT = lab.get("Input Data Type")
            cmd_cols = [lab.get(x) for x in ("Output Data Type", "Control Codes",
                                             "Command Times [s]", "Commanding Mode")]
            styles = [copy(ws.cell(HEADER_ROWS + 1, c)._style) for c in range(1, ncol + 1)]
            if ws.max_row > HEADER_ROWS:
                ws.delete_rows(HEADER_ROWS + 1, ws.max_row - HEADER_ROWS)

            tpl_src = DIG if klass == "D" else ANL
            rows = []
            for idxv, sigla in spec[klass]:
                tpl_key = sigla if sigla in tpl_src else FALLBACK.get(sigla)
                if tpl_key is None or tpl_key not in tpl_src:
                    print(f"  [{mod}] SEM TEMPLATE: {sigla} (idx {idxv}) — pulado")
                    continue
                row = [_subst(v, mapping) for v in tpl_src[tpl_key]]
                while len(row) < ncol:
                    row.append(None)
                row = row[:ncol]
                # nome único {prefix}_{SIGLA} (BC1 tem 62BF 2x → device -2)
                base_nome = f"{prefix}_{sigla}"
                n = seen.get(base_nome, 0) + 1
                seen[base_nome] = n
                nome = base_nome if n == 1 else f"LVA_{mod}_{device}-{n}_{sigla}"
                row[cN - 1] = nome
                if cRPN:
                    row[cRPN - 1] = nome
                if cRU:
                    row[cRU - 1] = RU
                if cAOR:
                    row[cAOR - 1] = AOR
                if cRPC:
                    row[cRPC - 1] = f"{nome}_{RU}"
                if cSC:
                    row[cSC - 1] = None
                # índice de entrada do spec (par mantido; single vs MultiCoord: se o
                # template é MultiCoord mas o spec dá 1 coord, degrada p/ SingleBit)
                shifted = _shift(idxv, delta)
                row[cIN - 1] = shifted
                if (klass == "D" and cIDT and str(row[cIDT - 1]) in ("MultiCoord", "DoubleBit")
                        and ";" not in str(shifted)):
                    row[cIDT - 1] = "SingleBit"
                if klass == "D":
                    out = spec["C"].get(sigla)
                    if out is not None:
                        ov = _shift(out, delta)
                        row[cOUT - 1] = f"{ov};{ov}"
                        # garante campos de comando (copia do DJF1 se template não tem)
                        if not any(row[c - 1] not in (None, "") for c in cmd_cols if c):
                            dj = DIG.get("DJF1") or []
                            for c in cmd_cols + [lab.get("Direction")]:
                                if c and c - 1 < len(dj) and dj[c - 1] not in (None, ""):
                                    row[c - 1] = dj[c - 1]
                    else:
                        # sem comando no spec → limpa campos de comando do template
                        row[cOUT - 1] = None
                        for c in cmd_cols:
                            if c:
                                row[c - 1] = None
                if klass == "A" and cESC and sigla in SCALE_1000:
                    row[cESC - 1] = 1000
                if cMM and sigla == "AJG2":
                    row[cMM - 1] = AJG2_MM
                rows.append(row)

            for i, row in enumerate(rows):
                for c in range(ncol):
                    cell = ws.cell(HEADER_ROWS + 1 + i, c + 1, value=row[c])
                    cell._style = copy(styles[c])

        buf = io.BytesIO()
        wb.save(buf)
        data = excel_native.resave_native(buf.getvalue())
        out = OUT_DIR / f"TDT_LVA_{mod}.xlsx"
        try:
            out.write_bytes(data)
        except PermissionError:
            out = OUT_DIR / f"TDT_LVA_{mod}_v02.xlsx"
            out.write_bytes(data)
        note = "  [NAO IMPORTAR: chefe ja aplicou no ADMS]" if mod in ("AL21", "AL22") else ""
        print(f"{mod}: {out.name} ({len(data)} b) dig={len(spec['D'])} anl={len(spec['A'])} "
              f"cmd={len(spec['C'])} reservas_puladas={len(reservas)}{note}")


if __name__ == "__main__":
    main(sys.argv[1:])
