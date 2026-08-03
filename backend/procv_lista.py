"""
procv_lista.py — O PROCV pedido pelo Cesar: cruza a LISTA DE PONTOS com o que
JA ESTA CARREGADO no ADMS e diz, ponto a ponto, o que falta criar.

  "gostaria que vc exportasse a TDT do que foi carregado ate agora da CAS, e
   fizesse um procv com a lista de pontos buscando o que falta criar no adms"
  "depois que criar faz isso pra gente revisar e nao deixar nada pra tras"

Ou seja: e uma ferramenta de CONFERENCIA, para rodar de novo depois de cada
rodada de criacao no modelo. A cada execucao ela re-cruza tudo e o que sobrar
e o que ainda falta.

Fontes:
  1) a lista corrigida (so as abas VISIVEIS — aba oculta nao existe)
  2) a TDT do que foi carregado
  3) o retorno do import (quais sinais o ADMS aceitou de fato)
  4) o modelo (quais dispositivos existem)

Uso: python procv_lista.py [tdt_carregada.xlsx] [retorno.csv]
Saida: C:/Users/egnpo/Downloads/CASCA_PROCV_LISTA.xlsx
"""
from __future__ import annotations
import collections, csv, io, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import excel_native
import casca_devmap as devmap

DOWN = Path("C:/Users/egnpo/Downloads")
OUT = DOWN / "CASCA_PROCV_LISTA.xlsx"
HR = 4
SKIP = {"Informações", "RELACAO RELES", "MAPA DE REDE", "Lista"}


def _atual(base: str) -> Path:
    o, n = DOWN / f"{base}.xlsx", DOWN / f"{base}_NOVA.xlsx"
    if n.exists() and (not o.exists() or n.stat().st_mtime > o.stat().st_mtime):
        return n
    return o


def ler_lista():
    """Todo ponto das abas VISIVEIS, com a aba/linha de origem."""
    wb = openpyxl.load_workbook(_atual("RGE ADMS_Lista Pontos Casca_CORRIGIDA"),
                                data_only=True)
    pts = []
    for sn in wb.sheetnames:
        if sn in SKIP or wb[sn].sheet_state != "visible":
            continue
        ws = wb[sn]
        hi = next((r for r in range(1, min(15, ws.max_row + 1))
                   if any(str(ws.cell(r, c).value or "").strip() == "SIGLA SINAL"
                          for c in range(1, ws.max_column + 1))), None)
        if hi is None:
            continue
        col = {str(ws.cell(hi, c).value or "").strip(): c
               for c in range(1, ws.max_column + 1)}
        if "NOME" not in col:
            continue
        for r in range(hi + 1, ws.max_row + 1):
            g = lambda k: ws.cell(r, col[k]).value if k in col else None
            tipo = str(g("TIPO") or "").strip()
            nome = str(g("NOME") or "").strip()
            if tipo not in ("D", "A", "C", "A/D") or not nome.startswith("CAS_"):
                continue
            pts.append({
                "aba": sn, "linha": r, "tipo": tipo, "nome": nome,
                "sigla": str(g("SIGLA SINAL") or "").strip(),
                "desc": str(g("DESCRIÇÃO DO PONTO") or "").strip(),
                "idx": g("INDEX DNP3"),
                "usado": str(g("Utilizado?") or "").strip().upper() == "SIM",
            })
    wb.close()
    return pts


def ler_tdt(p: Path):
    """Signal Name -> (Device Mapping, aba)."""
    out = {}
    if not p.exists():
        return out
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    for sn in ("DNP3_DiscreteSignals", "DNP3_AnalogSignals",
               "DNP3_DiscreteAnalog"):
        if sn not in wb.sheetnames:
            continue
        ls = list(wb[sn].iter_rows(values_only=True))
        ix = {n: i for i, n in enumerate(ls[HR - 1]) if n}
        for r in ls[HR:]:
            n = r[ix["Signal Name"]]
            if n:
                out[str(n)] = (str(r[ix["Device Mapping"]] or ""),
                               sn.replace("DNP3_", ""))
    wb.close()
    return out


def ler_retorno(p: Path):
    """Signal Name -> (entrou?, mensagem). Aceita ';'/portugues e ','/ingles."""
    if not p.exists():
        return {}
    txt = io.open(p, encoding="utf-8-sig", newline="").read()
    cab = txt.split("\n", 1)[0]
    d = ";" if cab.count(";") > cab.count(",") else ","
    out = {}
    for r in list(csv.reader(io.StringIO(txt), delimiter=d))[1:]:
        if len(r) > 3 and r[1]:
            ok = not r[0].strip().lower().startswith(("falh", "fail"))
            out[r[1]] = (ok, r[3])
    return out


def _norm(nome: str) -> str:
    """O NOME da lista vira o do ADMS (LT1 -> LTKVM, TR6 -> TR1)."""
    q = nome.split("_")
    if len(q) > 1 and q[1] in devmap.MODULO_EQUIV:
        v = devmap.MODULO_EQUIV[q[1]][0]
        if len(q) > 2 and q[2] == q[1]:
            q[2] = v
        q[1] = v
        return "_".join(q)
    return nome


def main(tdt_path: str, csv_path: str):
    pts = ler_lista()
    tdt = ler_tdt(Path(tdt_path))
    ret = ler_retorno(Path(csv_path))
    ids = devmap.modelo_new()[0]

    linhas = []
    for p in pts:
        nome_adms = _norm(p["nome"])
        dm, aba_tdt = tdt.get(nome_adms, ("", ""))
        entrou, msg = ret.get(nome_adms, (None, ""))

        # QUEM ESTA NA UTR: o sinal cujo mapeamento FALHOU. A TDT nunca
        # escreve "UTR_CAS_3" no Device Mapping — ela escreve o dispositivo
        # que DEVERIA existir; o ADMS e que, nao achando, pendura o sinal
        # direto na UTR_CAS_3. Entao "na UTR" == "falhou no import", e e o
        # que a arvore mostra em verde sob a UTR.
        if not p["usado"]:
            sit, acao = "Utilizado? = NAO", "nada — nao entra na TDT por decisao da lista"
        elif p["tipo"] == "C":
            sit, acao = "COMANDO", "nada — vira Output do sinal de status, nao e linha propria"
        elif not dm:
            sit, acao = "FORA DA TDT", "conferir: o ponto nao foi para nenhuma TDT"
        elif entrou is True:
            sit, acao = "OK — no ADMS", ""
        elif entrou is False:
            if "Could not find" in msg:
                sit, acao = "NA UTR — falta o dispositivo", f"CRIAR no ADMS: {dm}"
            elif "Found multiple" in msg:
                sit, acao = "NA UTR — ID duplicado", f"dar ID proprio ao 2o dispositivo com {dm}"
            elif "already mapped" in msg:
                sig = "_".join(nome_adms.split("_")[3:])
                sit, acao = "NA UTR — falta o estagio", f"CRIAR o estagio da funcao: {dm}_{sig}"
            else:
                sit, acao = "FALHOU", msg[:90]
        else:
            sit = "na TDT, sem retorno" if dm in ids else "FALTA O DISPOSITIVO"
            acao = "" if dm in ids else f"CRIAR no ADMS: {dm}"

        linhas.append([p["aba"], p["linha"], p["nome"], nome_adms, p["sigla"],
                       p["tipo"], p["desc"], p["idx"], dm, sit, acao])

    # ── planilha ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E78")
    cor = {"OK — no ADMS": "C6EFCE",
           "NA UTR — falta o dispositivo": "FFC7CE",
           "NA UTR — falta o estagio": "FFEB9C",
           "NA UTR — ID duplicado": "BDD7EE",
           "Utilizado? = NAO": "F2F2F2", "COMANDO": "F2F2F2",
           "FORA DA TDT": "FFC7CE", "FALHOU": "FFC7CE"}

    ws = wb.active
    ws.title = "1-PONTO A PONTO"
    ws.append(["Aba da lista", "Linha", "NOME na lista", "NOME no ADMS",
               "SIGLA", "Tipo", "Descricao", "Index DNP3",
               "Device Mapping", "SITUACAO", "O QUE FAZER"])
    for c in ws[1]:
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for r in linhas:
        ws.append(r)
        ws.cell(ws.max_row, 10).fill = PatternFill(
            "solid", fgColor=cor.get(r[9], "FFFFFF"))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    for col, w in zip("ABCDEFGHIJK",
                      (22, 7, 30, 30, 10, 6, 40, 11, 34, 22, 46)):
        ws.column_dimensions[col].width = w

    # ── o que criar, agrupado ────────────────────────────────────────────────
    ws2 = wb.create_sheet("2-O QUE CRIAR")
    ws2.append(["Dispositivo a CRIAR no ADMS", "Pontos que dependem",
                "Vao", "Exemplos"])
    for c in ws2[1]:
        c.font = hf; c.fill = hfill
    falta = collections.defaultdict(list)
    for r in linhas:
        if r[10].startswith("CRIAR"):
            falta[r[10].split(": ", 1)[1]].append(r[3])
    for dm, ns in sorted(falta.items(), key=lambda x: -len(x[1])):
        pp = dm.split("_")
        ws2.append([dm, len(ns), pp[1] if len(pp) > 1 else "",
                    ", ".join(sorted(ns)[:4])])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:D{ws2.max_row}"
    for col, w in zip("ABCD", (40, 12, 10, 62)):
        ws2.column_dimensions[col].width = w

    # ── resumo ───────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("0-RESUMO", 0)
    c = collections.Counter(r[9] for r in linhas)
    ws3.append(["PROCV: LISTA DE PONTOS x O QUE ESTA NO ADMS"])
    ws3.append([])
    ws3.append([f"Lista (abas visiveis): {len(linhas)} pontos"])
    ws3.append([])
    ws3.append(["SITUACAO", "Pontos"])
    for k, v in c.most_common():
        ws3.append([k, v])
    ws3.append([])
    ws3.append([f"Dispositivos a criar no ADMS: {len(falta)}"])
    ws3.append([])
    ws3.append(["Rode de novo depois de criar no modelo — o que sobrar na aba"])
    ws3.append(["'2-O QUE CRIAR' e o que ainda falta."])
    ws3["A1"].font = Font(bold=True, size=13)
    for row in ws3.iter_rows():
        if row[0].value == "SITUACAO":
            for x in row:
                x.font = hf; x.fill = hfill
    for col, w in zip("AB", (46, 12)):
        ws3.column_dimensions[col].width = w

    alvo = OUT
    try:
        wb.save(alvo)
    except PermissionError:
        alvo = OUT.with_name(OUT.stem + "_NOVA.xlsx")
        wb.save(alvo)
        print(f"ATENCAO: {OUT.name} aberto no Excel — gravado como {alvo.name}")
    alvo.write_bytes(excel_native.resave_native(alvo.read_bytes()))
    print(f"{alvo.name}: {len(linhas)} pontos da lista")
    for k, v in c.most_common():
        print(f"   {v:>5}  {k}")
    print(f"\n   dispositivos a criar no ADMS: {len(falta)}")


if __name__ == "__main__":
    a = sys.argv[1:]
    main(a[0] if a else str(DOWN / "CAS_TDT_COMPLETA_FUNCIONANDO_MENOS_O_QUE_TA_NA_UTR.xlsx"),
         a[1] if len(a) > 1 else str(DOWN / "erros.csv"))
