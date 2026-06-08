"""
build_index.py — varre a base completa UMA vez e cria data/sigla_index.json:
um template (linha tokenizada) por SIGLA, para discretos e analógicos. Usado pela
importação de listas de pontos (casa cada SIGLA da lista com os campos fixos reais
da base ADMS). SIGLA = NOME sem os 3 primeiros tokens (alias_modulo_device).
"""
from __future__ import annotations
import json, re, sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent
SRC = ROOT.parent.parent
BASE = SRC / "Export_base_Full__27_fev_2026.xlsx"

PROTO = (sys.argv[1] if len(sys.argv) > 1 else "dnp3").lower()
PREFIX = "IEC104" if PROTO == "iec104" else "DNP3"
SHEETS = (f"{PREFIX}_DiscreteSignals", f"{PREFIX}_AnalogSignals")
OUT = ROOT / "data" / (f"sigla_index_{PROTO}.json" if PROTO != "dnp3" else "sigla_index.json")


def make_tok(alias, module, device):
    prefix = f"{alias}_{module}_{device}"
    pairs = [(prefix, "<<PREFIX>>"), (alias, "<<ALIAS>>"), (module, "<<MODULE>>"), (str(device), "<<DEVICE>>")]
    def tok(v):
        if not isinstance(v, str):
            return v
        out = v
        for lit, ph in pairs:
            if lit:
                out = re.sub(r"(?<![A-Za-z0-9])" + re.escape(lit) + r"(?![A-Za-z0-9])", ph, out)
        return out
    return tok


def main():
    wb = openpyxl.load_workbook(BASE, read_only=True, data_only=True)
    index = {SHEETS[0]: {}, SHEETS[1]: {}}

    for sheet in SHEETS:
        # nº de colunas = largura da linha de rótulos (linha 4)
        hdr = next(wb[sheet].iter_rows(min_row=4, max_row=4, values_only=True))
        ncol = len(hdr)
        idx = index[sheet]
        for r in wb[sheet].iter_rows(min_row=5, values_only=True):
            nm = r[0]
            if not nm:
                continue
            p = str(nm).split("_")
            if len(p) < 4:
                continue
            sigla = "_".join(p[3:])
            if sigla in idx:
                continue
            alias, module, device = p[0], p[1], p[2]
            tok = make_tok(alias, module, device)
            trow = [tok(c) for c in r]
            while len(trow) < ncol:
                trow.append(None)
            idx[sigla] = trow

    OUT.write_text(json.dumps(index, ensure_ascii=False, default=str), encoding="utf-8")
    print(f"[{PROTO}] índice: {len(index[SHEETS[0]])} siglas discretas, "
          f"{len(index[SHEETS[1]])} analógicas -> {OUT.stat().st_size//1024} KB ({OUT.name})")


if __name__ == "__main__":
    main()
