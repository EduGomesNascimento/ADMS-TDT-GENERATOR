"""
build_tr.py — adiciona o tipo "Transformador (TR)" ao catálogo, extraindo da
base completa o TR mais completo (corpo TR{n} + lado AT + lado BT, digitais e
analógicos). Modelo CONSOLIDADO: usuário informa alias + número do trafo (N).

Tokenização específica do TR: alias da subestação -> <<ALIAS>>, e "TR{n}" ->
"TR<<N>>" (cobre TR{n}, TR{n}AT, TR{n}BT). Os ids de bay (52-x, 89-x, 29-x) são
mantidos do template.
"""
from __future__ import annotations
import json, re
from pathlib import Path
import openpyxl
from build_catalog import _lbl_idx, detect_command, load_dictionary

ROOT = Path(__file__).resolve().parent
SRC = ROOT.parent.parent
BASE = SRC / "Export_base_Full__27_fev_2026.xlsx"
CATALOG = ROOT / "data" / "catalog.json"

MOD_RE = re.compile(r"^TR(\d+)(AT|BT)?$")  # TR1, TR1AT, TR1BT


def family(parts):
    """retorna (n, side) se o módulo (parts[1]) pertence a um TR; senão None."""
    if len(parts) < 4:
        return None
    m = MOD_RE.match(parts[1])
    if not m:
        return None
    return m.group(1), (m.group(2) or "CORPO")


def main():
    cat = json.loads(CATALOG.read_text(encoding="utf-8"))
    cols = cat["columns"]
    dictionary = load_dictionary()
    wb = openpyxl.load_workbook(BASE, read_only=True, data_only=True)

    # 1ª passada: escolhe (alias, n) mais completo
    score = {}
    for sheet, klass in (("DNP3_DiscreteSignals", "discrete"), ("DNP3_AnalogSignals", "analog")):
        for r in wb[sheet].iter_rows(min_row=5, values_only=True):
            nm = r[0]
            if not nm:
                continue
            p = str(nm).split("_")
            fam = family(p)
            if not fam:
                continue
            key = (p[0], fam[0])
            d = score.setdefault(key, {"discrete": 0, "analog": 0, "sides": set()})
            d[klass] += 1
            d["sides"].add(fam[1])

    def rank(kv):
        d = kv[1]
        return (len(d["sides"]), d["analog"] > 0, d["discrete"] + d["analog"])
    best = max(score.items(), key=rank)
    alias, n = best[0]
    print(f"TR escolhido: alias={alias} n={n} -> {best[1]['discrete']} dig, {best[1]['analog']} anl, lados={best[1]['sides']}")

    tr_pfx = f"{alias}_TR{n}"

    def tokenize(v):
        if not isinstance(v, str):
            return v
        return v.replace(alias, "<<ALIAS>>").replace(f"TR{n}", "TR<<N>>")

    side_label = {"CORPO": "Corpo", "AT": "Alta (AT)", "BT": "Baixa (BT)"}
    signals = {"discrete": [], "analog": []}

    # 2ª passada: extrai linhas do TR escolhido
    for sheet, klass in (("DNP3_DiscreteSignals", "discrete"), ("DNP3_AnalogSignals", "analog")):
        lab = _lbl_idx(cols[sheet])
        seen = set()
        for r in wb[sheet].iter_rows(min_row=5, values_only=True):
            nm = r[0]
            if not nm or not str(nm).startswith(tr_pfx):
                continue
            p = str(nm).split("_")
            fam = family(p)
            if not fam or fam[0] != n:
                continue
            rel = str(nm)[len(alias) + 1:]  # relativo (sem alias_)
            if rel in seen:
                continue
            seen.add(rel)
            trow = [tokenize(c) for c in r]
            while len(trow) < len(cols[sheet]):
                trow.append(None)
            code = p[-1]
            info = dictionary.get(code.upper(), {})
            real_alias = r[2] if len(r) > 2 else None
            command = detect_command(r, lab) if klass == "discrete" else None
            signals[klass].append({
                "suffix": rel,
                "code": code,
                "group": side_label.get(fam[1], fam[1]),
                "description": info.get("description") or (str(real_alias) if real_alias else code),
                "aliasLabel": str(real_alias) if real_alias else "",
                "klass": info.get("klass", klass),
                "measurementType": r[4] if len(r) > 4 else None,
                "signalSubType": r[14] if len(r) > 14 else None,
                "phases": r[15] if len(r) > 15 else None,
                "hasCommand": command is not None,
                "command": command,
                "row": trow,
            })

    dev = {
        "id": "transformador",
        "label": "Transformador (TR)",
        "description": "Transformador de força — corpo + lados Alta (AT) e Baixa (BT) consolidados.",
        "source": tr_pfx,
        "consolidated": True,
        "defaults": {"transformerNumber": "1"},
        "signalCount": {"discrete": len(signals["discrete"]), "analog": len(signals["analog"])},
        "signals": signals,
    }
    cat["deviceTypes"] = [d for d in cat["deviceTypes"] if d["id"] != "transformador"] + [dev]
    CATALOG.write_text(json.dumps(cat, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"TR adicionado: {len(signals['discrete'])} dig, {len(signals['analog'])} anl")


if __name__ == "__main__":
    main()
