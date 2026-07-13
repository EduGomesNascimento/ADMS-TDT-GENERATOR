"""
make_cva_feeders.py — Clona a TDT do CVA AL11 para os demais alimentadores.

Regra (confirmada na lista 'CVA - Pontos Por Equipamentos DNP_V03 - COS'):
os índices UTR/COS são GLOBAIS e cada alimentador ocupa um bloco fixo:
  Analógicos : +10 por alimentador   (AL11: 0-7,  AL12: 10-17, ... AL23: 60-67)
  Comandos   : +10 por alimentador   (AL11: 0-4,  AL12: 10-14, ...)
  Digitais   : +35 por alimentador   (AL11: 0-28, AL12: 35-63, ...)

O que muda por TDT (o resto é clonado do AL11):
  - número do AL nos textos (AL11→ALxx, 52-11→52-xx, 11Q0→xxQ0)
  - Input Coordinates (offset do bloco)
  - Output Coordinates (offset do bloco de comando)
  - Remote Point Custom ID (recalculado = {novo_nome}__UTR_CVA_2)
  - Signal Custom ID limpo (ADMS gera o GUID do objeto novo)

Índices ESPECIAIS fora do bloco (DR=1199, 81 in=11257/out=1953) são mantidos
e reportados — confirmar manualmente (parecem legado de outro sistema).

Uso: python make_cva_feeders.py [AL12 AL13 ...]   (default: todos)
"""
from __future__ import annotations
import io, re, sys
from copy import copy
from pathlib import Path
import openpyxl
import excel_native

SRC = Path("C:/Users/egnpo/Downloads/UTR_CVA_2_AL11 (2).xlsx")
OUT_DIR = Path("C:/Users/egnpo/Downloads")
RU = "UTR_CVA_2"
HEADER_ROWS = 4

# ordem dos alimentadores nos blocos de índice (AL11 = bloco 0)
FEEDERS = ["AL11", "AL12", "AL13", "AL14", "AL21", "AL22", "AL23"]
STEP = {"analog_in": 10, "cmd_out": 10, "disc_in": 35}
# faixas do bloco do AL11 (valores fora = especiais, não deslocar)
BLOCK = {"analog_in": (0, 9), "cmd_out": (0, 9), "disc_in": (0, 34)}

# disjuntor FÍSICO de cada alimentador (confirmado na lista COS e na base antiga
# UTR_CVA_1 — o número do disjuntor NÃO segue o número do AL; formato sem zero
# à esquerda como a base ADMS nomeia: CVA_AL13_52-9_...)
BREAKER = {"AL11": "52-11", "AL12": "52-10", "AL13": "52-9", "AL14": "52-16",
           "AL21": "52-5", "AL22": "52-4", "AL23": "52-3"}


def _off(val, kind: str, k: int, specials: list, ctx: str):
    """Desloca um valor de coordenada ('n' ou 'a;b') pelo bloco do alimentador k."""
    if val in (None, ""):
        return val
    lo, hi = BLOCK[kind]
    step = STEP[kind] * k
    parts = str(val).split(";")
    out = []
    for p in parts:
        p = p.strip()
        if not re.fullmatch(r"-?\d+", p):
            return val
        n = int(p)
        if lo <= n <= hi:
            out.append(str(n + step))
        else:
            specials.append(f"{ctx}: coord especial {n} mantida")
            out.append(p)
    joined = ";".join(out)
    return int(joined) if ";" not in joined and str(val).find(";") < 0 else joined


def make_feeder(target: str) -> tuple[bytes, list]:
    k = FEEDERS.index(target)
    n_old, n_new = "11", target[2:]          # '12', '21', ...
    wb = openpyxl.load_workbook(SRC)         # clona TUDO (estilos/validações)
    specials: list = []

    # ordem importa: disjuntor físico ANTES do AL (52-11 contém '11')
    subs = [(f"52-{n_old}", BREAKER[target]), (f"AL{n_old}", f"AL{n_new}"),
            (f"{n_old}Q0", f"{n_new}Q0")]

    for sheet in ("DNP3_DiscreteSignals", "DNP3_AnalogSignals"):
        ws = wb[sheet]
        labels = [ws.cell(HEADER_ROWS, c).value for c in range(1, ws.max_column + 1)]
        L = {l: i + 1 for i, l in enumerate(labels) if l}
        c_name = L.get("Signal Name")
        c_in = L.get("Input Coordinates")
        c_out = L.get("Output Coordinates")
        c_rpc = L.get("Remote Point Custom ID")
        c_scid = L.get("Signal Custom ID")
        kind_in = "analog_in" if "Analog" in sheet else "disc_in"

        for r in range(HEADER_ROWS + 1, ws.max_row + 1):
            if not ws.cell(r, c_name).value:
                continue
            # 1) troca do número do AL em toda célula texto da linha
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    nv = v
                    for a, b in subs:
                        nv = nv.replace(a, b)
                    if nv != v:
                        ws.cell(r, c).value = nv
            nome = str(ws.cell(r, c_name).value)
            # 2) coordenadas
            if c_in:
                ws.cell(r, c_in).value = _off(ws.cell(r, c_in).value, kind_in,
                                              k, specials, nome)
            # fix de bug herdado do AL11: P e Q ambos in=6; pela lista COS a
            # Potência Reativa (Q) é o índice 7 do bloco
            if "Analog" in sheet and nome.endswith("_Q"):
                ws.cell(r, c_in).value = 7 + STEP["analog_in"] * k
                specials.append(f"{nome}: Q corrigido p/ {7 + STEP['analog_in']*k} "
                                f"(AL11 original tem P e Q duplicados em 6)")
            if c_out:
                ws.cell(r, c_out).value = _off(ws.cell(r, c_out).value, "cmd_out",
                                               k, specials, nome)
            # 3) Remote Point Custom ID recalculado; Signal Custom ID limpo
            if c_rpc and ws.cell(r, c_rpc).value:
                ws.cell(r, c_rpc).value = f"{nome}__{RU}"
            if c_scid:
                ws.cell(r, c_scid).value = None

    buf = io.BytesIO()
    wb.save(buf)
    data = excel_native.resave_native(buf.getvalue())
    return data, specials


if __name__ == "__main__":
    targets = sys.argv[1:] or [f for f in FEEDERS if f != "AL11"]
    for t in targets:
        data, specials = make_feeder(t)
        out = OUT_DIR / f"UTR_CVA_2_{t}.xlsx"
        out.write_bytes(data)
        print(f"{t}: {out.name} ({len(data)} bytes)")
        for s in sorted(set(specials)):
            print(f"   [ESPECIAL] {s}")
