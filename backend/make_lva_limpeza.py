"""
make_lva_limpeza.py — TDT de LIMPEZA: apaga os restos das importações antigas
(pré-V02) na UTR_LVA_2, usando a coluna Delete=Yes do formato TDT
("If set to Yes, both signal and remote point is going to be deleted...").

Apaga SÓ o que mudou de nome/estrutura na V02 (o resto atualiza em cima):
  - AL11: IN (analóg.), 50F, 50N, 51F, AJG2
  - AL12..AL15, AL23, AL24: 50F, 50N, 51F, 43LR  (renomeados 50F1/50N1/51F1/4RLR)
  - BC2 inteiro (módulo virou BC1): 24 discretos + 14 analógicos
NÃO toca em AL21/AL22 (do chefe).

Importar TDT_LVA_LIMPEZA.xlsx ANTES da TDT_LVA_COMPLETA_SEM_AL21_AL22.xlsx.
Uso: python make_lva_limpeza.py
"""
from __future__ import annotations
import io
from copy import copy
from pathlib import Path
import openpyxl
import excel_native
import make_lva_feeders as OLD
import make_lva_bc as OLDBC

SKEL = Path("C:/Users/egnpo/Downloads/TDT_LVA_20260720_v3.xlsx")
OUT = Path("C:/Users/egnpo/Downloads/TDT_LVA_LIMPEZA.xlsx")
HEADER_ROWS = 4

# módulo → sufixos a apagar ('*' = todos)
KILL = {
    "AL11": {"D": {"50F", "50N", "51F", "AJG2"}, "A": {"IN"}},
    **{m: {"D": {"50F", "50N", "51F", "43LR"}, "A": set()}
       for m in ("AL12", "AL13", "AL14", "AL15", "AL23", "AL24")},
    "BC2": {"D": {"*"}, "A": {"*"}},
}


def _suffix(nome: str) -> str:
    p = str(nome).split("_")
    return "_".join(p[3:]) if len(p) > 3 else p[-1]


def collect(tdt_bytes: bytes, kill: dict):
    """Extrai de uma TDT antiga as linhas cujo sufixo está na lista de exclusão."""
    wb = openpyxl.load_workbook(io.BytesIO(tdt_bytes), read_only=True, data_only=True)
    out = {"DNP3_DiscreteSignals": [], "DNP3_AnalogSignals": []}
    for sheet, klass in (("DNP3_DiscreteSignals", "D"), ("DNP3_AnalogSignals", "A")):
        ws = wb[sheet]
        targets = kill.get(klass, set())
        for r in ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True):
            if not r or not r[0]:
                continue
            if "*" in targets or _suffix(r[0]) in targets:
                out[sheet].append(list(r))
    wb.close()
    return out


def main():
    rows = {"DNP3_DiscreteSignals": [], "DNP3_AnalogSignals": []}
    # alimentadores antigos (gerador antigo reconstrói os nomes/linhas exatos)
    for mod in ("AL11", "AL12", "AL13", "AL14", "AL15", "AL23", "AL24"):
        data = OLD.make_feeder(mod)
        got = collect(data, KILL[mod])
        for sh in rows:
            rows[sh] += got[sh]
        print(f"{mod}: +{sum(len(v) for v in got.values())} p/ apagar")
    # BC2 antigo inteiro
    OLDBC.OUT = Path("C:/Users/egnpo/AppData/Local/Temp/claude/"
                     "C--Users-egnpo-Downloads-UI-DE-TDT/"
                     "98722265-60c2-4b93-8ad9-503ac25f1e70/scratchpad/_bc2_old.xlsx")
    OLDBC.main()
    got = collect(OLDBC.OUT.read_bytes(), KILL["BC2"])
    for sh in rows:
        rows[sh] += got[sh]
    print(f"BC2: +{sum(len(v) for v in got.values())} p/ apagar")

    wb = openpyxl.load_workbook(SKEL)
    for sheet in rows:
        ws = wb[sheet]
        lab = {ws.cell(HEADER_ROWS, c).value: c for c in range(1, ws.max_column + 1)
               if ws.cell(HEADER_ROWS, c).value}
        cDel = lab["Delete"]; cDelRP = lab.get("Delete Remote Point")
        cSC = lab.get("Signal Custom ID")
        ncol = ws.max_column
        styles = [copy(ws.cell(HEADER_ROWS + 1, c)._style) for c in range(1, ncol + 1)]
        if ws.max_row > HEADER_ROWS:
            ws.delete_rows(HEADER_ROWS + 1, ws.max_row - HEADER_ROWS)
        for i, row in enumerate(rows[sheet]):
            while len(row) < ncol:
                row.append(None)
            row[cDel - 1] = "Yes"                 # apaga sinal + remote point
            if cDelRP:
                row[cDelRP - 1] = "Yes"
            if cSC:
                row[cSC - 1] = None
            for c in range(ncol):
                cell = ws.cell(HEADER_ROWS + 1 + i, c + 1, value=row[c])
                cell._style = copy(styles[c])
        print(f"{sheet}: {len(rows[sheet])} linhas Delete=Yes")

    buf = io.BytesIO()
    wb.save(buf)
    data = excel_native.resave_native(buf.getvalue())
    OUT.write_bytes(data)
    print(f"salva: {OUT.name} ({len(data)} bytes)")


if __name__ == "__main__":
    main()
