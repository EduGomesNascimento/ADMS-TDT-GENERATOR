"""
make_snd_tpj_final.py — TDT DEFINITIVA do vão TPJ (SND) usando o CHANGESET como gabarito.

O changeset PT-MOD-SE-SND.xml contém os 102 sinais que já EXISTEM no modelo
(nomes SND_LTTPJ_* consolidados) e o Device Mapping exato de cada um.
A TDT deve usar ESSES nomes — assim o import casa com os objetos do modelo em
vez de criar sinais novos (que causavam os conflitos de client point).

Índices: re-sequenciados por tipo (UTR_SND_1 é nova). Comandáveis sem valor
real → filler 9599+. AOR: SAN Trans.

Uso: python make_snd_tpj_final.py
"""
from __future__ import annotations
import json, re
import xml.etree.ElementTree as ET
from pathlib import Path
import tdt_engine as E

XML = Path("C:/Users/egnpo/Downloads/PT-MOD-SE-SND.xml")
OUT = Path("C:/Users/egnpo/Downloads/TDT_SND_TPJ_final.xlsx")
DATA = Path(__file__).parent / "data"
AOR = "SAN Trans"
FILL_START = 9599

# sigla do MODELO sem template exato na base → template da variante mais próxima
# (o NOME do sinal continua o do modelo; só a linha-molde vem da variante)
FALLBACK = {
    "21_1": "67_1", "21_2": "67_2",
    "21_21FT": "21_21", "21_21NT": "21_21N",
    "21_50F1": "21_50_1", "21_50F2": "21_50_2",
    "21_51F": "21_51FN",
    "67_51FN": "21_51FN", "67_MTRF": "MTRF",
}


def load_changeset():
    """[(tipo A/D, nome, device_mapping)] — sinais que existem no modelo."""
    root = ET.parse(XML).getroot()
    out = []
    for rd in root.iter("ResourceDescription"):
        idel = rd.find("id")
        if idel is None or idel.get("type") not in ("DSIGNAL", "ASIGNAL"):
            continue
        props = {p.get("id"): p.get("value") for p in rd.iter("Property")}
        nm = props.get("IDOBJ_NAME")
        if nm:
            out.append(("A" if idel.get("type") == "ASIGNAL" else "D",
                        nm, props.get("SIGNAL_DEVICEMAPPING") or ""))
    return out


def main():
    idx = json.loads((DATA / "sigla_index.json").read_text(encoding="utf-8"))
    dig, anl = idx["DNP3_DiscreteSignals"], idx["DNP3_AnalogSignals"]

    cat = E.Catalog()
    SH = "DNP3_DiscreteSignals"
    i_odt = cat.label_index(SH, "Output Data Type")
    i_ctrl = cat.label_index(SH, "Control Codes")
    i_idt = cat.label_index(SH, "Input Data Type")

    def cmd_count(sigla):
        t = dig.get(sigla) or []
        ctrl = t[i_ctrl] if i_ctrl is not None and i_ctrl < len(t) else None
        if ctrl not in (None, ""):
            return len(str(ctrl).split(";"))
        odt = t[i_odt] if i_odt is not None and i_odt < len(t) else None
        return 1 if odt not in (None, "") else 0

    def is_double(sigla):
        """Ponto de posição (MultiCoord/DoubleBit) exige DUAS coords distintas."""
        t = dig.get(sigla) or []
        idt = t[i_idt] if i_idt is not None and i_idt < len(t) else None
        return str(idt) in ("MultiCoord", "DoubleBit")

    signals = load_changeset()
    print(f"changeset: {len(signals)} sinais do modelo")

    discrete, analog, missing = [], [], []
    seq = {"di": 0, "ai": 0}
    fill = FILL_START
    for tipo, nome, dm in sorted(signals, key=lambda x: (x[0], x[1])):
        parts = nome.split("_")
        sigla = "_".join(parts[3:]) if len(parts) > 3 else parts[-1]
        if tipo == "D" and sigla not in dig and sigla in FALLBACK:
            sigla = FALLBACK[sigla]
        if tipo == "A":
            if sigla not in anl:
                missing.append((tipo, nome, sigla)); continue
            analog.append({"sigla": sigla, "nome": nome, "escala": "",
                           "inCoord": seq["ai"], "aor": AOR, "deviceMapping": dm})
            seq["ai"] += 1
        else:
            if sigla not in dig:
                missing.append((tipo, nome, sigla)); continue
            cc = cmd_count(sigla)
            out = ""
            if cc:
                out = f"{fill};{fill}" if cc == 2 else fill
                fill += 1
            # ponto de posição (DoubleBit): 2 coordenadas DISTINTAS "n;n+1"
            if is_double(sigla):
                ic = f"{seq['di']};{seq['di'] + 1}"
                seq["di"] += 2
            else:
                ic = seq["di"]
                seq["di"] += 1
            discrete.append({"sigla": sigla, "nome": nome, "inCoord": ic,
                             "outCoord": out, "aor": AOR, "deviceMapping": dm})

    lista = {"discrete": discrete, "analog": analog, "discrete_analog": [],
             "inputErrors": [], "uncertain": set()}
    print(f"gerando: dig={len(discrete)} anl={len(analog)} | "
          f"SEM template na base: {len(missing)}")
    for t, nm, s in missing:
        print(f"   [FALTA NA BASE] {nm} (sigla {s})")

    tdt, report = E.generate_tdt_from_list(lista, protocol="dnp3", native=False)

    # ── pós-processamento (regras do validador ADMS) ─────────────────────────
    import io as _io
    import openpyxl as _px
    import excel_native
    wb = _px.load_workbook(_io.BytesIO(tdt))
    ws = wb["DNP3_DiscreteSignals"]
    lab = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)}
    cN, cST, cDM, cNV = lab["Signal Name"], lab["Signal Type"], lab["Device Mapping"], lab["Normal Value"]
    fixed_rt = fixed_nv = 0
    for r in range(5, ws.max_row + 1):
        nome = ws.cell(r, cN).value
        if not nome:
            continue
        dm = str(ws.cell(r, cDM).value or "")
        # RelayTrip só pode referenciar elemento de proteção; em device físico
        # (disjuntor/seccionadora) o tipo tem de ser Custom
        if str(ws.cell(r, cST).value) == "RelayTrip" and re.search(r"_(DJ|SEC)$", dm):
            ws.cell(r, cST).value = "Custom"
            fixed_rt += 1
        # DJF1: Normal Value deve corresponder ao normal Open do disjuntor no modelo
        if str(nome).endswith("_DJF1"):
            ws.cell(r, cNV).value = 1
            fixed_nv += 1
    print(f"pos-fix: RelayTrip->Custom={fixed_rt}, NormalValue DJF1->1={fixed_nv}")

    buf = _io.BytesIO()
    wb.save(buf)
    tdt = excel_native.resave_native(buf.getvalue())
    OUT.write_bytes(tdt)
    print(f"TDT salva: {OUT} ({len(tdt)} bytes) | "
          f"ok dig={report['discrete']['matched']} anl={report['analog']['matched']}")


if __name__ == "__main__":
    main()
