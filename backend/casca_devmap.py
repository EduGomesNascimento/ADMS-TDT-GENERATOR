"""
casca_devmap.py — Device Mapping da SE CASCA.

FONTE DA VERDADE: PT-MOD-SE-CASCA.xml (export do changeset do modelo
"Casca_Obra"). Cada elemento traz a propriedade 1224979098644840199, que é o
"ID de Mapeamento SCADA" — exatamente o texto que a coluna Device Mapping da
TDT tem que conter. Se o texto não existir lá, o ADMS responde
"Could not find any device that corresponds to Device Mapping: ...".

O QUE O MODELO TEM (clone da CASCA ATUAL):
  alimentadores AL12 AL13 AL14 AL15 AL21 · linhas LTSCO LTPRI LTMRU
  trafos TR1(3W) TR2(2W) + TR12AT TR1AT TR1BT TR2AT TR2BT
  barras B138 BP23 · TSA3 · LTGPR
O QUE A LISTA NOVA PEDE (subestação modernizada):
  AL12..AL15 AL21 AL24 AL25 AL26 · BC1(AL18) BC2(AL28) · LT1 LT2 LT3
  TR6 TR7 (+AT/BT) · IB20 · transferências 24-1(AL18) 24-2(TRF29)
  BP69 BP113.8 BP213.8 · TSA1 TSA2 (RET1 RET2)

Ou seja: boa parte dos dispositivos AINDA NÃO EXISTE no modelo. Isso não se
resolve na TDT — tem que ser criado no diagrama. O que este módulo faz:

  1) casa pelo texto exato quando ele já existe no modelo;
  2) casa por MÓDULO + SUFIXO quando só o número do equipamento mudou
     (o alimentador AL12 do modelo usa 52-2, a lista nova usa 52-12 — é o
     mesmo vão, então vale CAS_AL12_52-2_DJ);
  3) dentro de um módulo que existe, se o relé específico não existe cai para
     o relé genérico _PROT e depois para o disjuntor _DJ
     ("na dúvida, põe no disjuntor" — instrução do usuário);
  4) módulo inexistente: devolve o nome canônico e MARCA como pendente, para
     o relatório listar o que precisa ser criado no Casca_Obra.

Uso: from casca_devmap import resolver; dm, origem, pendente = resolver(nome, sigla)
"""
from __future__ import annotations
import collections
import json
import re
from pathlib import Path
import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# REGRA DO PROJETO (decisão do usuário):
#   "seguir com as informações do unifilar, só usando os sinais da lista mesmo,
#    mas o que está no campo real é o que está no adms/unifilar"
#
# Ou seja: a LISTA manda nos SINAIS (nome, tipo, índice, comando, escala) e o
# UNIFILAR/ADMS manda nos DISPOSITIVOS. O Device Mapping tem que apontar para
# o que existe no Casca_Obra, mesmo quando a lista usa outra numeração.
#
# A lista renumerou tudo em relação ao unifilar:
#   unifilar : BARRA P1/P2 138 kV · BARRA P3/T1 23 kV · TR1 15/20/25 MVA ·
#              TR2 10/12,5 MVA · LT KVM/PRI/SCO · AL12..AL15 AL21 · TSA-3
#   lista    : LT 69 kV (LT1 LT2 LT3) · TR 69/13,8 kV (TR6 TR7) · BP69 ·
#              BP1 13.8 / BP2 13.8 · AL12..AL15 AL21 AL24 AL25 AL26 ·
#              BC1 BC2 · interbarras 20 · transf. 24-1 24-2 · TSA1 TSA2
REAPROVEITAR_DISPOSITIVO_ANTIGO = True

# ── equivalência MÓDULO da lista -> MÓDULO do unifilar/ADMS ──────────────────
# (destino, confianca, evidencia). Editar SÓ aqui muda todo o mapeamento.
# Módulo que não está nesta tabela e não existe no modelo fica PENDENTE.
MODULO_EQUIV = {
    # ── linhas 138 kV: a chave de aterramento (SECG 29-x) manteve o número ──
    "LT1": ("LTSCO", "ALTA",
            "SECG 29-1 na lista = 29-01 do LT SCO no unifilar"),
    "LT2": ("LTPRI", "ALTA",
            "SECG 29-3 na lista = 29-03 do LT PRI no unifilar"),
    "LT3": ("LTMRU", "ALTA",
            "SECG 29-5 na lista = 29-05 do LT KVM (LTMRU no modelo)"),
    # ── transformadores: aba 'TR 1' = modulo 6, aba 'TR 2' = modulo 7 ──
    "TR6":   ("TR1",   "MEDIA", "aba 'TR 1' da lista -> TR1 do unifilar"),
    "TR6AT": ("TR1AT", "MEDIA", "lado AT do TR1"),
    "TR6BT": ("TR1BT", "MEDIA", "lado BT do TR1"),
    "TR7":   ("TR2",   "MEDIA", "aba 'TR 2' da lista -> TR2 do unifilar"),
    "TR7AT": ("TR2AT", "MEDIA", "lado AT do TR2"),
    "TR7BT": ("TR2BT", "MEDIA", "lado BT do TR2"),
    # ── barras ──
    "BP69":    ("B138", "MEDIA", "barra de ALTA; no unifilar e a BARRA P1 138 kV"),
    "BP113.8": ("BP23", "MEDIA", "barra de BAIXA 1; no unifilar e a BARRA P3 23 kV"),
    # BT23 nao existe na TDT antiga — foi criado agora no modelo. E a BARRA T1
    # 23 kV do unifilar, que corresponde a 2a barra de 13,8 kV da lista.
    "BP213.8": ("BT23", "MEDIA", "barra de BAIXA 2; no unifilar e a BARRA T1 23 kV"),
    # ── servico auxiliar: o unifilar so tem o TSA-3 ──
    "TSA1": ("TSA3", "BAIXA", "unico TSA do unifilar (TSA-3 45 kVA)"),
    "TSA":  ("TSA3", "BAIXA", "modulo generico das abas RET"),
    # SEM equivalente no unifilar (o vao nao existe no campo hoje):
    #   AL24 AL25 AL26 · AL18 (BC 1 e transf. 24-1) · AL28 (BC 2) ·
    #   TRF29 (transf. 24-2) · IB20 (interbarras BT) · BP213.8 · TSA2
    # Esses ficam com o nome canonico e entram na aba 13 do relatorio.
}

MODELO = Path("C:/Users/egnpo/Downloads/PT-MOD-SE-CASCA.xml")
# Export do modelo DEPOIS de o usuario renomear os dispositivos com _NEW.
# Usado so para CONFERIR (aba 21 do relatorio) — nao alimenta a resolucao.
MODELO_NEW = Path("C:/Users/egnpo/Downloads/PT-MOD-SE-CAS.xml")
TDT_ATUAL = Path("C:/Users/egnpo/Downloads/CASCA.xlsx")
PROP_SCADA_ID = "1224979098644840199"
HEADER_ROWS = 4

# ── sufixo (tipo de dispositivo) por sigla — aprendido da TDT atual da CASCA ──
_FAM = [
    # religador: na TDT atual TODO 79* aponta pro disjuntor, não pro relé
    (("79", "79OK", "79LO", "79_1", "79_2", "79RE", "79EP", "79BL"), "DJ"),
    (("IA", "IB", "IC", "IN", "IACC", "IBCC", "ICCC", "INCC", "I"), "TC"),
    (("P", "Q", "S", "FP", "COS"), "TC"),
    (("V", "VA", "VB", "VC", "VA_B", "VB_B", "VC_B", "VA_L", "VB_L", "VC_L",
      "VAB", "VBC", "VCA", "F", "FREQ"), "TP"),
    (("TAP", "CDC", "BCDC", "DCDC", "CDMT", "MDCD", "CDAM", "CDLR", "RLCD",
      "63CA", "63CD", "20CA", "20CD", "27CD", "59CD", "86CC", "71C", "71HI",
      "71LO", "TOC"), "COMTAP"),
    (("R90", "FC90", "DR90", "R90A", "R90B"), "TAP_REG"),
    (("VF", "FVF", "TOA", "TOD", "TOLE", "TEA", "TED", "TENR", "63T", "63TA",
      "63TD", "71T", "20A", "20D", "20TA", "20TD", "SCAR", "MEMB", "TOED",
      "DRT1", "DRT2", "FCT1", "FCT2", "FCPL", "86", "87", "87_T"), "TR"),
    (("SECF", "SECC", "SECT", "SECG", "SECB", "SELF", "SEC", "LIBM", "DSEC",
      "SLIB", "SECD"), "SEC"),
]
_POR_DEVICE = [("52-", "DJ"), ("89-", "SEC"), ("29-", "SEC"), ("24-", "DJ")]
_MEDIDA_TC = {"IA", "IB", "IC", "IN", "IACC", "IBCC", "ICCC", "INCC", "P", "Q", "S"}
_MEDIDA_TP = {"V", "VA", "VB", "VC", "VA_B", "VB_B", "VC_B", "VA_L", "VB_L",
              "VC_L", "VAB", "VBC", "VCA", "F"}
# ordem de rebaixamento dentro de um módulo que existe no modelo
# Rebaixamento dentro de um vão que EXISTE no unifilar mas não tem aquele
# dispositivo. Ex.: o TR1 do unifilar entra por seccionadora (89-12), não tem
# disjuntor de alta — os sinais do "52-4" da lista vão para a seccionadora.
# BP = barra (BUSBAR) e BT = barra de baixa; sao o mesmo papel para o
# rebaixamento — o vao BT23 do modelo so tem o sufixo BT.
_DEGRADA = {
    "COMTAP": ("TR", "DJ", "SEC"), "TAP_REG": ("TR", "DJ", "SEC"),
    "TR": ("BP", "BT", "DJ", "SEC"), "BP": ("BT", "TP", "TC", "DJ"),
    "BT": ("BP", "TP", "TC", "DJ"),
    "TC": ("TP", "DJ", "SEC"), "TP": ("TC", "DJ", "SEC"),
    "SEC": ("DJ", "TR", "BP", "BT"), "RET": ("TC", "DJ"),
    "DJ": ("SEC", "TR", "BP", "BT", "RET", "TC"),
}
# última tentativa: qualquer dispositivo do vão, nesta ordem de preferência
_ULTIMO_RECURSO = ("DJ", "SEC", "PROT", "TR", "BP", "BT", "TC", "TP", "RET")

_CACHE = {}


# ─── Device Mapping da TDT ORIGINAL da subestação ────────────────────────────
def dm_da_tdt_original():
    """(modulo, sigla) -> Device Mapping usado HOJE na CASCA.

    "o device mapping é o mesmo da tdt original da subestação": onde a TDT atual
    já define o Device Mapping de um sinal, é esse que vale — ele funciona em
    produção. Isso corrige casos que nenhuma regra adivinha, por exemplo:
      - as medidas (IA/IB/IC/P/Q) apontam pro DISJUNTOR do vão, não pro TC
      - auxiliares de seccionadora às vezes apontam pro disjuntor
      - o TAP do TR1 usa CAS_TR1_TR1_COMTAP / _TAP_REG, dispositivos que nem
        aparecem no XML do changeset (o export é um delta, não o modelo inteiro)
    """
    if "orig" in _CACHE:
        return _CACHE["orig"]
    por = collections.defaultdict(collections.Counter)
    todos = set()
    if TDT_ATUAL.exists():
        wb = openpyxl.load_workbook(TDT_ATUAL, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if "Signals" not in sn and "DiscreteAnalog" not in sn:
                continue
            rows = list(wb[sn].iter_rows(values_only=True))
            if len(rows) <= HEADER_ROWS:
                continue
            hdr = [str(c or "").strip() for c in rows[HEADER_ROWS - 1]]
            ix = {n: i for i, n in enumerate(hdr) if n}
            cD = ix.get("Device Mapping")
            if cD is None:
                continue
            for r in rows[HEADER_ROWS:]:
                if not r or not r[0]:
                    continue
                nome, dm = str(r[0]).strip(), str(r[cD] or "").strip()
                p = nome.split("_")
                if not dm or not nome.startswith("CAS_") or len(p) < 4:
                    continue
                por[(p[1], "_".join(p[3:]))][dm] += 1
                todos.add(dm)
        wb.close()
    _CACHE["orig"] = ({k: c.most_common(1)[0][0] for k, c in por.items()}, todos)
    return _CACHE["orig"]


# ─── catálogo do modelo (XML do changeset) ───────────────────────────────────
def catalogo():
    if "cat" in _CACHE:
        return _CACHE["cat"]
    validos, por_mod, tipos = set(), collections.defaultdict(dict), {}
    if MODELO.exists():
        txt = MODELO.read_text(encoding="utf-8-sig", errors="replace")
        for b in re.findall(r"<ResourceDescription>(.*?)</ResourceDescription>",
                            txt, re.S):
            m = re.search(rf'id="{PROP_SCADA_ID}" value="([^"]*)"', b)
            if not m or not m.group(1):
                continue
            dm = m.group(1)
            t = re.search(r'type="([^"]+)"', b)
            validos.add(dm)
            tipos.setdefault(dm, t.group(1) if t else "?")
            p = dm.split("_")
            if len(p) >= 3:
                por_mod[p[1]].setdefault("_".join(p[3:]) if len(p) > 3 else "", dm)
    # O XML e um DELTA do changeset, nao o modelo inteiro: dispositivos que a
    # TDT atual ja referencia (ex.: CAS_TR1_TR1_COMTAP) existem no modelo mesmo
    # sem aparecer no export. Contam como validos e alimentam o rebaixamento.
    _, do_tdt = dm_da_tdt_original()
    for dm in do_tdt:
        validos.add(dm)
        tipos.setdefault(dm, "(da TDT atual)")
        p = dm.split("_")
        if len(p) >= 3:
            por_mod[p[1]].setdefault("_".join(p[3:]) if len(p) > 3 else "", dm)
    _CACHE["cat"] = {"validos": validos, "por_mod": dict(por_mod), "tipos": tipos}
    return _CACHE["cat"]


# ─── sufixo por sigla (aprendido da TDT atual) ───────────────────────────────
def _aprender_siglas():
    if "sig" in _CACHE:
        return _CACHE["sig"]
    per = collections.defaultdict(collections.Counter)
    if TDT_ATUAL.exists():
        wb = openpyxl.load_workbook(TDT_ATUAL, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if "Signals" not in sn and "DiscreteAnalog" not in sn:
                continue
            rows = list(wb[sn].iter_rows(values_only=True))
            if len(rows) <= HEADER_ROWS:
                continue
            hdr = [str(c or "").strip() for c in rows[HEADER_ROWS - 1]]
            ix = {n: i for i, n in enumerate(hdr) if n}
            cD = ix.get("Device Mapping")
            if cD is None:
                continue
            for r in rows[HEADER_ROWS:]:
                if not r or not r[0]:
                    continue
                nome, dm = str(r[0]).strip(), str(r[cD] or "").strip()
                p = nome.split("_")
                if not dm or not nome.startswith("CAS_") or len(p) < 4:
                    continue
                pref = f"CAS_{p[1]}_{p[2]}_"
                if dm.startswith(pref):
                    per["_".join(p[3:])][dm[len(pref):]] += 1
        wb.close()
    _CACHE["sig"] = {s: c.most_common(1)[0] for s, c in per.items()}  # (sufixo, n)
    return _CACHE["sig"]


def sufixo(sig: str, device: str) -> tuple[str, str]:
    """Tipo de dispositivo que leva esta sigla.

    Ordem das fontes, da mais forte para a mais fraca:
      1) TDT atual da CASCA, com PELO MENOS 2 ocorrencias — convencao local
      2) familia conhecida (medida, trafo, seccionadora, religador)
      3) BASE COMPLETA (27 subestacoes) — 2692 siglas
      4) TDT da CASCA com 1 ocorrencia so
      5) pickup/alarme, medida, tipo do equipamento, codigo ANSI
    O passo 4 vem DEPOIS da base de proposito: uma amostra unica da CASCA nao
    deve derrubar a convencao de 27 subestacoes. Foi o que acontecia com o
    '81' — a CASCA tem UM caso (CAS_LTPRI_LTPRI_81 -> PROT_81SU) e a base
    inteira diz PROT_81, que e o rele que existe no modelo.
    """
    tab = _aprender_siglas()
    base = convencao_base()["sigla_sufixo"]
    if sig in tab and tab[sig][1] >= 2:
        return tab[sig][0], "TDT atual"
    for fam, suf in _FAM:
        if sig in fam:
            return suf, "familia"
    if sig in base:
        return base[sig], "base completa"
    if sig in tab:
        return tab[sig][0], "TDT atual (1 caso)"
    # PICKUP (P_) e ALARME (A_) tem dispositivo PROPRIO na CASCA:
    #   CAS_LTPRI_LTPRI_P_51N1 -> CAS_LTPRI_LTPRI_P_PROT_51N1
    #   CAS_LTPRI_LTPRI_A_51N1 -> CAS_LTPRI_LTPRI_A_PROT_51N1
    # Mandar o pickup para o mesmo PROT_51N1 do trip faz o ADMS recusar com
    # "Same signal ... was already mapped on device".
    if len(sig) > 2 and sig[:2] in ("P_", "A_"):
        resto = sig[2:]
        if resto[:1].isdigit():
            return f"{sig[0]}_PROT_{resto}", "pickup/alarme"
        if resto in tab:
            return tab[resto][0], "TDT atual (base)"
        for fam, suf in _FAM:
            if resto in fam:
                return suf, "familia (base)"
        if resto in base:
            return base[resto], "base completa (base)"
        return "DJ", "default"
    if sig in _MEDIDA_TC:
        return "TC", "medida"
    if sig in _MEDIDA_TP:
        return "TP", "medida"
    d = str(device or "")
    for pre, suf in _POR_DEVICE:
        if d.startswith(pre):
            if sig[:1].isdigit():
                return f"PROT_{sig}", "ANSI"
            return suf, "device"
    if sig[:1].isdigit():
        return f"PROT_{sig}", "ANSI"
    return "DJ", "default"


# ─── resolução contra o modelo ───────────────────────────────────────────────
def catalogo_estrito() -> set[str]:
    """SÓ os Device Mappings que existem na TDT atual da CASCA (CASCA.xlsx).

    "todos os device mapping necessarios estao aqui. nao existe outros."
    São 332 valores CAS_*; os ~158 numéricos são dos religadores RAD_* e não
    servem para a UTR nova.
    """
    if "estrito" not in _CACHE:
        _, todos = dm_da_tdt_original()
        _CACHE["estrito"] = {d for d in todos if d.startswith("CAS_")}
    return _CACHE["estrito"]


def resolver_estrito(nome: str, sigla: str) -> tuple[str | None, str]:
    """(device_mapping, origem) usando SOMENTE o catálogo da CASCA.xlsx.
    Devolve (None, motivo) quando nao ha alvo valido — o sinal fica de fora."""
    val = catalogo_estrito()
    p = str(nome).split("_")
    alias = p[0] if p else "CAS"
    mod = p[1] if len(p) > 1 else ""
    dev = p[2] if len(p) > 2 else mod
    suf, _ = sufixo(sigla, dev)

    equiv = MODULO_EQUIV.get(mod)
    mod_alvo = equiv[0] if equiv else mod
    nota = f" [{mod}->{mod_alvo}]" if equiv else ""

    # 1) o que a TDT atual usa para este vao + esta sigla
    orig, _t = dm_da_tdt_original()
    achado = orig.get((mod_alvo, sigla))
    if achado in val:
        return achado, f"TDT atual: (vao, sigla){nota}"
    # 2) mesmo vao, mesmo device, mesmo sufixo
    cand = f"{alias}_{mod_alvo}_{dev}_{suf}"
    if cand in val:
        return cand, f"TDT atual: exato{nota}"
    # 3) mesmo vao + mesmo sufixo, equipamento renumerado
    do_vao = {d: d.split("_", 3)[3] if d.count("_") > 2 else ""
              for d in val if d.split("_")[1:2] == [mod_alvo]}
    porsuf = {}
    for d, s in do_vao.items():
        porsuf.setdefault(s, d)
    if suf in porsuf:
        return porsuf[suf], f"TDT atual: equip. renumerado{nota}"
    # 4) rele especifico inexistente -> rele generico do vao
    if suf.startswith("PROT_") and "PROT" in porsuf:
        return porsuf["PROT"], f"TDT atual: rele generico{nota}"
    # 5) rebaixa por tipo de dispositivo, depois ultimo recurso
    for alt in _DEGRADA.get(suf, ()) + _ULTIMO_RECURSO:
        if alt in porsuf:
            return porsuf[alt], f"TDT atual: fallback {alt}{nota}"
    return None, (f"vao {mod} nao tem equivalente na TDT atual da CASCA"
                  if not do_vao else f"vao {mod_alvo} nao tem {suf}")


def resolver(nome: str, sigla: str) -> tuple[str, str, str]:
    """(device_mapping, origem_da_regra, pendencia) — pendencia vazia = ok."""
    cat = catalogo()
    p = str(nome).split("_")
    alias = p[0] if p else "CAS"
    mod = p[1] if len(p) > 1 else ""
    dev = p[2] if len(p) > 2 else mod
    suf, origem = sufixo(sigla, dev)
    canonico = f"{alias}_{mod}_{dev}_{suf}"

    # módulo da lista -> módulo do unifilar (o campo real manda no dispositivo)
    equiv = MODULO_EQUIV.get(mod)
    mod_alvo = equiv[0] if equiv else mod
    nota = f" [{mod}->{mod_alvo}]" if equiv else ""

    # 0) O QUE A TDT ORIGINAL JA USA para este vão + esta sigla. Tem prioridade
    #    sobre qualquer regra: é o valor que funciona em produção hoje.
    orig_por_chave, _ = dm_da_tdt_original()
    achado = orig_por_chave.get((mod_alvo, sigla))
    if achado:
        return achado, f"TDT original da CASCA{nota}", ""

    if not cat["validos"]:                       # sem XML: fica no canônico
        return canonico, origem, ""
    if canonico in cat["validos"]:
        return canonico, f"{origem} + modelo (exato)", ""

    domod = cat["por_mod"].get(mod_alvo) if REAPROVEITAR_DISPOSITIVO_ANTIGO else None
    if domod:
        # o alvo pode ter o sufixo com o device do unifilar; tenta o exato dele
        if equiv and f"{alias}_{mod_alvo}_{dev}_{suf}" in cat["validos"]:
            return f"{alias}_{mod_alvo}_{dev}_{suf}", \
                f"{origem} + unifilar (exato){nota}", ""
        # 1) mesmo sufixo, equipamento renumerado (AL12: 52-12 -> 52-2)
        if suf in domod:
            return domod[suf], f"{origem} + unifilar (equip. renumerado){nota}", ""
        # 2) relé específico não existe -> relé genérico do vão
        if suf.startswith("PROT_") and "PROT" in domod:
            return domod["PROT"], f"{origem} + unifilar (rele generico){nota}", \
                f"rele {suf} nao existe no vao {mod_alvo} do unifilar"
        # 3) rebaixa por tipo de dispositivo
        for alt in _DEGRADA.get(suf, ()):
            if alt in domod:
                return domod[alt], f"{origem} + unifilar (fallback {alt}){nota}", \
                    f"{suf} nao existe no vao {mod_alvo} do unifilar"
        # 4) último recurso: qualquer dispositivo do vão
        for alt in _ULTIMO_RECURSO:
            if alt in domod:
                return domod[alt], f"{origem} + unifilar (ultimo recurso {alt}){nota}", \
                    f"{suf} nao existe no vao {mod_alvo} do unifilar"
    existe_mod = mod_alvo in cat["por_mod"]
    return canonico, origem, (f"criar {suf} no vao {mod_alvo} do Casca_Obra"
                              if existe_mod
                              else f"vao {mod} nao existe no unifilar/ADMS")


def ambiguos_no_modelo() -> dict[str, list[tuple[str, str]]]:
    """IDs de Mapeamento SCADA usados por MAIS DE UM elemento DENTRO do próprio
    Casca_Obra. Nesses o ADMS responde "Found multiple devices..." e a coluna
    Substation não resolve (os dois candidatos estão na mesma subestação) —
    só dando ID único a cada dispositivo no modelo."""
    if "amb" in _CACHE:
        return _CACHE["amb"]
    por = collections.defaultdict(list)
    if MODELO.exists():
        txt = MODELO.read_text(encoding="utf-8-sig", errors="replace")
        for b in re.findall(r"<ResourceDescription>(.*?)</ResourceDescription>",
                            txt, re.S):
            m = re.search(rf'id="{PROP_SCADA_ID}" value="([^"]+)"', b)
            if not m:
                continue
            t = re.search(r'type="([^"]+)"', b)
            n = re.search(r'id="IDOBJ_NAME" value="([^"]*)"', b)
            por[m.group(1)].append((t.group(1) if t else "?",
                                    n.group(1) if n else ""))
    _CACHE["amb"] = {k: v for k, v in por.items() if len(v) > 1}
    return _CACHE["amb"]


def nome_do_dispositivo(dm: str) -> str:
    """Nome (IDOBJ_NAME) do dispositivo do Casca_Obra que tem este ID de
    Mapeamento SCADA — só quando é UM só dentro do modelo.

    Serve para preencher a coluna Device da TDT. Como o Casca_Obra é CÓPIA da
    CASCA, os dois têm o MESMO ID; Substation sozinho não vinha resolvendo os
    sinais discretos, e o par (Substation + Device) aperta a busca.
    Nos 7 IDs que se repetem DENTRO do próprio Casca_Obra fica em branco: ali
    há dois candidatos legítimos e a lista não diz qual é qual.
    """
    if "nomedisp" not in _CACHE:
        por = collections.defaultdict(set)
        if MODELO.exists():
            txt = MODELO.read_text(encoding="utf-8-sig", errors="replace")
            for b in re.findall(r"<ResourceDescription>(.*?)</ResourceDescription>",
                                txt, re.S):
                m = re.search(rf'id="{PROP_SCADA_ID}" value="([^"]+)"', b)
                n = re.search(r'id="IDOBJ_NAME" value="([^"]*)"', b)
                if m and n and n.group(1):
                    por[m.group(1)].add(n.group(1))
        _CACHE["nomedisp"] = {k: next(iter(v)) for k, v in por.items() if len(v) == 1}
    return _CACHE["nomedisp"].get(dm, "")


def modelo_new():
    """Do PT-MOD-SE-CAS.xml (modelo ja com _NEW): (ids_validos, ids_duplicados).
    ids_duplicados = os que aparecem em 2+ dispositivos (o ADMS da 'Found
    multiple')."""
    if "modelo_new" not in _CACHE:
        por = collections.defaultdict(int)
        if MODELO_NEW.exists():
            txt = MODELO_NEW.read_text(encoding="utf-8-sig", errors="replace")
            for m in re.findall(rf'id="{PROP_SCADA_ID}" value="([^"]+)"', txt):
                por[m] += 1
        _CACHE["modelo_new"] = (set(por), {k for k, v in por.items() if v > 1})
    return _CACHE["modelo_new"]


def _vao_suf(dm: str) -> tuple[str, str]:
    """'CAS_LTKVM_52-1_DJ_NEW' -> ('LTKVM', 'DJ_NEW'); o device fica de fora."""
    p = dm.split("_")
    if len(p) < 4:
        return (p[1] if len(p) > 1 else "", "")
    return p[1], "_".join(p[3:])


def indice_modelo():
    """Do modelo ja renomeado: (vao, sufixo) -> ID completo, e as ALIAS de vao.

    O vao pode aparecer com DOIS nomes: a LT KVM tem as seccionadoras/TC sob
    'LTMRU' e o disjuntor/reles sob 'LTKVM'. A alias e derivada da TDT antiga —
    la o sinal CAS_LTKVM_LTKVM_PROT_21F aponta para CAS_LTMRU_LTMRU_PROT_21F,
    o que prova que os dois nomes sao o mesmo vao.
    """
    if "idx_modelo" not in _CACHE:
        val, _dup = modelo_new()
        por = {}
        for dm in val:
            v, s = _vao_suf(dm)
            if s:
                por.setdefault((v, s), dm)
        # Alias vao-do-sinal <-> vao-do-dispositivo, aprendida da TDT antiga.
        # So conta como alias o par com EVIDENCIA FORTE: >=5 sinais e >=60% dos
        # sinais daquele vao. Sem isso entram pares que sao mero fallback
        # (TR1BT->TR1 com 21%, AL13->AL12 com 2%...). Na CASCA passam so
        # LTKVM<->LTMRU (67 sinais, 94%) e DJHIB<->LTPRI (7, 87%).
        par = collections.Counter(); tot_vao = collections.Counter()
        alias = collections.defaultdict(set)
        if TDT_ATUAL.exists():
            wb = openpyxl.load_workbook(TDT_ATUAL, read_only=True, data_only=True)
            for sn in wb.sheetnames:
                if "Signals" not in sn and "DiscreteAnalog" not in sn:
                    continue
                rows = list(wb[sn].iter_rows(values_only=True))
                if len(rows) <= HEADER_ROWS:
                    continue
                hdr = [str(c or "").strip() for c in rows[HEADER_ROWS - 1]]
                ix = {n: i for i, n in enumerate(hdr) if n}
                cD = ix.get("Device Mapping")
                if cD is None:
                    continue
                for r in rows[HEADER_ROWS:]:
                    if not r or not r[0]:
                        continue
                    nome, dm = str(r[0]).strip(), str(r[cD] or "").strip()
                    pn, pd = nome.split("_"), dm.split("_")
                    if (len(pn) > 1 and len(pd) > 1 and nome.startswith("CAS_")
                            and dm.startswith("CAS_")):
                        tot_vao[pn[1]] += 1
                        if pn[1] != pd[1]:
                            par[(pn[1], pd[1])] += 1
            wb.close()
        for (a, b), n in par.items():
            if n >= 5 and n >= 0.60 * tot_vao[a]:
                alias[a].add(b)
                alias[b].add(a)
        _CACHE["idx_modelo"] = (por, {k: sorted(v) for k, v in alias.items()})
    return _CACHE["idx_modelo"]


def no_modelo(dm_com_sufixo: str) -> tuple[str | None, str]:
    """Acha o ID REAL do modelo para um Device Mapping ja com o sufixo.
    Devolve (id_real, nota). Tenta o texto exato e depois os nomes ALIAS do vao
    (LTMRU <-> LTKVM), porque o mesmo vao pode ter dois nomes no modelo."""
    por, alias = indice_modelo()
    if not por:
        return dm_com_sufixo, ""                 # sem XML: aceita como esta
    v, s = _vao_suf(dm_com_sufixo)
    if (v, s) in por and por[(v, s)] == dm_com_sufixo:
        return dm_com_sufixo, ""
    if (v, s) in por:
        return por[(v, s)], f"equipamento renumerado no modelo"
    for outro in alias.get(v, []):
        if (outro, s) in por:
            return por[(outro, s)], f"vao {v} esta como {outro} no modelo"
    return None, f"nao existe no modelo"


def so_no_modelo(nome: str, sigla: str, sufixo_id: str = "_NEW"):
    """Ultimo recurso: procura o dispositivo DIRETO no modelo novo, para os vaos
    que a TDT antiga nao tinha (ex.: o BT23 — BARRA T1 23 kV — foi criado agora
    e nao existe no CASCA.xlsx, entao o catalogo estrito nunca o acharia).
    Devolve (id_real_sem_sufixo, nota) ou (None, motivo)."""
    por, alias = indice_modelo()
    if not por:
        return None, "sem modelo"
    p = str(nome).split("_")
    alias_ = p[0] if p else "CAS"
    mod = p[1] if len(p) > 1 else ""
    dev = p[2] if len(p) > 2 else mod
    equiv = MODULO_EQUIV.get(mod)
    alvo = equiv[0] if equiv else mod
    suf, _o = sufixo(sigla, dev)
    for vao in [alvo] + list(alias.get(alvo, [])):
        for s in (f"{suf}{sufixo_id}",):
            if (vao, s) in por:
                dm = por[(vao, s)]
                base = dm[:-len(sufixo_id)] if dm.endswith(sufixo_id) else dm
                return base, f"so no modelo novo [{mod}->{vao}]"
        # rebaixa dentro do vao, so com o que o modelo tem
        do_vao = {k[1]: v for k, v in por.items() if k[0] == vao}
        for alt in _DEGRADA.get(suf, ()) + _ULTIMO_RECURSO:
            if f"{alt}{sufixo_id}" in do_vao:
                dm = do_vao[f"{alt}{sufixo_id}"]
                base = dm[:-len(sufixo_id)] if dm.endswith(sufixo_id) else dm
                return base, f"so no modelo novo, fallback {alt} [{mod}->{vao}]"
    return None, f"vao {alvo} nao existe nem no modelo novo"


def convencao_base():
    """Convencao de TODAS as subestacoes, extraida da base completa
    (backend/_extrai_base.py -> data/convencao_dm.json):
        sigla_sufixo — 2692 siglas -> tipo de dispositivo que a levam
        quota        — (tipo, Signal Type) -> maximo observado na base inteira
    A CASCA sozinha da 156 siglas; a base da 2692. Serve de segunda fonte
    quando a sigla nao aparece na TDT da propria CASCA.
    """
    if "conv" not in _CACHE:
        f = Path(__file__).parent / "data" / "convencao_dm.json"
        try:
            d = json.loads(f.read_text(encoding="utf-8"))
        except Exception:                                   # noqa: BLE001
            d = {"sigla_sufixo": {}, "quota": {}}
        _CACHE["conv"] = d
    return _CACHE["conv"]


def quota_por_tipo():
    """(tipo de dispositivo, Signal Type) -> quantos sinais cabem.

    Aprendido da TDT ATUAL da CASCA. E a regra que o ADMS aplica quando diz
    "Same signal ... was already mapped on device". Os numeros observados:
        PROT  RelayTrip 1 · Enabled 1 · Custom 1
        DJ    Custom 25 · MeasuredValue 5 · SwitchStatus 2 · Local 1
              (NENHUM RelayTrip nem Enabled — um disjuntor nao carrega trip
               de protecao neste modelo; por isso rebaixar um 51F para o DJ e
               rejeitado)
        SEC   Custom 3 · SwitchStatus 2
        TC/TP MeasuredValue 3..5      TR  Custom 46
    Tipo de dispositivo nao listado, ou par (tipo, Signal Type) inexistente na
    TDT antiga, significa que aquele sinal NAO vai naquele dispositivo.
    """
    if "quota" in _CACHE:
        return _CACHE["quota"]
    por = collections.Counter()
    if TDT_ATUAL.exists():
        wb = openpyxl.load_workbook(TDT_ATUAL, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if "Signals" not in sn and "DiscreteAnalog" not in sn:
                continue
            rows = list(wb[sn].iter_rows(values_only=True))
            if len(rows) <= HEADER_ROWS:
                continue
            hdr = [str(c or "").strip() for c in rows[HEADER_ROWS - 1]]
            ix = {n: i for i, n in enumerate(hdr) if n}
            cD, cT = ix.get("Device Mapping"), ix.get("Signal Type")
            if cD is None or cT is None:
                continue
            for r in rows[HEADER_ROWS:]:
                if not r or not r[0]:
                    continue
                d, t = str(r[cD] or "").strip(), str(r[cT] or "").strip()
                if d.startswith("CAS_"):
                    por[(d, t)] += 1
        wb.close()
    mx = collections.defaultdict(int)
    for (d, t), v in por.items():
        mx[(tipo_dispositivo(d), t)] = max(mx[(tipo_dispositivo(d), t)], v)
    # une com a convencao da BASE COMPLETA: o que vale e o maior dos dois. A
    # base confirma a regra principal — em NENHUMA subestacao um DISJUNTOR
    # carrega RelayTrip ou Enabled.
    for k, v in convencao_base().get("quota", {}).items():
        td, _, st = k.partition("|")
        mx[(td, st)] = max(mx[(td, st)], v)
    _CACHE["quota"] = dict(mx)
    return _CACHE["quota"]


def tipo_dispositivo(dm: str) -> str:
    """Sufixo que diz QUE tipo de dispositivo e: DJ, SEC, PROT, TR, TC, TP..."""
    s = dm.split("_", 3)[-1] if dm.count("_") > 2 else ""
    for p in ("PROT", "P_PROT", "A_PROT"):
        if s.startswith(p):
            return "PROT"
    return s.replace("_NEW", "") or "?"


def container_da_subestacao() -> tuple[str, str]:
    """(nome, custom id) da SUBSTATION do modelo — é ela que o campo Container
    da aba DNP3_RTUs referencia (no esqueleto da LVA era 'LAGOA VERMELHA 1').
    Mandar o Custom ID junto evita o erro 'Mandatory reference ... not found',
    que reprova a RTU e, em cascata, TODOS os sinais dela."""
    if "cont" in _CACHE:
        return _CACHE["cont"]
    nome = cid = ""
    if MODELO.exists():
        txt = MODELO.read_text(encoding="utf-8-sig", errors="replace")
        for b in re.findall(r"<ResourceDescription>(.*?)</ResourceDescription>",
                            txt, re.S):
            if 'type="SUBSTATION"' not in b:
                continue
            n = re.search(r'id="IDOBJ_NAME" value="([^"]*)"', b)
            c = re.search(r'id="IDOBJ_CUSTOMID" value="([^"]*)"', b)
            nome = n.group(1) if n else ""
            cid = c.group(1) if c else ""
            break
    _CACHE["cont"] = (nome, cid)
    return _CACHE["cont"]


def device_mapping(nome: str, sigla: str) -> tuple[str, str]:
    dm, origem, _ = resolver(nome, sigla)
    return dm, origem


if __name__ == "__main__":
    cat = catalogo()
    print(f"modelo: {len(cat['validos'])} IDs de mapeamento SCADA")
    print(f"modulos: {sorted(cat['por_mod'])}")
    print(f"siglas aprendidas da TDT atual: {len(_aprender_siglas())}")
