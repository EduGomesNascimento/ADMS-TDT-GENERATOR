"""
probability_report.py — Gera o relatório de probabilidades de mapeamento.

Formato igual ao TDT_GPR_dnp3_PROBABILIDADES_v2.xlsx:
  - Resumo   : stats por tipo de sinal
  - Discretos : tabela colorida com confiança
  - Analógicos: idem
"""
from __future__ import annotations

import io
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from ai_mapper import MappedSignal

# Paleta
_F_HEADER = PatternFill(fill_type='solid', fgColor='2F5496')
_F_ALTA   = PatternFill(fill_type='solid', fgColor='C6EFCE')
_F_MEDIA  = PatternFill(fill_type='solid', fgColor='FFEB9C')
_F_BAIXA  = PatternFill(fill_type='solid', fgColor='FFC7CE')
_F_SEM    = PatternFill(fill_type='solid', fgColor='D9D9D9')
_FONT_HDR = Font(color='FFFFFF', bold=True, size=10)
_THIN     = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)

_CONF_FILL = {'ALTA': _F_ALTA, 'MÉDIA': _F_MEDIA, 'BAIXA': _F_BAIXA, 'SEM': _F_SEM}


def _header_row(ws, headers: list):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _F_HEADER
        c.font = _FONT_HDR
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = _THIN


def _data_row(ws, row_idx: int, values: list, fill=None):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=col, value=v)
        if fill:
            c.fill = fill
        c.border = _THIN
        c.alignment = Alignment(vertical='center')


def _col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def build_probability_xlsx(
    mapped: list[MappedSignal],
    alias: str = '',
    source_file: str = '',
) -> bytes:
    wb = openpyxl.Workbook()

    discrete = [m for m in mapped if m.signal_type not in ('analog',)]
    analog   = [m for m in mapped if m.signal_type == 'analog']

    def _stats(lst):
        total    = len(lst)
        w_addr   = sum(1 for m in lst if m.dnp3_addr is not None)
        alta     = sum(1 for m in lst if m.confidence_label == 'ALTA')
        media    = sum(1 for m in lst if m.confidence_label == 'MÉDIA')
        baixa    = sum(1 for m in lst if m.confidence_label == 'BAIXA')
        sem      = sum(1 for m in lst if m.confidence_label == 'SEM')
        return total, w_addr, alta, media, baixa, sem

    # ── Resumo ───────────────────────────────────────────────────────────────
    ws_r = wb.active
    ws_r.title = 'Resumo'
    title = f"Mapeamento DNP3 → TDT RTU {alias}"
    if source_file:
        title += f" — revisão completa contra {source_file}"
    ws_r.merge_cells('A1:H1')
    c = ws_r['A1']
    c.value = title
    c.font  = Font(bold=True, size=12)
    ws_r.row_dimensions[1].height = 20

    headers_r = ['Tipo', 'Total', 'Com endereço', 'ALTA', 'MÉDIA', 'BAIXA', 'Sem corresp.', 'Saídas (cmds)']
    for col, h in enumerate(headers_r, 1):
        cell = ws_r.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
    rows_r = [
        ('Discretos (Sinalização/Comando)', *_stats(discrete)),
        ('Analógicos (Medição)', *_stats(analog)),
    ]
    for r_idx, row in enumerate(rows_r, start=4):
        for col, v in enumerate(row, 1):
            ws_r.cell(row=r_idx, column=col, value=v)
    ws_r.column_dimensions['A'].width = 38

    # Notas
    ws_r.cell(row=7, column=1, value='Notas:').font = Font(bold=True)
    ws_r.cell(row=8, column=1,
              value='• ALTA (≥90%) = mapeamento certo. MÉDIA (70-89%) = revisar. BAIXA (<70%) = verificar manualmente.')
    ws_r.cell(row=9, column=1,
              value='• "Sem corresp." = sinal não encontrado na base ADMS. Adicione à base e reprocesse.')

    # ── Discretos ────────────────────────────────────────────────────────────
    ws_d = wb.create_sheet('Discretos')
    _header_row(ws_d, [
        'Sinal TDT', 'Bay', 'Descrição TDT',
        'End. DNP3 Entrada', 'End. DNP3 Saída (comando)',
        'Ponto na UTR (lista mestre)', 'Descrição na lista mestre',
        'Prob. (%)', 'Confiança', 'Alternativa',
    ])
    for r_idx, m in enumerate(discrete, start=2):
        tdt = f"{alias}_{m.module}_{m.module}_{m.sigla}" if m.sigla else ''
        fill = _CONF_FILL.get(m.confidence_label, _F_SEM)
        out_coord = m.dnp3_addr if m.signal_type == 'command' else None
        in_coord  = m.dnp3_addr if m.signal_type != 'command' else None
        _data_row(ws_d, r_idx, [
            tdt, m.module, m.sigla_desc or '',
            in_coord, out_coord,
            m.utr_id, m.description,
            m.confidence, m.confidence_label, m.alternative,
        ], fill)
    _col_widths(ws_d, [32, 10, 28, 16, 22, 30, 35, 9, 10, 20])
    ws_d.freeze_panes = 'A2'

    # ── Analógicos ───────────────────────────────────────────────────────────
    ws_a = wb.create_sheet('Analógicos')
    _header_row(ws_a, [
        'Sinal TDT', 'Bay', 'Descrição TDT',
        'End. DNP3', 'End. DNP3 Saída',
        'Ponto na UTR (lista mestre)', 'Descrição na lista mestre',
        'Prob. (%)', 'Confiança', 'Alternativa',
    ])
    for r_idx, m in enumerate(analog, start=2):
        tdt = f"{alias}_{m.module}_{m.module}_{m.sigla}" if m.sigla else ''
        fill = _CONF_FILL.get(m.confidence_label, _F_SEM)
        _data_row(ws_a, r_idx, [
            tdt, m.module, m.sigla_desc or '',
            m.dnp3_addr, None,
            m.utr_id, m.description,
            m.confidence, m.confidence_label, m.alternative,
        ], fill)
    _col_widths(ws_a, [32, 10, 28, 10, 14, 30, 35, 9, 10, 20])
    ws_a.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
