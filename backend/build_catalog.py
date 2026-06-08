"""
build_catalog.py
=================
Camada de APRENDIZAGEM. Lê as TDTs reais + o dicionário oficial de sinais
(RGE ADMS) e produz `data/catalog.json` — a base de conhecimento que a aplicação
usa para liberar os sinais por tipo de equipamento e clonar fielmente as linhas.

Princípios (descobertos analisando os arquivos reais):
- Cada TDT tem as sheets DNP3_DiscreteSignals e DNP3_AnalogSignals com 4 linhas
  de cabeçalho (seção / tabela / código de campo / rótulo) e dados a partir da 5ª.
- O nome do sinal segue: {ALIAS}_{MODULO}_{DEVICE}_{SUFIXO}. O sufixo é a chave
  do tipo de sinal (mapeado no dicionário).
- Campos que dependem da identidade do equipamento são detectados de forma
  GENÉRICA por tokenização: toda célula de texto que contém o prefixo/alias/
  módulo/device é convertida em placeholders (<<PREFIX>>, <<ALIAS>>, ...).
  Na exportação esses placeholders são substituídos pelos dados do novo
  equipamento. Assim "descobrimos quais campos mudam" sem hardcode.
"""
from __future__ import annotations
import json
import re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent
SRC = ROOT.parent.parent  # pasta "UI DE TDT" com os arquivos originais
DATA = ROOT / "data"
DATA.mkdir(exist_ok=True)

CLEAN_TDT = SRC / "TDT_DNP3_UTR_FWB_AL13_re.xlsx"
MULTI_TDT = SRC / "Export_TDT_DNP3_UTR_CNO_NEW_ALs_BCs_IB_Transf_LTSAS_LTKGT_TSA.xlsx"
DICT_XLSX = SRC / "RGE ADMS_Lista Sinais Padrão_v1.xlsx"

HEADER_ROWS = 4  # nº de linhas de cabeçalho antes dos dados
DATA_START = HEADER_ROWS  # índice 0-based da 1ª linha de dados

# ---------------------------------------------------------------------------
# Equipamentos-fonte: de onde extrair o catálogo de cada tipo.
# (alias, modulo, device) permitem identificar o prefixo e isolar o sufixo
# real do sinal mesmo quando o sufixo contém "_" (ex.: 79_1).
# ---------------------------------------------------------------------------
SOURCE_DEVICES = [
    {
        "id": "alimentador",
        "label": "Alimentador",
        "description": "Saída de distribuição (feeder) protegida por disjuntor.",
        "file": "CLEAN",
        "alias": "FWB", "module": "AL13", "device": "52-13",
        "sheets": ["DNP3_DiscreteSignals", "DNP3_AnalogSignals"],
        "defaults": {"module": "AL01", "device": "52-01"},
    },
    {
        "id": "tsa",
        "label": "Transformador Auxiliar (TSA)",
        "description": "Transformador de Serviço Auxiliar — alimenta os serviços auxiliares da subestação.",
        "file": "MULTI",
        "alias": "CNO_NEW", "module": "TRANSF", "device": "24-1",
        "sheets": ["DNP3_DiscreteSignals"],
        "defaults": {"module": "TSA", "device": "24-1"},
    },
]


def norm(v):
    """Normaliza valor de célula para algo serializável em JSON, preservando tipo."""
    return v


def load_dictionary():
    """suffix(UPPER) -> {description, signalType, measurementType, unit, phases, klass}"""
    wb = openpyxl.load_workbook(DICT_XLSX, read_only=True, data_only=True)
    out = {}

    def add(sheet, klass):
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            suf = str(r[0]).strip()
            entry = {
                "description": (str(r[1]).strip() if len(r) > 1 and r[1] else ""),
                "klass": klass,
            }
            if klass == "analog" and len(r) >= 6:
                entry["signalTypePt"] = (str(r[2]).strip() if r[2] else "")
                entry["measurementPt"] = (str(r[3]).strip() if r[3] else "")
                entry["unit"] = (str(r[4]).strip() if r[4] else "")
                entry["phasesPt"] = (str(r[5]).strip() if r[5] else "")
            elif klass == "discrete" and len(r) >= 3:
                entry["signalTypePt"] = (str(r[2]).strip() if r[2] else "")
            out.setdefault(suf.upper(), entry)

    add("Analog", "analog")
    add("Discrete", "discrete")
    add("Discrete Analog", "discrete_analog")
    return out


def sheet_meta(ws):
    """Extrai as 4 linhas de cabeçalho e a metadados de coluna de uma sheet."""
    rows = list(ws.iter_rows(values_only=True))
    header = [list(rows[i]) for i in range(HEADER_ROWS)]
    section, table, codes, labels = header
    columns = []
    for ci in range(len(labels)):
        columns.append({
            "index": ci,
            "section": section[ci] if ci < len(section) else None,
            "table": table[ci] if ci < len(table) else None,
            "code": codes[ci] if ci < len(codes) else None,
            "label": labels[ci] if ci < len(labels) else None,
        })
    return header, columns, rows


def build_tokenizer(alias, module, device):
    """Retorna (tokenize, prefix). Converte tokens da identidade em placeholders."""
    prefix = f"{alias}_{module}_{device}"
    # ordem importa: mais específico primeiro
    repls = [
        (prefix, "<<PREFIX>>"),
        (alias, "<<ALIAS>>"),
        (module, "<<MODULE>>"),
        (str(device), "<<DEVICE>>"),
    ]

    def tokenize(v):
        if not isinstance(v, str):
            return v
        out = v
        for lit, ph in repls:
            if lit:
                out = out.replace(lit, ph)
        return out

    return tokenize, prefix


def _lbl_idx(columns):
    return {c["label"]: c["index"] for c in columns if c["label"]}


def detect_command(r, lab):
    """Identifica se o sinal possui comando (output) e resume seu formato.

    Campos de comando descobertos nas TDTs reais:
      Direction=ReadWrite/Write, Commanding Mode, Output Data Type,
      Output Coordinates, Control Codes, Command Times.
    """
    def g(label):
        i = lab.get(label)
        return r[i] if i is not None and i < len(r) else None

    direction = g("Direction")
    out_dt = g("Output Data Type")
    ctrl = g("Control Codes")
    times = g("Command Times [s]")
    out_coord = g("Output Coordinates")
    mode = g("Commanding Mode")
    timeout = g("Command Timeout [s]")

    has_cmd = bool(
        (direction in ("Write", "ReadWrite"))
        or (out_dt not in (None, ""))
        or (ctrl not in (None, ""))
    )
    if not has_cmd:
        return None

    ctrl_list = [c for c in str(ctrl).split(";")] if ctrl not in (None, "") else []
    return {
        "outputDataType": out_dt,
        "controlCodes": ctrl if ctrl not in (None, "") else None,
        "commandTimes": times if times not in (None, "") else None,
        "commandingMode": mode if mode not in (None, "") else None,
        "commandTimeout": timeout if timeout not in (None, "") else None,
        "coordCount": max(1, len(ctrl_list)) if ctrl_list else 1,
        "templateCoord": out_coord if out_coord not in (None, "") else None,
    }


def extract_signals(ws, columns, alias, module, device, dictionary, klass_default):
    """Extrai sinais cujo Signal Name começa com o prefixo do equipamento."""
    rows = list(ws.iter_rows(values_only=True))
    tokenize, prefix = build_tokenizer(alias, module, device)
    lab = _lbl_idx(columns)
    name_col = 0
    signals = []
    seen = set()
    for ri in range(DATA_START, len(rows)):
        r = rows[ri]
        name = r[name_col]
        if not name:
            continue
        name = str(name)
        if not name.startswith(prefix + "_"):
            continue
        suffix = name[len(prefix) + 1:]
        if suffix in seen:
            continue
        seen.add(suffix)
        # tokeniza a linha inteira
        trow = [tokenize(c) for c in r]
        # garante comprimento alinhado às colunas
        while len(trow) < len(columns):
            trow.append(None)
        info = dictionary.get(suffix.upper(), {})
        # Signal Alias real (col 2) costuma ser a melhor descrição
        real_alias = r[2] if len(r) > 2 else None
        command = detect_command(r, lab) if klass_default == "discrete" else None
        signals.append({
            "suffix": suffix,
            "description": info.get("description") or (str(real_alias) if real_alias else suffix),
            "aliasLabel": str(real_alias) if real_alias else "",
            "klass": info.get("klass", klass_default),
            "measurementType": r[4] if len(r) > 4 else None,
            "signalSubType": r[14] if len(r) > 14 else None,
            "phases": r[15] if len(r) > 15 else None,
            "hasCommand": command is not None,
            "command": command,
            "row": trow,
        })
    return signals


def main():
    dictionary = load_dictionary()
    print(f"dicionário: {len(dictionary)} sufixos")

    wb_clean = openpyxl.load_workbook(CLEAN_TDT, read_only=True, data_only=True)
    wb_multi = openpyxl.load_workbook(MULTI_TDT, read_only=True, data_only=True)

    # cabeçalhos/colunas vêm do arquivo LIMPO (golden template, tem as 2 sheets)
    headers = {}
    columns = {}
    for sh in ["DNP3_DiscreteSignals", "DNP3_AnalogSignals"]:
        h, cols, _ = sheet_meta(wb_clean[sh])
        headers[sh] = h
        columns[sh] = cols

    device_types = []
    for dev in SOURCE_DEVICES:
        wb = wb_clean if dev["file"] == "CLEAN" else wb_multi
        sigs = {"discrete": [], "analog": []}
        for sh in dev["sheets"]:
            if sh not in wb.sheetnames:
                continue
            klass = "analog" if "Analog" in sh else "discrete"
            extracted = extract_signals(
                wb[sh], columns[sh], dev["alias"], dev["module"], dev["device"],
                dictionary, klass,
            )
            sigs[klass] = extracted
        device_types.append({
            "id": dev["id"],
            "label": dev["label"],
            "description": dev["description"],
            "source": f'{dev["alias"]}_{dev["module"]}_{dev["device"]}',
            "defaults": dev["defaults"],
            "signalCount": {
                "discrete": len(sigs["discrete"]),
                "analog": len(sigs["analog"]),
            },
            "signals": sigs,
        })
        print(f'{dev["label"]}: {len(sigs["discrete"])} digitais, {len(sigs["analog"])} analógicos')

    catalog = {
        "meta": {
            "headerRows": HEADER_ROWS,
            "referenceTemplate": "reference_template.xlsx",
            "infoSheet": "DMSMatchingTemplateInfo",
            "signalSheets": ["DNP3_DiscreteSignals", "DNP3_AnalogSignals"],
            "identityRule": "Signal name = {ALIAS}_{MODULE}_{DEVICE}_{SUFFIX}; suffix after last device token is the signal type.",
        },
        "headers": headers,
        "columns": columns,
        "deviceTypes": device_types,
    }

    out = DATA / "catalog.json"
    out.write_text(json.dumps(catalog, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"catálogo escrito em {out} ({out.stat().st_size//1024} KB)")


if __name__ == "__main__":
    main()
