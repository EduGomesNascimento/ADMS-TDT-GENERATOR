"""
build_tap.py — adiciona suporte ao tipo Discrete Analog (A/D), cujo único sinal é
o TAP (posição do comutador), que pertence ao Transformador (lado AT).

Faz 3 coisas:
1. Adiciona a sheet DNP3_DiscreteAnalog ao reference_template.xlsx (estrutura real
   de 48 colunas, estilos reaproveitados da AnalogSignals).
2. Registra as colunas/headers do DiscreteAnalog no catalog.json.
3. Insere o sinal TAP (tokenizado p/ o TR consolidado) no tipo Transformador.
"""
from __future__ import annotations
import json, re
from copy import copy
from pathlib import Path
import openpyxl
from build_catalog import _lbl_idx, detect_command, load_dictionary

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
CATALOG = DATA / "catalog.json"
TEMPLATE = DATA / "reference_template.xlsx"
RAW = DATA / "_tap_raw.json"
DA_SHEET = "DNP3_DiscreteAnalog"


def main():
    raw = json.loads(RAW.read_text(encoding="utf-8"))
    header = raw["header"]      # 4 linhas (section/table/code/label)
    tap = raw["tap"]
    ncol = len(header[3])

    # ---- 1. adicionar a sheet DNP3_DiscreteAnalog ao template ----
    wb = openpyxl.load_workbook(TEMPLATE)
    if DA_SHEET in wb.sheetnames:
        del wb[DA_SHEET]
    base = wb["DNP3_AnalogSignals"]
    da = wb.copy_worksheet(base)
    da.title = DA_SHEET
    # ajustar nº de colunas
    cur = da.max_column
    if cur > ncol:
        da.delete_cols(ncol + 1, cur - ncol)
    # sobrescrever as 4 linhas de cabeçalho com os rótulos reais do DA
    for ri in range(4):
        for ci in range(ncol):
            v = header[ri][ci] if ci < len(header[ri]) else None
            da.cell(row=ri + 1, column=ci + 1, value=(None if v in (None, "") else v))
    # limpar linhas de dados (mantém a 5ª p/ estilos)
    if da.max_row > 5:
        da.delete_rows(6, da.max_row - 5)
    wb.save(TEMPLATE)

    # ---- 2. registrar colunas no catálogo ----
    cat = json.loads(CATALOG.read_text(encoding="utf-8"))
    section, table, codes, labels = header
    columns = [{"index": ci,
                "section": section[ci] if ci < len(section) else None,
                "table": table[ci] if ci < len(table) else None,
                "code": codes[ci] if ci < len(codes) else None,
                "label": labels[ci] if ci < len(labels) else None}
               for ci in range(ncol)]
    cat["columns"][DA_SHEET] = columns
    cat["headers"][DA_SHEET] = header
    if DA_SHEET not in cat["meta"]["signalSheets"]:
        cat["meta"]["signalSheets"].append(DA_SHEET)

    # ---- 3. inserir TAP no Transformador (tokenizado p/ TR consolidado) ----
    dictionary = load_dictionary()
    alias, module, device = tap["alias"], tap["module"], tap["device"]
    n = re.match(r"TR(\d+)", module).group(1)

    def tok(v):
        if not isinstance(v, str):
            return v
        out = v.replace(alias, "<<ALIAS>>").replace(f"TR{n}", "TR<<N>>")
        return out

    trow = [tok(c) for c in tap["row"]]
    while len(trow) < ncol:
        trow.append(None)
    lab = _lbl_idx(columns)
    sig = {
        "suffix": "TAP", "code": "TAP", "group": "Comutador (A/D)",
        "description": dictionary.get("TAP", {}).get("description") or "Posição do comutador (TAP)",
        "aliasLabel": "TAP", "klass": "discrete_analog",
        "measurementType": tap["row"][lab.get("Measurement Type")],
        "signalSubType": tap["row"][lab.get("Signal Type")],
        "phases": tap["row"][lab.get("Phases")],
        "hasCommand": False, "command": None, "row": trow,
    }
    for d in cat["deviceTypes"]:
        if d["id"] == "transformador":
            d["signals"]["discrete_analog"] = [sig]
            d["signalCount"]["discrete_analog"] = 1

    CATALOG.write_text(json.dumps(cat, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"DA sheet adicionada ({ncol} cols); TAP inserido no Transformador.")
    print("sheets template:", openpyxl.load_workbook(TEMPLATE).sheetnames)


if __name__ == "__main__":
    main()
