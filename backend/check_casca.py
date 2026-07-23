"""
check_casca.py — conferencia independente dos 3 arquivos da CASCA.

Valida, sem reaproveitar a logica do gerador:
  1) lista CORRIGIDA: nenhuma linha de sinal sem index; por grupo DNP3 as
     coordenadas sao unicas e contiguas
  2) TDT: nomes unicos, Input Coordinates unicas por aba, nada vazio,
     Device Mapping preenchido, Remote Unit/AOR/container corretos
  3) TDT x lista: todo sinal Utilizado?=SIM esta na TDT com a MESMA coordenada

Uso: python check_casca.py
"""
from __future__ import annotations
import collections
from pathlib import Path
import openpyxl

D = Path("C:/Users/egnpo/Downloads")
LISTA = D / "RGE ADMS_Lista Pontos Casca_CORRIGIDA.xlsx"
TDT = D / "TDT_CASCA_UTR_CAS_3.xlsx"
SKIP = {"Informações", "RELACAO RELES", "MAPA DE REDE", "Lista"}
HR = 4
RU, AOR, CONT = "UTR_CAS_3", "CAS Trans", "Cas_Obra"
ALIAS = "CAS"


def _norm_alias(nome: str) -> str:
    """A lista tem 4 linhas com o alias de OUTRA SE (IMA_...). O gerador troca
    para CAS_ ao montar a TDT — aqui a conferencia faz o mesmo, senao acusa
    'ausente na TDT' de mentira."""
    return (nome if not nome or nome.startswith(f"{ALIAS}_")
            else "_".join([ALIAS] + nome.split("_")[1:]))

erros = []


def falha(msg):
    erros.append(msg)
    print(f"  ERRO: {msg}")


def _nums(v):
    out = []
    for x in str(v).split(";"):
        try:
            out.append(int(float(x.strip())))
        except (ValueError, TypeError):
            pass
    return out


def ler_lista():
    wb = openpyxl.load_workbook(LISTA, data_only=True)
    pts = []
    for sn in wb.sheetnames:
        if sn in SKIP:
            continue
        ws = wb[sn]
        hi = next((r for r in range(1, min(15, ws.max_row + 1))
                   if any(str(ws.cell(r, c).value or "").strip() == "SIGLA SINAL"
                          for c in range(1, ws.max_column + 1))), None)
        if hi is None:
            continue
        col = {str(ws.cell(hi, c).value or "").strip(): c
               for c in range(1, ws.max_column + 1)}
        if "INDEX DNP3" not in col:
            continue
        for r in range(hi + 1, ws.max_row + 1):
            gv = lambda k: ws.cell(r, col[k]).value if k in col else None
            tipo = str(gv("TIPO") or "").strip()
            if tipo not in ("D", "A", "C", "A/D"):
                continue
            pts.append({
                "sheet": sn, "linha": r, "tipo": tipo,
                "sigla": str(gv("SIGLA SINAL") or "").strip(),
                "nome": _norm_alias(str(gv("NOME") or "").strip()),
                "idx": gv("INDEX DNP3"),
                "desc": str(gv("DESCRIÇÃO DO PONTO") or "").strip(),
                "escala": gv("Escala"),
                "tipoPt": str(gv("Tipo") or "").strip(),
                "usado": str(gv("Utilizado?") or "").strip().upper() == "SIM",
            })
    return pts


def ler_tdt():
    wb = openpyxl.load_workbook(TDT, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        if not sn.startswith("DNP3_") or sn == "DNP3_RTUs":
            continue
        ws = wb[sn]
        hdr = {str(ws.cell(HR, c).value or "").strip(): c
               for c in range(1, ws.max_column + 1)}
        if "Signal Name" not in hdr:
            continue
        linhas = []
        for r in range(HR + 1, ws.max_row + 1):
            nome = ws.cell(r, hdr["Signal Name"]).value
            if not nome:
                continue
            linhas.append({k: ws.cell(r, c).value for k, c in hdr.items()})
        out[sn] = linhas
    # RTU
    ws = wb["DNP3_RTUs"]
    hdr = {str(ws.cell(HR, c).value or "").strip(): c
           for c in range(1, ws.max_column + 1)}
    out["_RTU"] = {k: ws.cell(HR + 1, c).value for k, c in hdr.items()}
    # abas de configuracao que NAO podem sobrar do esqueleto
    out["_CFG"] = {}
    for sn in ("DNP3_TCPLinks", "DNP3_UDPLinks", "DNP3_ScanGroups"):
        if sn not in wb.sheetnames:
            continue
        w = wb[sn]
        out["_CFG"][sn] = [
            [w.cell(r, c).value for c in range(1, w.max_column + 1)]
            for r in range(HR + 1, w.max_row + 1)
            if any(w.cell(r, c).value not in (None, "")
                   for c in range(1, w.max_column + 1))]
    return out


def main():
    print("=== 1) LISTA CORRIGIDA ===")
    pts = ler_lista()
    sem = [p for p in pts if not _nums(p["idx"])]
    print(f"  linhas de sinal: {len(pts)} | sem index: {len(sem)}")
    if sem:
        falha(f"{len(sem)} linhas sem index — ex.: "
              + "; ".join(f"{p['sheet']}!L{p['linha']}" for p in sem[:5]))
    por = collections.defaultdict(list)
    for p in pts:
        ns = _nums(p["idx"])
        # A/D (TAP) lê ponto analogico -> divide o espaco de indices com A
        g = "A" if p["tipo"] in ("A", "A/D") else p["tipo"]
        # comando usa n;n (mesma coord repetida) -> conta uma vez
        por[g] += sorted(set(ns)) if g == "C" else ns
    for t, v in sorted(por.items()):
        s = sorted(v)
        dup = [x for x, n in collections.Counter(s).items() if n > 1]
        cont = s == list(range(s[0], s[0] + len(s))) if s else True
        print(f"  {t:<4} {len(s):>5} coords  {s[0]}..{s[-1]}  "
              f"duplicadas={len(dup)}  contiguo={cont}")
        if dup:
            falha(f"grupo {t}: coordenadas duplicadas {dup[:10]}")
        if not cont:
            falha(f"grupo {t}: sequencia com buraco")

    print("=== 2) TDT ===")
    tdt = ler_tdt()
    rtu = tdt.pop("_RTU")
    cfg = tdt.pop("_CFG", {})
    for k, esperado in (("Remote Unit (Terminal Server) Name", RU),
                        ("Remote Unit AOR Group", AOR),
                        ("Container Name", CONT)):
        got = str(rtu.get(k) or "")
        print(f"  RTU {k}: {got!r}")
        if got != esperado:
            falha(f"RTU {k} = {got!r}, esperado {esperado!r}")
    # sem Container Custom ID o ADMS reprova a RTU e derruba TODOS os sinais
    ccid = str(rtu.get("Container Custom ID") or "")
    print(f"  RTU Container Custom ID: {ccid!r}")
    if not ccid:
        falha("RTU sem Container Custom ID — o ADMS responde 'Mandatory "
              "reference nao encontrada' e nenhum sinal entra")
    # nenhuma sobra do esqueleto da LVA nas abas de configuracao
    for sn, linhas in cfg.items():
        alheias = [l for l in linhas
                   if any(RU not in str(v or "") and "LVA" in str(v or "")
                          for v in l)]
        print(f"  {sn}: {len(linhas)} linha(s)")
        for l in alheias:
            falha(f"{sn}: linha de OUTRA subestacao ({l[0]!r}) — importar isso "
                  f"alteraria a UTR da LVA")
    nomes = collections.Counter()
    total = 0
    # AnalogSignals e DiscreteAnalog dividem o espaco de indices analogicos
    espaco = collections.defaultdict(collections.Counter)
    for sn, linhas in tdt.items():
        coords = espaco["D" if "Discrete" in sn and "Analog" not in sn else "A"]
        for x in linhas:
            total += 1
            nomes[str(x.get("Signal Name"))] += 1
            ic = x.get("Input Coordinates")
            if ic in (None, ""):
                falha(f"{sn}: {x.get('Signal Name')} sem Input Coordinates")
            else:
                for n in _nums(ic):
                    coords[n] += 1
            for campo in ("Device Mapping", "Remote Unit", "Signal AOR Group",
                          "Remote Point Custom ID"):
                if not str(x.get(campo) or "").strip():
                    falha(f"{sn}: {x.get('Signal Name')} sem {campo}")
            if str(x.get("Remote Unit") or "") != RU:
                falha(f"{sn}: {x.get('Signal Name')} Remote Unit != {RU}")
            if str(x.get("Signal AOR Group") or "") != AOR:
                falha(f"{sn}: {x.get('Signal Name')} AOR != {AOR}")
        print(f"  {sn:<24} {len(linhas):>5} sinais")
    for g, coords in espaco.items():
        dup = [c for c, n in coords.items() if n > 1]
        print(f"  espaco DNP3 {g}: {len(coords)} coords | duplicadas: {len(dup)}")
        if dup:
            falha(f"espaco {g}: Input Coordinates repetidas {sorted(dup)[:10]}")
    dupn = [n for n, c in nomes.items() if c > 1]
    print(f"  total {total} sinais | nomes duplicados: {len(dupn)}")
    if dupn:
        falha(f"nomes duplicados na TDT: {dupn[:10]}")

    print("=== 3) TDT x LISTA ===")
    porcoord = {}
    porsinal = {}
    for sn, linhas in tdt.items():
        for x in linhas:
            porcoord[str(x.get("Signal Name"))] = str(x.get("Input Coordinates"))
            porsinal[str(x.get("Signal Name"))] = x
    faltando, divergentes = [], []
    usados = set()          # cada sinal da TDT casa com UMA linha da lista
    for p in pts:
        if not p["usado"] or p["tipo"] == "C" or not p["sigla"]:
            continue
        # nome repetido: o 2º+ vira CAS_MOD_DEV-2_SIGLA na TDT
        pn = p["nome"].split("_")
        cands = [p["nome"]] + ["_".join(pn[:2] + [f"{pn[2]}-{k}"] + pn[3:])
                               for k in (2, 3, 4)] if len(pn) > 3 else [p["nome"]]
        # casa com o sinal cujo NOME é um dos candidatos E cuja coordenada bate;
        # so se nenhum bater é que aponta divergencia (independe da ordem em que
        # o gerador decidiu quem fica com o nome canonico)
        livres = [c for c in cands if c in porcoord and c not in usados]
        hit = next((c for c in livres
                    if porcoord[c].strip() == str(p["idx"]).strip()), None)
        if hit is not None:
            usados.add(hit)
            continue
        if livres:
            usados.add(livres[0])
            divergentes.append((p, porcoord[livres[0]]))
        else:
            faltando.append(p)
    print(f"  sinais SIM na lista: "
          f"{sum(1 for p in pts if p['usado'] and p['tipo'] != 'C' and p['sigla'])}")
    print(f"  ausentes na TDT: {len(faltando)} | coordenada divergente: {len(divergentes)}")
    for p in faltando[:8]:
        falha(f"ausente na TDT: {p['sheet']}!L{p['linha']} {p['nome']}")
    for p, c in divergentes[:8]:
        falha(f"coord divergente: {p['nome']} lista={p['idx']} tdt={c}")

    # mapa NOME-da-lista -> sinal da TDT (reaproveita o casamento acima)
    casado = {}
    usados2 = set()
    for p in pts:
        if not p["usado"] or p["tipo"] == "C" or not p["sigla"]:
            continue
        pn = p["nome"].split("_")
        cands = ([p["nome"]] + ["_".join(pn[:2] + [f"{pn[2]}-{k}"] + pn[3:])
                                for k in (2, 3, 4)]) if len(pn) > 3 else [p["nome"]]
        livres = [c for c in cands if c in porsinal and c not in usados2]
        hit = next((c for c in livres
                    if porcoord[c].strip() == str(p["idx"]).strip()), livres[0] if livres else None)
        if hit:
            usados2.add(hit)
            casado[(p["sheet"], p["linha"])] = hit

    print("=== 4) COMANDOS (linhas C -> Output Coordinates) ===")
    # comando da lista: linha C com o MESMO NOME de uma linha D
    cmds = {}
    for p in pts:
        if p["tipo"] == "C" and p["usado"] and p["sigla"]:
            cmds.setdefault(p["nome"], p)
    ok = errado = 0
    por_mod_sigla = {}
    for p in pts:
        if p["usado"] and p["tipo"] == "D":
            pn = p["nome"].split("_")
            por_mod_sigla.setdefault((pn[1] if len(pn) > 1 else "", p["sigla"]), p)
    portador = {}          # NOME do comando -> sinal da TDT que o carrega
    for nome, c in cmds.items():
        alvo = next((p for p in pts
                     if p["usado"] and p["tipo"] == "D" and p["nome"] == nome), None)
        if alvo is not None:
            sig_nome = casado.get((alvo["sheet"], alvo["linha"]))
        else:
            # 2) comando no disjuntor, status no modulo (CAS_LT1_52-1_25IE)
            cn = nome.split("_")
            alt = por_mod_sigla.get((cn[1] if len(cn) > 1 else "", c["sigla"]))
            sig_nome = casado.get((alt["sheet"], alt["linha"])) if alt else None
            # 3) comando puro: sinal proprio criado com o mesmo nome
            if sig_nome is None and nome in porsinal:
                sig_nome = nome
        if sig_nome is None:
            errado += 1
            falha(f"comando sem portador na TDT: {nome} ({c['sheet']}!L{c['linha']})")
            continue
        portador[nome] = sig_nome
        got = str(porsinal[sig_nome].get("Output Coordinates") or "").strip()
        esperado = str(c["idx"]).strip()
        if got == esperado:
            ok += 1
        else:
            errado += 1
            falha(f"Output de {nome} (em {sig_nome}): TDT={got!r} lista={esperado!r}")
    print(f"  comandos na lista: {len(cmds)} | corretos: {ok} | problemas: {errado}")
    # o inverso: sinal com Output que nao corresponde a nenhuma linha C
    carregam = set(portador.values())
    sobra = [n for n, x in porsinal.items()
             if str(x.get("Output Coordinates") or "").strip() and n not in carregam]
    print(f"  sinais com Output sem comando na lista: {len(sobra)}")
    for n in sobra[:8]:
        falha(f"tem Output mas nao ha linha C na lista: {n}")
    # sinais criados so para carregar comando puro (Input de preenchimento)
    filler = sorted(n for n, x in porsinal.items() if _nums(x.get("Input Coordinates"))
                    and min(_nums(x.get("Input Coordinates"))) >= 9000)
    print(f"  sinais com Input de preenchimento (>=9000): {len(filler)}")
    for n in filler:
        print(f"     {n} -> in={porsinal[n].get('Input Coordinates')} "
              f"out={porsinal[n].get('Output Coordinates')}")

    print("=== 5) DESCRICAO / ESCALA / TIPO DE PONTO ===")
    dif_desc = dif_esc = dif_multi = 0
    for p in pts:
        k = casado.get((p["sheet"], p["linha"]))
        if not k:
            continue
        x = porsinal[k]
        if p["desc"] and str(x.get("Signal Alias") or "").strip() != p["desc"]:
            dif_desc += 1
            if dif_desc <= 5:
                falha(f"descricao: {k} TDT={x.get('Signal Alias')!r} lista={p['desc']!r}")
        if p["tipo"] in ("A", "A/D") and p["escala"] not in (None, "", "-"):
            if str(x.get("Scaling Factor") or "").strip() != str(p["escala"]).strip():
                dif_esc += 1
                if dif_esc <= 5:
                    falha(f"escala: {k} TDT={x.get('Scaling Factor')!r} lista={p['escala']!r}")
        # Multi Coord na lista -> 2 coordenadas na TDT
        if p["tipoPt"].upper().startswith("MULTI"):
            if len(_nums(x.get("Input Coordinates"))) != 2:
                dif_multi += 1
                falha(f"Multi Coord com 1 coordenada: {k} = {x.get('Input Coordinates')!r}")
    print(f"  descricoes diferentes: {dif_desc} | escalas diferentes: {dif_esc} | "
          f"Multi Coord errados: {dif_multi}")

    print("=== 6) SINAL NA TDT QUE NAO EXISTE NA LISTA ===")
    legitimos = set(casado.values()) | set(portador.values())
    orfaos = [n for n in porsinal if n not in legitimos]
    print(f"  sinais orfaos: {len(orfaos)}")
    for n in orfaos[:8]:
        falha(f"na TDT mas nao casou com a lista: {n}")

    print("=== 7) CONTAGEM POR ABA ===")
    esperado = collections.Counter(
        p["sheet"] for p in pts if p["usado"] and p["tipo"] != "C" and p["sigla"])
    obtido = collections.Counter(
        p["sheet"] for p in pts if casado.get((p["sheet"], p["linha"])))
    for sn in sorted(esperado):
        e, o = esperado[sn], obtido[sn]
        marca = "ok" if e == o else "<<< DIFERENTE"
        print(f"  {sn:<24} lista={e:>4}  tdt={o:>4}  {marca}")
        if e != o:
            falha(f"aba {sn}: lista tem {e} sinais, TDT tem {o}")

    print()
    print("OK — nenhum problema encontrado" if not erros
          else f"{len(erros)} PROBLEMA(S) encontrado(s)")


if __name__ == "__main__":
    main()
