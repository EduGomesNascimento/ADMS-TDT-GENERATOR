"""
make_casca.py — TDT DNP3 da SE CASCA (UTR nova UTR_CAS_3, do zero).

Fonte da verdade: "RGE ADMS_Lista Pontos Casca.xlsx". Cada aba de módulo traz
Utilizado?/TIPO/SIGLA/NOME/Escala/Níveis Lógicos/Control Code/INDEX DNP3.
Um sinal comandável aparece em 2 linhas com o MESMO NOME: a `D` dá o índice de
ENTRADA, a `C` o de COMANDO. `A` = analógico, `A/D` = TAP (DiscreteAnalog).

As UTRs atuais (IEC104 UTR_CAS_1 / religadores RAD_*) são só referência — a nova
não interage com elas.

REALOCAÇÃO: a lista tem índices duplicados (duas cadeias de alocação que se
sobrepõem + duplicatas dentro do módulo). Mantemos o 1º ocupante de cada
coordenada e realocamos os demais para a 1ª coordenada livre acima do topo em
uso. Toda realocação é registrada no relatório.

Saídas:
  TDT_CASCA_UTR_CAS_3.xlsx       — a TDT
  CASCA_RELATORIO.xlsx           — inconsistências + de-para das realocações

Uso: python make_casca.py
"""
from __future__ import annotations
import io, json, re
from collections import defaultdict, OrderedDict, Counter
from copy import copy
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
import excel_native
import tdt_engine as E
import casca_devmap as devmap

LISTA = Path("C:/Users/egnpo/Downloads/RGE ADMS_Lista Pontos Casca.xlsx")
SKEL = Path("C:/Users/egnpo/Downloads/TDT_LVA_AL24.xlsx")   # esqueleto c/ DNP3_RTUs
OUT_TDT = Path("C:/Users/egnpo/Downloads/TDT_CASCA_UTR_CAS_3.xlsx")
OUT_REL = Path("C:/Users/egnpo/Downloads/CASCA_RELATORIO.xlsx")
DATA = Path(__file__).parent / "data"

# DE-PARA das siglas da lista sem template na base — usa a linha-molde da
# variante de PERFIL TÉCNICO equivalente (o NOME e a SIGLA da lista são
# preservados, e o Signal Alias vem da DESCRIÇÃO da lista; do molde só vem a
# configuração: Signal Type / Input Data Type / Message Mapping).
#   FGOO 'FALHA GOOSE'                   -> FCOM 'FALHA COMUNICACAO'  (Custom, NORMAL@FALHA)
#   TOC  '26 - ALARME TEMP OLEO CDC'     -> TOA  '26 ALARME'          (Custom, NORMAL@ATUADO)
#   81A  '81 - FUNCAO RECOMPOSICAO ERAC' -> 81   '81 FUNCAO'          (Enabled, comandável)
#   81P/81F/81C (estados do ERAC)        -> FALH 'FALHA GERAL'        (Custom status)
# 81E1 NÃO serve p/ 81P/81F/81C: é RelayTrip, e RelayTrip apontando p/ device
# físico é rejeitado pelo validador do ADMS (visto na SND).
FALLBACK_SIGLA = {
    "FGOO": "FCOM", "TOC": "TOA",
    "81A": "81", "81P": "FALH", "81F": "FALH", "81C": "FALH",
    "86RM": "86",          # 'REARME 86' — só existe como comando; molde = 86 BLOQUEIO
}

# Índice de ENTRADA de preenchimento para pontos que só existem como COMANDO
# (o validador do ADMS exige Input Coordinates em todo sinal discreto).
# Convenção do usuário: "se for impossível achar index, bote números BEEEM
# altos como 9599 ou 9999".
FILLER_BASE = 9599

RU = "UTR_CAS_3"
AOR = "CAS Trans"
FABRICANTE = "ELIPSE"
CONTAINER = "Casca_Obra"        # container da RTU nova (informado pelo usuário)
HEADER_ROWS = 4
SKIP_SHEETS = {"Informações", "RELACAO RELES", "MAPA DE REDE", "Lista"}


# ─── leitura da lista ────────────────────────────────────────────────────────
def read_lista():
    wb = openpyxl.load_workbook(LISTA, read_only=True, data_only=True)
    pts = []
    for sn in wb.sheetnames:
        if sn in SKIP_SHEETS:
            continue
        rows = list(wb[sn].iter_rows(values_only=True))
        hi = next((i for i, r in enumerate(rows[:14])
                   if r and any(str(c or "").strip() == "SIGLA SINAL" for c in r)), None)
        if hi is None:
            continue
        hdr = [str(c or "").strip() for c in rows[hi]]
        ix = {n: i for i, n in enumerate(hdr) if n}
        def g(r, k):
            i = ix.get(k)
            return r[i] if i is not None and i < len(r) else None
        for ri, r in enumerate(rows[hi + 1:], hi + 2):
            if not r:
                continue
            nome = str(g(r, "NOME") or "").strip()
            tipo = str(g(r, "TIPO") or "").strip()
            sigla = str(g(r, "SIGLA SINAL") or "").strip()
            idxv = str(g(r, "INDEX DNP3") or "").strip()
            # TODA linha que ocupa endereço recebe coordenada — inclusive as
            # Utilizado?=NÃO e as RESERVAS (sem sigla, mas com índice na lista
            # original). O usuário pediu "nada de pontas soltas na lista".
            # Só o que entra na TDT é filtrado depois, por p["usado"].
            if tipo not in ("D", "A", "C", "A/D") or nome == "NOME":
                continue
            reserva = not sigla or not nome
            if reserva and not nome and not _nums(idxv):
                continue                      # linha vazia de verdade: ignora
            pts.append({
                "sheet": sn, "linha": ri,
                "usado": (not reserva
                          and str(g(r, "Utilizado?") or "").strip().upper() == "SIM"),
                "reserva": reserva,
                "tipo": tipo,
                "sigla": str(g(r, "SIGLA SINAL") or "").strip(),
                "nome": nome,
                "idx": str(g(r, "INDEX DNP3") or "").strip(),
                "tipoPt": str(g(r, "Tipo") or "").strip(),
                "n0": str(g(r, "Nível Lógico 0") or "").strip(),
                "n1": str(g(r, "Nível Lógico 1") or "").strip(),
                "escala": g(r, "Escala"),
                "cc": str(g(r, "Control Code / Qualificador") or "").strip(),
                "desc": str(g(r, "DESCRIÇÃO DO PONTO") or "").strip(),
            })
    wb.close()
    return pts


def _nums(idx: str):
    """Coordenadas de um INDEX DNP3 ('97', '97;98', e floats do Excel '97.0')."""
    out = []
    for x in str(idx).split(";"):
        x = x.strip()
        if not x:
            continue
        try:
            out.append(int(float(x)))
        except (ValueError, TypeError):
            pass
    return out


# ─── diagnóstico das duplicatas da lista ─────────────────────────────────────
def diagnosticar(pts):
    """Evidencia as coordenadas repetidas na lista ORIGINAL (célula a célula)."""
    ocup = defaultdict(list)
    for p in pts:
        g = "A" if p["tipo"] in ("A", "A/D") else p["tipo"]   # mesmo espaço DNP3
        for n in _nums(p["idx"]):
            ocup[(g, n)].append(p)
    dups = []
    for (g, n), lst in sorted(ocup.items(), key=lambda x: (x[0][0], x[0][1])):
        if len(lst) > 1:
            abas = {p["sheet"] for p in lst}
            for p in lst:
                dups.append({**p, "grupo": g, "coord": n,
                             "escopo": "ENTRE MODULOS" if len(abas) > 1 else "DENTRO DO MODULO",
                             "conflita_com": ", ".join(
                                 f'{q["sheet"]}!N{q["linha"]} {q["nome"]}'
                                 for q in lst if q is not p)})
    semidx = [{**p, "grupo": p["tipo"]} for p in pts if not _nums(p["idx"])]
    return dups, semidx


# ─── re-sequenciamento contínuo das coordenadas ──────────────────────────────
def sequenciar(pts):
    """Re-sequencia TODAS as coordenadas em ordem, sem buracos nem repetição.

    Ordena pelo índice ORIGINAL (preserva o arranjo pretendido) e distribui
    sequencialmente: onde havia 490,490,491 passa a 490,491,492. Cada módulo
    continua de onde o anterior parou. Pontos de posição (Multi Coord) consomem
    2 coordenadas; comando usa a mesma repetida (n;n). Pontos com INDEX inválido
    (#REF!) entram no fim da fila do seu grupo.

    GRUPOS DNP3 (espaços de índice independentes):
      D            binary input
      A + A/D      analog input — o TAP (DiscreteAnalog) lê um ponto ANALÓGICO,
                   então divide a numeração com os analógicos comuns. Na lista
                   original isso causava colisão (TAP do TR6 = 70 = IA do TR7AT).
      C            binary output (comandos)
    """
    final, mapa = {}, []
    for g, tipos in (("D", ("D",)), ("A", ("A", "A/D")), ("C", ("C",))):
        lst = [p for p in pts if p["tipo"] in tipos]
        if not lst:
            continue
        # ordena pelo índice original; sem índice (#REF!) vai pro fim
        def chave(p):
            ns = _nums(p["idx"])
            return (0, ns[0], p["linha"]) if ns else (1, 0, p["linha"])
        lst = sorted(lst, key=chave)
        prox = min((n for p in lst for n in _nums(p["idx"])), default=0)
        for p in lst:
            multi = str(p.get("tipoPt", "")).upper().startswith("MULTI")
            if g == "C":
                val = f"{prox};{prox}"; prox += 1
            elif multi:
                val = f"{prox};{prox + 1}"; prox += 2
            else:
                val = str(prox); prox += 1
            final[(p["sheet"], p["linha"])] = val
            mapa.append({**p, "grupo": g, "de": p["idx"], "para": val,
                         "mudou": "SIM" if str(p["idx"]).strip() != val else "nao"})
    return final, mapa


# ─── relatório ───────────────────────────────────────────────────────────────
def gerar_relatorio(pts, mapa, dups, semidx, sem_tpl, nomes_dup, renomeados=(),
                    fallback_rows=(), limpos=(), dm=None, cmds=None):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    bold = Font(bold=True); hdrfill = PatternFill("solid", fgColor="DDEBF7")
    warn = PatternFill("solid", fgColor="FFF2CC")

    def sheet(title, cols, rows, fills=None):
        ws = wb.create_sheet(title[:31])
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            ws.cell(1, c).font = bold; ws.cell(1, c).fill = hdrfill
        for r in rows:
            ws.append(r)
        for i, w in enumerate([14, 34, 10, 12, 14, 14, 40], 1):
            if i <= len(cols):
                ws.column_dimensions[chr(64 + i)].width = w
        return ws

    mudou = [m for m in mapa if m["mudou"] == "SIM"]
    sheet("0-LEIA-ME",
          ["O PROBLEMA E A SOLUCAO"],
          [["PROBLEMA ENCONTRADO NA LISTA DE PONTOS"],
           ["1) Coordenadas REPETIDAS: o mesmo INDEX DNP3 aparece em 2+ sinais."],
           ["   - ENTRE MODULOS: duas cadeias de alocacao se sobrepoem."],
           ["     ex.: AL 13!N66 (CAS_AL13_52-13_SGFT) e INTERBARRAS BT!N41"],
           ["          (CAS_IB20_52-20_FA) usam os DOIS a coordenada 556."],
           ["   - DENTRO DO MODULO: no TR 2, a partir da linha 79, o indice sobe"],
           ["     de 2 em 2 (L79 MOLA=416 e L80 SF6A=416) — formula arrastada."],
           ["2) Coordenadas INVALIDAS: 172 pontos com #REF! (formula quebrada)."],
           [""],
           ["POR QUE ISSO QUEBRA: no DNP3 cada ponto de um grupo (entrada binaria,"],
           ["entrada analogica, saida) precisa de um indice UNICO na UTR. Coordenada"],
           ["repetida = o ADMS rejeita ('does not have unique input coordinates')."],
           [""],
           ["COMO FOI CORRIGIDO"],
           ["Re-sequenciamento CONTINUO: as coordenadas foram redistribuidas em"],
           ["ordem, sem buracos e sem repeticao, preservando o arranjo original"],
           ["(ordenado pelo indice antigo). Onde havia 490,490,491 passou a ser"],
           ["490,491,492. Cada modulo continua de onde o anterior parou."],
           ["Ponto de posicao (Multi Coord) consome 2 coordenadas; comando usa a"],
           ["mesma repetida (n;n). Os #REF! entraram no fim da fila do seu grupo."],
           [""],
           ["O QUE VOCE RECEBE"],
           ["- TDT_CASCA_UTR_CAS_3.xlsx ....... a TDT ja com as coordenadas novas"],
           ["- RGE ADMS_Lista Pontos Casca_CORRIGIDA.xlsx ... a lista com a coluna"],
           ["  INDEX DNP3 ja arrumada (mesma estrutura da original)"],
           ["- Aba 2 deste relatorio .......... de-para completo (antes -> depois)"],
           ["- Aba 3 .......................... evidencia das coordenadas repetidas"],
           [""],
           ["SEM PONTAS SOLTAS"],
           ["TODA linha de sinal recebe coordenada — inclusive as marcadas"],
           ["Utilizado? = NAO. Elas nao entram na TDT (nao sao sinais ativos),"],
           ["mas RESERVAM o endereco, entao a numeracao nao muda quando forem"],
           ["ativadas. Nenhum sinal fica sem index."],
           ["NOME repetido nao e mais descartado: o 2o sinal recebe sufixo no"],
           ["device (CAS_TR7_TR7-2_TOD) e entra na TDT — aba 6."],
           [""],
           ["ATENCAO — ERRO DE ORIGEM NAS ABAS 'BC 1' E 'BC 2'"],
           ["Nessas abas a celula MODULO esta escrita 'AL' (e nao 'BC'), entao a"],
           ["formula do NOME gera CAS_AL18_... e CAS_AL28_... . O AL18 colide com"],
           ["a aba TRANSFERENCIA 24-01, que e o alimentador 18 de verdade. Aqui o"],
           ["nome limpo ficou com quem tinha indice na lista (TRANSFERENCIA) e o"],
           ["do BC 1 recebeu sufixo. CORRIJA A CELULA NA PLANILHA DE ORIGEM se o"],
           ["banco de capacitores deve se chamar BC18/BC28 no modelo."],
           [""],
           ["DEVICE MAPPING — conferido contra o MODELO"],
           ["Fonte da verdade: PT-MOD-SE-CASCA.xml (changeset do Casca_Obra)."],
           ["Cada dispositivo de la tem um 'ID de Mapeamento SCADA'; a coluna"],
           ["Device Mapping da TDT precisa conter EXATAMENTE esse texto, senao o"],
           ["ADMS responde 'Could not find any device that corresponds to...'."],
           [""],
           ["PROBLEMA: o Casca_Obra e um clone da CASCA ATUAL. Ele tem"],
           ["  AL12 AL13 AL14 AL15 AL21 · LTSCO LTPRI LTMRU · TR1 TR2 (+AT/BT)"],
           ["  B138 BP23 TSA3"],
           ["e a lista NOVA pede tambem"],
           ["  AL24 AL25 AL26 · BC1(AL18) BC2(AL28) · LT1 LT2 LT3 · TR6 TR7"],
           ["  IB20 · 24-1(AL18) 24-2(TRF29) · BP69 BP113.8 BP213.8 · TSA1 TSA2"],
           ["Esses dispositivos AINDA NAO EXISTEM no diagrama — nenhuma TDT"],
           ["resolve isso, tem que ser criado no modelo. A aba 13 lista um por um."],
           [""],
           ["O QUE FOI RESOLVIDO AUTOMATICAMENTE"],
           ["  1) texto igual ao do modelo -> usado direto"],
           ["  2) so o numero do equipamento mudou (AL12 usa 52-2 no modelo e"],
           ["     52-12 na lista) -> vale o texto do modelo"],
           ["  3) rele especifico inexistente -> rele generico _PROT do vao"],
           ["  4) ainda sem alternativa -> DISJUNTOR (instrucao do usuario)"],
           ["Auditoria: aba 10 (sinal a sinal), 11 (resumo), 14 (rebaixados)."],
           [""],
           ["A LISTA CORRIGIDA ABRE SEM REPARO"],
           ["As formulas foram congeladas no valor calculado e o vinculo externo"],
           ["da planilha original foi removido — era ele que fazia o Excel pedir"],
           ["reparo. Indices antigos em linhas nao utilizadas: ver aba 9."],
           [""],
           ["ATENCAO: as coordenadas tem que bater com o que for configurado na UTR"],
           ["ELIPSE. Use a lista CORRIGIDA como referencia para parametrizar a UTR."]])

    sheet("1-Resumo",
          ["Item", "Qtde", "Observação"],
          [["Pontos na lista (TODOS)", len(pts), "toda linha com SIGLA+NOME+TIPO"],
           ["  .. Utilizado? = SIM", sum(1 for p in pts if p.get("usado")), "entram na TDT"],
           ["  .. Utilizado? = NAO", sum(1 for p in pts if not p.get("usado")),
            "NAO entram na TDT, mas RESERVAM coordenada"],
           ["Coordenadas RE-SEQUENCIADAS", len(mapa), "todas — ver aba 2 (de-para)"],
           ["  .. que MUDARAM de valor", len(mudou), "as demais ficaram iguais"],
           ["Coordenadas REPETIDAS na lista", len(dups), "ver aba 3 — evidencia"],
           ["Pontos com INDEX invalido (#REF!)", len(semidx), "ver aba 4"],
           ["Siglas SEM template na base", len(sem_tpl), "0 = todas resolvidas"],
           ["NOMES duplicados (renomeados)", len(renomeados), "ver aba 6 — todos na TDT"],
           ["Index antigo limpo (nao utilizado)", len(limpos), "ver aba 9"],
           ["UTR", RU, f"nova, DNP3, {FABRICANTE}, AOR {AOR}"],
           ["Container da RTU", CONTAINER, "informado pelo usuario"]])

    sheet("2-DePara coordenadas",
          ["Aba", "Linha", "Tipo", "SIGLA", "NOME", "Index ORIGINAL", "Index NOVO", "Mudou?"],
          [[m["sheet"], m["linha"], m["grupo"], m["sigla"], m["nome"], m["de"], m["para"],
            m["mudou"]] for m in mapa])

    sheet("3-Coords REPETIDAS (evidencia)",
          ["Escopo", "Tipo", "Coordenada", "Aba", "Linha", "NOME", "Conflita com"],
          [[d["escopo"], d["grupo"], d["coord"], d["sheet"], d["linha"], d["nome"],
            d["conflita_com"]] for d in dups])

    sheet("4-INDEX invalido (#REF!)",
          ["Aba", "Linha", "Tipo", "SIGLA", "NOME", "Valor na lista"],
          [[x["sheet"], x["linha"], x["grupo"], x["sigla"], x["nome"], x["idx"]]
           for x in semidx])

    sheet("5-Siglas sem template",
          ["SIGLA", "Qtde pontos", "Abas", "NOMEs (exemplos)"],
          [[s, d["n"], ", ".join(sorted(d["sheets"])[:5]), ", ".join(d["nomes"][:3])]
           for s, d in sorted(sem_tpl.items(), key=lambda x: -x[1]["n"])])

    sheet("7-Nomes duplicados",
          ["NOME", "Tipo", "Ocorrencias", "Abas"],
          nomes_dup)

    sheet("6-Renomeados (nome dup)",
          ["Aba", "Linha", "Tipo", "SIGLA", "NOME na lista", "NOME na TDT", "Descricao"],
          [[d["sheet"], d["linha"], d["tipo"], d["sigla"], d["nome"], d["novo"],
            d.get("desc", "")] for d in renomeados])

    sheet("8-Siglas por equivalencia",
          ["SIGLA da lista", "Molde usado", "Qtde", "Descricao na lista"],
          fallback_rows)

    sheet("9-Index antigo limpo",
          ["Aba", "Linha", "Tipo", "SIGLA", "Utilizado?", "Index ANTIGO removido"],
          list(limpos))

    cmds = cmds or {"realoc": [], "orfaos": []}
    sheet("12-Comandos resolvidos",
          ["Situacao", "Aba", "Linha", "SIGLA", "NOME do comando",
           "Sinal que recebeu o comando", "Coord de comando", "Input usado"],
          [["comando no DJ, status no modulo", c["sheet"], c["linha"], c["sigla"],
            c["nome"], c["alvo"], c["coord"], "(do proprio status)"]
           for c in cmds["realoc"]]
          + [["COMANDO PURO: sinal criado", c["sheet"], c["linha"], c["sigla"],
              c["nome"], c["nome"], c["coord"],
              f"{c.get('input', '')} (preenchimento)"] for c in cmds["orfaos"]])

    dm = dm or {"linhas": [], "origem": Counter(), "pendentes": []}
    sheet("10-Device Mapping",
          ["Aba", "Linha", "Tipo", "SIGLA", "NOME na TDT", "Device Mapping",
           "Origem da regra", "Situacao no modelo"],
          dm["linhas"])

    # o que precisa ser CRIADO no Casca_Obra para os sinais mapearem
    falta = OrderedDict()
    for x in dm.get("pendentes", []):
        if not x["pend"].startswith("MODULO"):
            continue
        d = falta.setdefault(x["dm"], {"n": 0, "mod": x["mod"], "dev": x["dev"],
                                       "ex": []})
        d["n"] += 1
        if len(d["ex"]) < 3:
            d["ex"].append(x["nome"])
    tipo_disp = {"DJ": "Disjuntor (BREAKER)", "SEC": "Seccionadora (DISCONNECTOR)",
                 "TC": "Transformador de corrente (CURRENTTR)",
                 "TP": "Transformador de potencial (POTENTIALTR)",
                 "TR": "Transformador de potencia (POWERTR)",
                 "COMTAP": "Comutador de tap", "TAP_REG": "Regulador (REGCTRL)",
                 "BP": "Barra (BUSBAR)", "RET": "Retificador"}
    sheet("13-CRIAR no Casca_Obra",
          ["Device Mapping (ID de Mapeamento SCADA)", "Modulo", "Equipamento",
           "Tipo de dispositivo", "Sinais que dependem", "Exemplos de sinal"],
          [[k, v["mod"], v["dev"],
            tipo_disp.get(k.split("_")[-1],
                          "Rele de protecao (PROTECTEQP)" if "_PROT" in k else "?"),
            v["n"], ", ".join(v["ex"])]
           for k, v in sorted(falta.items(), key=lambda x: -x[1]["n"])])

    sheet("14-DM rebaixado",
          ["NOME na TDT", "SIGLA", "Device Mapping usado", "Motivo"],
          [[x["nome"], x["sigla"], x["dm"], x["pend"]]
           for x in dm.get("pendentes", []) if not x["pend"].startswith("MODULO")])
    sheet("11-DM origem (resumo)",
          ["Origem", "Qtde", "O que significa"],
          [[o, n, {"TDT atual": "sigla existe na TDT atual da CASCA — regra copiada",
                   "TDT atual (base)": "pickup/alarme: seguiu a sigla base da TDT atual",
                   "familia": "sigla da mesma familia (medida/trafo/secc/religador)",
                   "pickup/alarme": "P_/A_ de codigo ANSI -> {P|A}_PROT_{codigo}",
                   "medida": "grandeza eletrica -> TC (corrente/potencia) ou TP (tensao)",
                   "ANSI": "codigo ANSI -> rele do proprio device (PROT_)",
                   "device": "pelo equipamento: 52-*=DJ, 89-*/29-*=SEC",
                   "default": "sem regra: vai pro DISJUNTOR (instrucao do usuario)"}
                  .get(o, "")]
           for o, n in dm["origem"].most_common()])

    buf = io.BytesIO(); wb.save(buf)
    OUT_REL.write_bytes(buf.getvalue())


# ─── lista de pontos corrigida ───────────────────────────────────────────────
def gerar_lista_corrigida(mapa):
    """Copia a lista original e reescreve a coluna INDEX DNP3 com as coords novas.
    Casa por POSIÇÃO (aba, linha) — a coluna NOME é fórmula, então casar por nome
    falharia ao abrir o arquivo preservando fórmulas.

    A lista original tem VÍNCULOS EXTERNOS (xl/externalLinks/externalLink1.xml).
    O openpyxl reescreve essa parte sem o cache de valores → o Excel abre pedindo
    reparo. Como a saída é para consulta humana, congelamos toda fórmula no seu
    valor calculado e removemos os vínculos externos."""
    porpos = {(m["sheet"], m["linha"]): m["para"] for m in mapa}
    wb = openpyxl.load_workbook(LISTA)
    wbv = openpyxl.load_workbook(LISTA, data_only=True)   # valores em cache
    n = 0
    limpos = []
    for sn in wb.sheetnames:
        if sn in SKIP_SHEETS:
            continue
        ws = wb[sn]
        hi = None
        for r in range(1, min(15, ws.max_row + 1)):
            if any(str(ws.cell(r, c).value or "").strip() == "SIGLA SINAL"
                   for c in range(1, ws.max_column + 1)):
                hi = r; break
        if hi is None:
            continue
        col = {str(ws.cell(hi, c).value or "").strip(): c
               for c in range(1, ws.max_column + 1)}
        cI = col.get("INDEX DNP3")
        if not cI:
            continue
        cU, cT, cS = col.get("Utilizado?"), col.get("TIPO"), col.get("SIGLA SINAL")
        wsv = wbv[sn]
        for r in range(hi + 1, ws.max_row + 1):
            v = porpos.get((sn, r))
            if v is not None:
                ws.cell(r, cI).value = v
                n += 1
                continue
            # linha NÃO utilizada que ainda carrega o índice ANTIGO → limpa,
            # senão colide com a numeração nova e induz erro na parametrização
            old = wsv.cell(r, cI).value
            if old in (None, "") or not _nums(old):
                continue
            limpos.append([sn, r,
                           str(wsv.cell(r, cT).value or "").strip() if cT else "",
                           str(wsv.cell(r, cS).value or "").strip() if cS else "",
                           str(wsv.cell(r, cU).value or "").strip() if cU else "",
                           str(old)])
            ws.cell(r, cI).value = None
    # congela fórmulas no valor calculado (inclui a coluna NOME) e corta os
    # vínculos externos + nomes definidos que apontam para outra pasta ('[')
    frozen = 0
    for sn in wb.sheetnames:
        ws, wsv = wb[sn], wbv[sn]
        for row in ws.iter_rows():
            for cel in row:
                if isinstance(cel.value, str) and cel.value.startswith("="):
                    cel.value = wsv.cell(cel.row, cel.column).value
                    frozen += 1
    wb._external_links = []
    for dn in [k for k, v in wb.defined_names.items() if "[" in str(v.value)]:
        del wb.defined_names[dn]

    out = LISTA.with_name(LISTA.stem + "_CORRIGIDA.xlsx")
    buf = io.BytesIO(); wb.save(buf)
    data = buf.getvalue()
    try:
        data = excel_native.resave_native(data)          # grava formato nativo
    except Exception as e:                               # noqa: BLE001
        print(f"  (resave nativo indisponivel: {e})")
    try:
        out.write_bytes(data)
    except PermissionError:                              # aberta no Excel
        out = out.with_name(out.stem + "_NOVA.xlsx")
        out.write_bytes(data)
    print(f"lista corrigida: {out.name} ({n} coordenadas escritas, "
          f"{len(limpos)} indices antigos limpos em linhas nao utilizadas, "
          f"{frozen} formulas congeladas, vinculos externos removidos)")
    return limpos


# ─── TDT ─────────────────────────────────────────────────────────────────────
def main():
    pts = read_lista()
    n_sim = sum(1 for p in pts if p["usado"])
    print(f"lista: {len(pts)} pontos ({n_sim} Utilizado?=SIM, {len(pts) - n_sim} NAO) "
          f"— TODOS recebem coordenada")

    idx = json.loads((DATA / "sigla_index.json").read_text(encoding="utf-8"))
    TPL = {"D": idx["DNP3_DiscreteSignals"], "A": idx["DNP3_AnalogSignals"],
           "A/D": idx.get("DNP3_DiscreteAnalog", {}), "C": idx["DNP3_DiscreteSignals"]}

    # NOMES duplicados: o ADMS exige nome único. Em vez de descartar, o 2º+ sinal
    # ganha um sufixo no DEVICE do NOME (CAS_TR7_TR7-2_TOD) — o Device Mapping
    # continua apontando pro device REAL, então nada se perde na TDT.
    # Quem fica com o nome canônico: a linha que TEM índice válido na lista
    # original (a que o autor de fato endereçou). Ex.: CAS_AL18_AL18_IA existe
    # em "BC 1" (com #REF!, módulo escrito errado como AL) e em
    # "TRANSFERENCIA 24-01" (com índice) — o nome limpo fica com a segunda.
    vistos = {}; renomeados = []; nome_tdt = {}
    ordem = sorted((p for p in pts if p["usado"] and p["tipo"] != "C"),
                   key=lambda p: (0 if _nums(p["idx"]) else 1, p["sheet"], p["linha"]))
    for p in ordem:
        k = (p["tipo"], p["nome"])
        n = vistos.get(k, 0) + 1
        vistos[k] = n
        if n == 1:
            continue
        parts = p["nome"].split("_")
        if len(parts) >= 3:
            parts[2] = f"{parts[2]}-{n}"
        novo = "_".join(parts)
        nome_tdt[(p["sheet"], p["linha"])] = novo
        renomeados.append({**p, "novo": novo})
    if renomeados:
        print(f"nome duplicado RENOMEADO (todos entram na TDT): {len(renomeados)}")

    dups, semidx = diagnosticar(pts)
    print(f"diagnostico: {len(dups)} ocorrencias de coord repetida, {len(semidx)} com #REF!")
    final, mapa = sequenciar(pts)
    print(f"re-sequenciadas: {len(mapa)} coords ({sum(1 for m in mapa if m['mudou']=='SIM')} mudaram)")

    # lista de pontos CORRIGIDA (mesma estrutura, INDEX DNP3 arrumado)
    limpos = gerar_lista_corrigida(mapa)

    # siglas sem template + nomes duplicados
    sem_tpl = defaultdict(lambda: {"n": 0, "sheets": set(), "nomes": []})
    for p in pts:
        if p["tipo"] == "C" or not p["usado"]:
            continue
        if p["sigla"] not in TPL.get(p["tipo"], {}) and FALLBACK_SIGLA.get(p["sigla"]) not in TPL.get(p["tipo"], {}):
            d = sem_tpl[p["sigla"]]; d["n"] += 1; d["sheets"].add(p["sheet"])
            if len(d["nomes"]) < 3: d["nomes"].append(p["nome"])
    cnt = defaultdict(list)
    for p in pts:
        cnt[(p["nome"], p["tipo"])].append(p["sheet"])
    nomes_dup = [[n, t, len(s), ", ".join(sorted(set(s)))] for (n, t), s in cnt.items() if len(s) > 1]

    _rel = lambda fb, dmr, cm: gerar_relatorio(pts, mapa, dups, semidx, sem_tpl,
                                               nomes_dup, renomeados, fb, limpos,
                                               dmr, cm)

    # ── comandos: cada linha C precisa achar o sinal D que vai carregá-lo ──
    # 1) NOME idêntico (caso normal)
    # 2) mesmo MÓDULO + mesma SIGLA — a lista às vezes põe o comando no
    #    disjuntor e o status no módulo (CAS_LT1_52-1_25IE x CAS_LT1_LT1_25IE)
    # 3) sobrou: é comando PURO, vira um sinal novo com Input de preenchimento
    cmd = {}                       # NOME do sinal D -> coordenada de comando
    cmd_orfaos = []                # linhas C sem nenhum D -> sinal proprio
    cmd_realoc = []                # casadas pela regra 2 (vao pro relatorio)
    d_por_nome = {p["nome"]: p for p in pts if p["usado"] and p["tipo"] == "D"}
    d_por_mod_sigla = {}
    for p in pts:
        if p["usado"] and p["tipo"] == "D":
            parts = p["nome"].split("_")
            d_por_mod_sigla.setdefault((parts[1] if len(parts) > 1 else "",
                                        p["sigla"]), p)
    for p in pts:
        if p["tipo"] != "C" or not p["usado"] or not p["sigla"]:
            continue
        coord = final[(p["sheet"], p["linha"])]
        alvo = d_por_nome.get(p["nome"])
        if alvo is None:
            parts = p["nome"].split("_")
            alvo = d_por_mod_sigla.get((parts[1] if len(parts) > 1 else "", p["sigla"]))
            if alvo is not None:
                cmd_realoc.append({**p, "alvo": alvo["nome"], "coord": coord})
        if alvo is None:
            cmd_orfaos.append({**p, "coord": coord})
            continue
        cmd.setdefault(alvo["nome"], coord)
    if cmd_realoc:
        print(f"comando casado por MODULO+SIGLA (nome do device difere): {len(cmd_realoc)}")
    # comando puro vira um sinal discreto próprio, com Input de preenchimento
    sinteticos = []
    for i, o in enumerate(cmd_orfaos):
        chave = (f"{o['sheet']} (comando)", o["linha"])
        o["input"] = str(FILLER_BASE + i)
        final[chave] = o["input"]
        cmd.setdefault(o["nome"], o["coord"])
        sinteticos.append({**o, "tipo": "D", "sheet": chave[0], "linha": chave[1],
                           "idx": o["input"], "tipoPt": "Ponto Simples"})
    pts_tdt = pts + sinteticos
    if cmd_orfaos:
        print(f"comando PURO (sem sinal de status na lista): {len(cmd_orfaos)} "
              f"-> sinal novo com Input {FILLER_BASE}+")

    # monta as linhas da TDT
    wb = openpyxl.load_workbook(SKEL)
    plano = [("DNP3_DiscreteSignals", "D"), ("DNP3_AnalogSignals", "A"),
             ("DNP3_DiscreteAnalog", "A/D")]
    gerados = defaultdict(int); pulados = []; usou_fallback = []
    dm_rows = []; dm_origem = Counter(); dm_pendentes = []
    for sheet, tipo in plano:
        ws = wb[sheet]
        lab = {ws.cell(HEADER_ROWS, c).value: c for c in range(1, ws.max_column + 1)
               if ws.cell(HEADER_ROWS, c).value}
        ncol = ws.max_column
        styles = [copy(ws.cell(HEADER_ROWS + 1, c)._style) for c in range(1, ncol + 1)]
        if ws.max_row > HEADER_ROWS:
            ws.delete_rows(HEADER_ROWS + 1, ws.max_row - HEADER_ROWS)
        L = lambda n: lab.get(n)
        rows = []
        for p in pts_tdt:
            if p["tipo"] != tipo or not p["usado"]:
                continue
            tpl = TPL[tipo].get(p["sigla"])
            if not tpl:
                alt = FALLBACK_SIGLA.get(p["sigla"])
                tpl = TPL[tipo].get(alt) if alt else None
                if tpl:
                    usou_fallback.append({**p, "molde": alt})
            if not tpl:
                pulados.append(p)
                continue
            # NOME final (2ª ocorrência de um nome repetido leva sufixo no device)
            nome = nome_tdt.get((p["sheet"], p["linha"]), p["nome"])
            parts = p["nome"].split("_")           # device REAL, p/ o Device Mapping
            alias = parts[0]; mod = parts[1] if len(parts) > 1 else ""
            dev = parts[2] if len(parts) > 2 else ""
            mapping = {"<<PREFIX>>": f"{alias}_{mod}_{dev}", "<<ALIAS>>": alias,
                       "<<MODULE>>": mod, "<<DEVICE>>": dev, "<<N>>": "1"}
            row = [E._subst(v, mapping) for v in tpl]
            while len(row) < ncol:
                row.append(None)
            row = row[:ncol]
            def put(name, val):
                c = L(name)
                if c: row[c - 1] = val
            put("Signal Name", nome); put("Remote Point Name", nome)
            # descrição AUTORITATIVA da lista (o molde pode ser de sigla equivalente)
            if p.get("desc"):
                put("Signal Alias", p["desc"])
            put("Signal Custom ID", None)
            put("Remote Point Custom ID", f"{nome}_{RU}")
            put("Remote Unit", RU); put("Signal AOR Group", AOR)
            # Device Mapping SEMPRE sobrescrito: o molde traz o DM da subestação
            # de ORIGEM (lixo). Aqui vale a regra da PRÓPRIA CASCA, aprendida da
            # TDT atual — ver casca_devmap.py.
            dm, dm_o, dm_pend = devmap.resolver(p["nome"], p["sigla"])
            put("Device Mapping", dm)
            dm_origem[dm_o] += 1
            dm_rows.append([p["sheet"], p["linha"], p["tipo"], p["sigla"], nome, dm,
                            dm_o, dm_pend or "ok"])
            if dm_pend:
                dm_pendentes.append({"nome": nome, "sigla": p["sigla"], "dm": dm,
                                     "pend": dm_pend, "sheet": p["sheet"],
                                     "linha": p["linha"], "mod": mod, "dev": dev})
            put("Input Coordinates", final[(p["sheet"], p["linha"])])
            if p["escala"] not in (None, "", "-") and tipo in ("A", "A/D"):
                put("Scaling Factor", p["escala"])
            if tipo == "D":
                out = cmd.get(p["nome"])          # comando casa pelo NOME original
                if out:
                    put("Output Coordinates", out)
                else:  # sem comando: vira status puro (senao o validador exige Output)
                    for k in ("Output Coordinates", "Output Data Type", "Control Codes",
                              "Command Times [s]", "Commanding Mode"):
                        put(k, None)
                    ref = TPL["D"].get("MOLA") or []
                    for k in ("Direction", "Command Timeout [s]", "Scan After Command"):
                        c = L(k)
                        if c and c - 1 < len(ref):
                            row[c - 1] = ref[c - 1]
            rows.append(row)
        for i, row in enumerate(rows):
            for c in range(ncol):
                cell = ws.cell(HEADER_ROWS + 1 + i, c + 1, value=row[c])
                cell._style = copy(styles[c])
        gerados[sheet] = len(rows)

    # UTR nova
    ws = wb["DNP3_RTUs"]
    lab = {ws.cell(HEADER_ROWS, c).value: c for c in range(1, ws.max_column + 1)
           if ws.cell(HEADER_ROWS, c).value}
    r = HEADER_ROWS + 1
    def putr(name, val):
        c = lab.get(name)
        if c: ws.cell(r, c).value = val
    putr("Remote Unit (Terminal Server) Name", RU)
    putr("Remote Unit Alias", f"{RU}__")
    putr("Remote Unit Custom ID", None)          # ADMS gera
    putr("Remote Unit AOR Group", AOR)
    putr("Remote Unit Description", FABRICANTE)
    putr("Container Name", CONTAINER)
    putr("Container Custom ID", None)

    fbc = Counter((f["sigla"], f["molde"]) for f in usou_fallback)
    fbd = {f["sigla"]: f["desc"] for f in usou_fallback}
    _rel([[sg, mo, n, fbd.get(sg, "")] for (sg, mo), n in sorted(fbc.items(), key=lambda x: -x[1])],
         {"linhas": dm_rows, "origem": dm_origem, "pendentes": dm_pendentes},
         {"realoc": cmd_realoc, "orfaos": cmd_orfaos})
    print(f"relatorio: {OUT_REL.name} ({len(mapa)} coords, {len(dups)} repetidas, "
          f"{len(usou_fallback)} por equivalencia)")
    print(f"device mapping: {len(dm_rows)} sinais | origem: {dm_origem.most_common()}")

    buf = io.BytesIO(); wb.save(buf)
    OUT_TDT.write_bytes(excel_native.resave_native(buf.getvalue()))
    print(f"TDT: {OUT_TDT.name} | {dict(gerados)} | container={CONTAINER} | "
          f"pulados (sem template): {len(pulados)}")
    for p in pulados[:20]:
        print(f"   PULADO {p['sheet']}!L{p['linha']} {p['tipo']} {p['sigla']} {p['nome']}")


if __name__ == "__main__":
    main()
