"""
tdt_engine.py
=============
Motor de exportação ("clonador inteligente"). Carrega o template de referência
real (.xlsx) preservando 100% da estrutura/estilos e a sheet de validações
DMSMatchingTemplateInfo, depois escreve apenas os sinais selecionados aplicando
a substituição dinâmica dos campos dependentes da identidade do equipamento.

Garantias:
- Mantém as 4 linhas de cabeçalho exatamente como no original.
- Mantém DMSMatchingTemplateInfo intacta (enumerações/validações do ADMS).
- Copia o estilo da 1ª linha de dados original para cada nova linha (fonte,
  preenchimento, bordas, formato numérico) → formatação idêntica.
- Substitui placeholders <<PREFIX>>/<<ALIAS>>/<<MODULE>>/<<DEVICE>> pelos dados
  do novo equipamento e re-sequencia Custom IDs e coordenadas DNP3.
"""
from __future__ import annotations
import io
import json
import re
from copy import copy
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
CATALOG_PATH = DATA / "catalog.json"
TEMPLATE_PATH = DATA / "reference_template.xlsx"
SIGLA_INDEX_PATH = DATA / "sigla_index.json"

HEADER_ROWS = 4

# Presets de formato de comando (output) descobertos nas TDTs reais.
# count = nº de coordenadas/control-codes; times = tempos de pulso.
COMMAND_PRESETS = {
    "template": None,  # mantém o formato original do sinal
    "trip_close": {"controlCodes": "TripPulseOn;ClosePulseOn", "commandTimes": "0.25;0.25", "outputDataType": "SingleCoord", "count": 2},
    "close_close": {"controlCodes": "ClosePulseOn;ClosePulseOn", "commandTimes": "0.25;0.25", "outputDataType": "MultiCoord", "count": 2},
    "single_close": {"controlCodes": "ClosePulseOn", "commandTimes": "0.25", "outputDataType": "SingleCoord", "count": 1},
    "single_trip": {"controlCodes": "TripPulseOn", "commandTimes": "0.25", "outputDataType": "SingleCoord", "count": 1},
    "latch": {"controlCodes": "LatchOn;LatchOff", "commandTimes": "", "outputDataType": "MultiCoord", "count": 2},
}


class Catalog:
    def __init__(self):
        self.raw = json.loads(CATALOG_PATH.read_text(encoding="utf-8"))
        self.columns = self.raw["columns"]
        self._by_id = {d["id"]: d for d in self.raw["deviceTypes"]}

    def device(self, device_id):
        return self._by_id.get(device_id)

    def label_index(self, sheet, label):
        for c in self.columns[sheet]:
            if c["label"] == label:
                return c["index"]
        return None


def _fit_table(ws, last_data_row):
    """Estende (ou ajusta) o ref da Tabela do Excel da sheet para cobrir todas as
    linhas escritas — senão a formatação (faixa banded) para no fim do template."""
    end_row = max(HEADER_ROWS + 1, last_data_row)  # cabeçalho na linha 4, dados a partir da 5
    for name in list(ws.tables.keys()):
        tbl = ws.tables[name]
        m = re.match(r"^([A-Z]+\d+):([A-Z]+)\d+$", str(tbl.ref))
        if m:
            tbl.ref = f"{m.group(1)}:{m.group(2)}{end_row}"


def _subst(value, mapping):
    if not isinstance(value, str):
        return value
    out = value
    for ph, real in mapping.items():
        out = out.replace(ph, real)
    return out


def generate_tdt(config: dict) -> bytes:
    """
    config = {
      deviceTypeId, alias, module, device,
      remoteUnit?,                       # default UTR_{alias}_1
      customIdStart?: int (default 1),
      coordStart?: {discrete:int, analog:int},
      signals: [ {sheet, suffix, inputCoord?, customId?} ... ]
    }
    """
    cat = Catalog()
    dev = cat.device(config["deviceTypeId"])
    if not dev:
        raise ValueError(f"tipo de equipamento desconhecido: {config['deviceTypeId']}")

    alias = config["alias"].strip()
    module = (config.get("module") or "").strip()
    device = (config.get("device") or "").strip()
    new_prefix = f"{alias}_{module}_{device}"
    remote_unit = config.get("remoteUnit") or f"UTR_{alias}_1"

    mapping = {
        "<<PREFIX>>": new_prefix,
        "<<ALIAS>>": alias,
        "<<MODULE>>": module,
        "<<DEVICE>>": device,
        "<<N>>": str(config.get("transformerNumber") or "1"),
    }

    custom_start = int(config.get("customIdStart") or 1)
    coord_start = config.get("coordStart") or {}
    cmd_coord_start = config.get("commandCoordStart")
    cmd_coord_start = int(cmd_coord_start) if cmd_coord_start not in (None, "") else None

    # índice rápido dos sinais do catálogo por (sheet, suffix)
    sig_index = {}
    for klass, sheet in (("discrete", "DNP3_DiscreteSignals"), ("analog", "DNP3_AnalogSignals"),
                         ("discrete_analog", "DNP3_DiscreteAnalog")):
        for s in dev["signals"].get(klass, []):
            sig_index[(sheet, s["suffix"])] = s

    # agrupa seleção por sheet preservando a ordem do catálogo
    selected = {"DNP3_DiscreteSignals": [], "DNP3_AnalogSignals": [], "DNP3_DiscreteAnalog": []}
    for sel in config["signals"]:
        key = (sel["sheet"], sel["suffix"])
        if key in sig_index:
            selected[sel["sheet"]].append((sig_index[key], sel))

    wb = openpyxl.load_workbook(TEMPLATE_PATH)  # mantém estilos e validações

    for sheet, items in selected.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        cols = cat.columns[sheet]
        ncol = len(cols)
        klass = "analog" if "Analog" in sheet else "discrete"

        # índices de colunas chave
        idx_rp_custom = cat.label_index(sheet, "Remote Point Custom ID")
        idx_signal_custom = cat.label_index(sheet, "Signal Custom ID")
        idx_in_coord = cat.label_index(sheet, "Input Coordinates")
        idx_remote_unit = cat.label_index(sheet, "Remote Unit")
        idx_aor = cat.label_index(sheet, "Signal AOR Group")
        idx_out_coord = cat.label_index(sheet, "Output Coordinates")
        idx_out_dt = cat.label_index(sheet, "Output Data Type")
        idx_ctrl = cat.label_index(sheet, "Control Codes")
        idx_times = cat.label_index(sheet, "Command Times [s]")

        # captura estilos da 1ª linha de dados (linha 5 = header+1)
        style_row = HEADER_ROWS + 1
        styles = []
        for ci in range(ncol):
            c = ws.cell(row=style_row, column=ci + 1)
            styles.append(copy(c._style))

        # limpa todas as linhas de dados existentes
        last = ws.max_row
        if last > HEADER_ROWS:
            ws.delete_rows(HEADER_ROWS + 1, last - HEADER_ROWS)

        # sequência de Custom ID (DS=discrete, AS=analog)
        seq_tag = "AS" if klass == "analog" else "DS"
        custom_seq = custom_start
        coord_seq = int(coord_start.get(klass, 0) or 0)
        cmd_seq = cmd_coord_start

        for offset, (sig, sel) in enumerate(items):
            excel_row = HEADER_ROWS + 1 + offset
            row_vals = [_subst(v, mapping) for v in sig["row"]]
            while len(row_vals) < ncol:
                row_vals.append(None)

            # Signal Custom ID (GUID do objeto): ADMS gera — deixa vazio
            if idx_signal_custom is not None:
                row_vals[idx_signal_custom] = None

            # Remote Unit (garante consistência mesmo se template divergir)
            if idx_remote_unit is not None and isinstance(row_vals[idx_remote_unit], str):
                row_vals[idx_remote_unit] = remote_unit

            # Signal AOR Group: troca o 1º token (ex.: "SCD Distr") pelo alias
            if idx_aor is not None and isinstance(row_vals[idx_aor], str) and row_vals[idx_aor].strip():
                parts = row_vals[idx_aor].split(" ", 1)
                rest = (" " + parts[1]) if len(parts) > 1 else ""
                row_vals[idx_aor] = f"{alias}{rest}"

            # Remote Point Custom ID: re-sequencia
            if idx_rp_custom is not None and row_vals[idx_rp_custom]:
                cid = sel.get("customId")
                if cid:
                    row_vals[idx_rp_custom] = cid
                else:
                    row_vals[idx_rp_custom] = f"{alias}_{seq_tag}_{custom_seq:05d}"
                    custom_seq += 1

            # Input Coordinates: override > auto-sequência > valor do template
            if idx_in_coord is not None:
                ic = sel.get("inputCoord")
                if ic is not None and ic != "":
                    row_vals[idx_in_coord] = ic
                elif "discrete" in coord_start or "analog" in coord_start:
                    row_vals[idx_in_coord] = coord_seq
                    coord_seq += 1
                # senão mantém o valor original tokenizado (numérico do template)

            # ----- Comando (output) -----
            cmd = sig.get("command")
            if cmd and idx_out_coord is not None:
                fmt = sel.get("commandFormat") or "template"
                preset = COMMAND_PRESETS.get(fmt)
                if preset:
                    count = preset["count"]
                    if idx_ctrl is not None:
                        row_vals[idx_ctrl] = preset["controlCodes"]
                    if idx_times is not None:
                        row_vals[idx_times] = preset["commandTimes"]
                    if idx_out_dt is not None:
                        row_vals[idx_out_dt] = preset["outputDataType"]
                else:
                    count = cmd.get("coordCount") or 1  # mantém formato do template

                # Output Coordinates: override > auto-sequência > valor do template
                oc = sel.get("outputCoord")
                if oc is not None and oc != "":
                    row_vals[idx_out_coord] = oc
                elif cmd_seq is not None:
                    row_vals[idx_out_coord] = f"{cmd_seq};{cmd_seq}" if count == 2 else cmd_seq
                    cmd_seq += 1
                # senão mantém o output coord do template

            for ci in range(ncol):
                cell = ws.cell(row=excel_row, column=ci + 1, value=row_vals[ci])
                cell._style = copy(styles[ci])

        _fit_table(ws, HEADER_ROWS + len(items))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ===========================================================================
# Importação de LISTA DE PONTOS (formato padrão de entrada)
# ===========================================================================

def _col(headers, *keywords):
    """acha o índice da coluna cujo cabeçalho contém todas as keywords."""
    for i, h in enumerate(headers):
        hs = str(h or "").upper()
        if all(k.upper() in hs for k in keywords):
            return i
    return None


def parse_points_list(file_bytes: bytes) -> dict:
    """Lê a lista de pontos (.xlsx) -> {discrete:[...], analog:[...]}."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    out = {"discrete": [], "analog": [], "discrete_analog": [], "inputErrors": []}
    names = {n.lower(): n for n in wb.sheetnames}

    def find(*cands):
        for c in cands:
            if c in names:
                return names[c]
        return None

    sd = find("discreto", "discretos", "digital", "digitais")
    if sd:
        ws = wb[sd]
        rows = list(ws.iter_rows(values_only=True))
        h = rows[0]
        ci_sig = _col(h, "SIGLA"); ci_nome = _col(h, "NOME")
        ci_in = _col(h, "ENTRADA"); ci_cmd = _col(h, "COMANDO"); ci_aor = _col(h, "AOR")
        for r in rows[1:]:
            if not r or not r[ci_nome if ci_nome is not None else 1]:
                continue
            out["discrete"].append({
                "sigla": str(r[ci_sig]).strip() if ci_sig is not None and r[ci_sig] else "",
                "nome": str(r[ci_nome]).strip(),
                "inCoord": r[ci_in] if ci_in is not None else None,
                "outCoord": r[ci_cmd] if ci_cmd is not None else None,
                "aor": str(r[ci_aor]).strip() if ci_aor is not None and r[ci_aor] else None,
            })

    sa = find("analogicos", "analógicos", "analogico", "analog", "analogicas")
    if sa:
        ws = wb[sa]
        rows = list(ws.iter_rows(values_only=True))
        h = rows[0]
        ci_sig = _col(h, "SIGLA"); ci_nome = _col(h, "NOME")
        ci_esc = _col(h, "ESCALA"); ci_in = _col(h, "INDEX"); ci_aor = _col(h, "AOR")
        for r in rows[1:]:
            if not r or not r[ci_nome if ci_nome is not None else 1]:
                continue
            out["analog"].append({
                "sigla": str(r[ci_sig]).strip() if ci_sig is not None and r[ci_sig] else "",
                "nome": str(r[ci_nome]).strip(),
                "escala": r[ci_esc] if ci_esc is not None else None,
                "inCoord": r[ci_in] if ci_in is not None else None,
                "aor": str(r[ci_aor]).strip() if ci_aor is not None and r[ci_aor] else None,
            })

    # aba Digital-Analógico (A/D), ex.: TAP — mesmas colunas do analógico
    sda = find("discretoanalogico", "discreto analogico", "digital-analogico",
               "digitalanalogico", "a/d", "discreteanalog")
    if sda:
        ws = wb[sda]
        rows = list(ws.iter_rows(values_only=True))
        h = rows[0]
        ci_sig = _col(h, "SIGLA"); ci_nome = _col(h, "NOME")
        ci_esc = _col(h, "ESCALA"); ci_in = _col(h, "INDEX"); ci_aor = _col(h, "AOR")
        for r in rows[1:]:
            if not r or not r[ci_nome if ci_nome is not None else 1]:
                continue
            out["discrete_analog"].append({
                "sigla": str(r[ci_sig]).strip() if ci_sig is not None and r[ci_sig] else "",
                "nome": str(r[ci_nome]).strip(),
                "escala": r[ci_esc] if ci_esc is not None else None,
                "inCoord": r[ci_in] if ci_in is not None else None,
                "aor": str(r[ci_aor]).strip() if ci_aor is not None and r[ci_aor] else None,
            })

    # aba de erros do pré-processador (sinais rejeitados antes de chegar aqui)
    se = find("erros", "erro", "errors")
    if se:
        rows = list(wb[se].iter_rows(values_only=True))
        h = rows[0] if rows else []
        c_sig = _col(h, "SIGLA"); c_mod = _col(h, "MODUL", "MÓDUL"); c_tipo = _col(h, "TIPO")
        c_mot = _col(h, "MOTIVO"); c_aba = _col(h, "ABA"); c_lin = _col(h, "LINHA")
        for r in rows[1:]:
            if not r or all(v in (None, "") for v in r):
                continue
            out["inputErrors"].append({
                "aba": str(r[c_aba]).strip() if c_aba is not None and r[c_aba] else "",
                "linha": r[c_lin] if c_lin is not None else None,
                "sigla": str(r[c_sig]).strip() if c_sig is not None and r[c_sig] else "",
                "modulo": str(r[c_mod]).strip() if c_mod is not None and r[c_mod] else "",
                "tipo": str(r[c_tipo]).strip() if c_tipo is not None and r[c_tipo] else "",
                "motivo": str(r[c_mot]).strip() if c_mot is not None and r[c_mot] else "",
            })
    return out


PROTOCOLS = {
    "dnp3": {"template": TEMPLATE_PATH, "index": SIGLA_INDEX_PATH,
             "discrete": "DNP3_DiscreteSignals", "analog": "DNP3_AnalogSignals",
             "discrete_analog": "DNP3_DiscreteAnalog"},
    "iec104": {"template": DATA / "reference_template_iec104.xlsx",
               "index": DATA / "sigla_index_iec104.json",
               "discrete": "IEC104_DiscreteSignals", "analog": "IEC104_AnalogSignals",
               "discrete_analog": "IEC104_DiscreteAnalog"},
}


def _labels(ws):
    """label -> índice, lido da linha de rótulos (linha HEADER_ROWS)."""
    row = next(ws.iter_rows(min_row=HEADER_ROWS, max_row=HEADER_ROWS, values_only=True))
    return {str(v).strip(): i for i, v in enumerate(row) if v}


def build_import_report(parsed: dict, protocol: str = "dnp3") -> bytes:
    """Gera um relatório .xlsx: Resumo, Revisão da Lista (todos os pontos com
    status) e Problemas (não casados, duplicados, erros de origem)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    proto = PROTOCOLS.get(protocol, PROTOCOLS["dnp3"])
    index = json.loads(Path(proto["index"]).read_text(encoding="utf-8"))
    didx, aidx = index[proto["discrete"]], index[proto["analog"]]

    from collections import Counter
    allnames = [it["nome"] for it in parsed["discrete"] + parsed["analog"] if it["nome"]]
    dupset = {n for n, c in Counter(allnames).items() if c > 1}

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2563EB")
    fills = {"OK": PatternFill("solid", fgColor="DCFCE7"),
             "NAO_ENCONTRADO": PatternFill("solid", fgColor="FEE2E2"),
             "DUPLICADO": PatternFill("solid", fgColor="FEF3C7")}

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    # ---- Resumo ----
    ws = wb.active; ws.title = "Resumo"
    nd, na = len(parsed["discrete"]), len(parsed["analog"])
    md = sum(1 for it in parsed["discrete"] if it["sigla"] in didx)
    ma = sum(1 for it in parsed["analog"] if it["sigla"] in aidx)
    aliases = sorted({n.split("_")[0] for n in allnames})
    info = [
        ("Protocolo", protocol.upper()),
        ("Subestações", ", ".join(aliases)),
        ("Sinais digitais (total / casados)", f"{nd} / {md}"),
        ("Sinais analógicos (total / casados)", f"{na} / {ma}"),
        ("Não encontrados na base ADMS", (nd - md) + (na - ma)),
        ("Nomes duplicados na lista", len(dupset)),
        ("Sinais rejeitados na origem (aba Erros)", len(parsed.get("inputErrors", []))),
    ]
    ws["A1"] = "Relatório de Importação — TDT ADMS"; ws["A1"].font = Font(bold=True, size=14)
    for i, (k, v) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 40

    # ---- Revisão da Lista ----
    ws = wb.create_sheet("Revisão da Lista")
    cols = ["#", "Tipo", "SIGLA", "Nome do Sinal", "Entrada", "Comando/Index", "Escala", "AOR", "Status"]
    ws.append(cols)
    n = 0
    for klass, items, idxset in (("Digital", parsed["discrete"], didx), ("Analógico", parsed["analog"], aidx)):
        for it in items:
            n += 1
            if it["nome"] in dupset:
                st = "DUPLICADO"
            elif it["sigla"] in idxset:
                st = "OK"
            else:
                st = "NAO_ENCONTRADO"
            ws.append([
                n, klass, it["sigla"], it["nome"],
                it.get("inCoord"),
                it.get("outCoord") if klass == "Digital" else it.get("inCoord"),
                it.get("escala") if klass == "Analógico" else None,
                it.get("aor"),
                {"OK": "OK", "NAO_ENCONTRADO": "Não encontrado na base", "DUPLICADO": "Nome duplicado"}[st],
            ])
            ws.cell(row=ws.max_row, column=9).fill = fills[st]
    style_header(ws, len(cols))
    for col, w in zip("ABCDEFGHI", [6, 11, 12, 34, 12, 14, 8, 14, 24]):
        ws.column_dimensions[col].width = w

    # ---- Problemas ----
    ws = wb.create_sheet("Problemas")
    ws.append(["Categoria", "SIGLA", "Nome / Detalhe", "Motivo"])
    for it in parsed["discrete"] + parsed["analog"]:
        if it["nome"] and it["sigla"] not in (didx if it in parsed["discrete"] else aidx):
            pass  # tratado abaixo de forma robusta
    for it in parsed["discrete"]:
        if it["sigla"] not in didx:
            ws.append(["Não encontrado", it["sigla"], it["nome"], "SIGLA ausente na base ADMS"])
    for it in parsed["analog"]:
        if it["sigla"] not in aidx:
            ws.append(["Não encontrado", it["sigla"], it["nome"], "SIGLA ausente na base ADMS"])
    for nme in sorted(dupset):
        ws.append(["Duplicado", "", nme, "Nome repetido na lista (gera linha duplicada)"])
    for e in parsed.get("inputErrors", []):
        ws.append(["Rejeitado na origem", e.get("sigla", ""), f"{e.get('modulo','')} {e.get('tipo','')}".strip(), e.get("motivo", "")])
    style_header(ws, 4)
    for col, w in zip("ABCD", [20, 12, 40, 48]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def generate_tdt_from_list(parsed: dict, protocol: str = "dnp3"):
    """Gera a TDT a partir da lista de pontos. Retorna (bytes, report)."""
    proto = PROTOCOLS.get(protocol, PROTOCOLS["dnp3"])
    index = json.loads(Path(proto["index"]).read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(proto["template"])
    report = {"discrete": {"matched": 0, "unmatched": []},
              "analog": {"matched": 0, "unmatched": []},
              "discrete_analog": {"matched": 0, "unmatched": []}}
    custom_seq = {}  # (alias, tag) -> contador

    plan = [("discrete", proto["discrete"], "DS"),
            ("analog", proto["analog"], "AS"),
            ("discrete_analog", proto.get("discrete_analog"), "DA")]

    # pré-limpa sheets de sinais que NÃO terão dados (remove linhas-fantasma do template)
    written = {sheet for klass, sheet, _ in plan if parsed.get(klass) and sheet in index}
    for _k, sh, _t in plan:
        if sh and sh in wb.sheetnames and sh not in written:
            ws0 = wb[sh]
            if ws0.max_row > HEADER_ROWS:
                ws0.delete_rows(HEADER_ROWS + 1, ws0.max_row - HEADER_ROWS)

    for klass, sheet, tag in plan:
        items = parsed.get(klass, [])
        if not items or not sheet or sheet not in wb.sheetnames or sheet not in index:
            continue
        ws = wb[sheet]
        lab = _labels(ws)
        ncol = max(lab.values()) + 1 if lab else 0
        def L(*cands):
            for c in cands:
                if c in lab:
                    return lab[c]
            return None
        idx_name = L("Signal Name"); idx_rpname = L("Remote Point Name")
        idx_in = L("Input Coordinates", "Input Coordinate")
        idx_out = L("Output Coordinates", "Output Coordinate")
        idx_aor = L("Signal AOR Group"); idx_ru = L("Remote Unit")
        idx_cid = L("Remote Point Custom ID"); idx_scale = L("Scaling Factor")
        idx_scustom = L("Signal Custom ID")

        style_row = HEADER_ROWS + 1
        styles = [copy(ws.cell(row=style_row, column=c + 1)._style) for c in range(ncol)]
        last = ws.max_row
        if last > HEADER_ROWS:
            ws.delete_rows(HEADER_ROWS + 1, last - HEADER_ROWS)

        out_row = 0
        for it in items:
            tpl = index[sheet].get(it["sigla"])
            if not tpl:
                report[klass]["unmatched"].append({"sigla": it["sigla"], "nome": it["nome"]})
                continue
            nome = it["nome"]
            p = nome.split("_")
            alias = p[0] if p else nome
            module = p[1] if len(p) > 1 else ""
            device = p[2] if len(p) > 2 else ""
            mapping = {"<<PREFIX>>": f"{alias}_{module}_{device}", "<<ALIAS>>": alias,
                       "<<MODULE>>": module, "<<DEVICE>>": device, "<<N>>": ""}
            row = [_subst(c, mapping) for c in tpl]
            while len(row) < ncol:
                row.append(None)

            if idx_name is not None: row[idx_name] = nome
            if idx_rpname is not None: row[idx_rpname] = nome
            if idx_scustom is not None: row[idx_scustom] = None  # GUID do objeto: ADMS gera
            if idx_ru is not None and isinstance(row[idx_ru], str): row[idx_ru] = f"UTR_{alias}_1"
            # coordenadas: a LISTA é autoritativa (em branco => limpa)
            if idx_in is not None:
                row[idx_in] = it.get("inCoord") if it.get("inCoord") not in (None, "") else None
            if klass == "discrete" and idx_out is not None:
                row[idx_out] = it.get("outCoord") if it.get("outCoord") not in (None, "") else None
            if klass in ("analog", "discrete_analog") and idx_scale is not None and it.get("escala") not in (None, ""):
                row[idx_scale] = it["escala"]
            if idx_aor is not None and it.get("aor"):
                row[idx_aor] = f"{alias} {it['aor']}"
            if idx_cid is not None and row[idx_cid]:
                k = (alias, tag)
                custom_seq[k] = custom_seq.get(k, 0) + 1
                row[idx_cid] = f"{alias}_{tag}_{custom_seq[k]:05d}"

            er = HEADER_ROWS + 1 + out_row
            for c in range(ncol):
                cell = ws.cell(row=er, column=c + 1, value=row[c])
                cell._style = copy(styles[c])
            out_row += 1
            report[klass]["matched"] += 1

        _fit_table(ws, HEADER_ROWS + out_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), report


if __name__ == "__main__":
    # round-trip de fidelidade: reconstrói a TDT FWB original
    cat = Catalog()
    dev = cat.device("alimentador")
    signals = []
    for s in dev["signals"]["discrete"]:
        signals.append({"sheet": "DNP3_DiscreteSignals", "suffix": s["suffix"]})
    for s in dev["signals"]["analog"]:
        signals.append({"sheet": "DNP3_AnalogSignals", "suffix": s["suffix"]})
    data = generate_tdt({
        "deviceTypeId": "alimentador",
        "alias": "FWB", "module": "AL13", "device": "52-13",
        "signals": signals,
    })
    out = ROOT / "roundtrip_FWB.xlsx"
    out.write_bytes(data)
    print(f"round-trip salvo em {out} ({len(data)//1024} KB)")
