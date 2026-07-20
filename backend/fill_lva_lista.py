"""
fill_lva_lista.py — Preenche SIGLA + INDEX DNP3 + EQUIPAMENTO dos alimentadores
que faltam na Lista de pontos LVA_V02, ESPELHANDO as TDTs já importadas no ADMS.

Fonte da verdade: aba AL21 (o chefe preencheu; as TDTs geradas usaram este
gabarito e foram importadas OK). Cada feeder recebe os mesmos SIGLAs/índices
relativos do AL21, deslocados pelo bloco do feeder:
  AL12=100 · AL13=200 · AL14=300 · AL15=400 · AL24=700 · AL23=800
(bases confirmadas pelos marcadores '26_49 Excluido' já presentes na lista.)

Casa linha-a-linha por (tipo normalizado, descrição normalizada). Tipos crus do
AL12 (SP/DP/ME_FL) são mapeados. Linhas sem correspondência no AL21 (secc
feeder-específicas, ERAC, pickups, '79 Em Progresso'...) ficam vazias — como o
próprio AL21/AL22 do chefe. NÃO toca em AL11/AL21/AL22/BC1 (já prontos).

Uso: python fill_lva_lista.py
"""
from __future__ import annotations
import re
from pathlib import Path
import openpyxl

SRC = Path("C:/Users/egnpo/Downloads/Lista de pontos LVA_V02 (1).xlsx")
OUT = Path("C:/Users/egnpo/Downloads/Lista de pontos LVA_V02_preenchida.xlsx")

BASES = {"AL12": 100, "AL13": 200, "AL14": 300, "AL15": 400, "AL24": 700, "AL23": 800}
AL21_BASE = 500
TYPE_MAP = {"A": "A", "D": "D", "C": "C", "SP": "D", "DP": "D",
            "ME_FL": "A", "ME": "A", "DC": "C", "SC": "C"}


def _norm(d) -> str:
    d = str(d or "").strip().upper()
    d = re.sub(r"52-?\d+", "52X", d)
    d = re.sub(r"29-?\d+", "29X", d)
    d = re.sub(r"BI\d\.\d", "BIX", d)
    return re.sub(r"\s+", " ", d)


def _cols(ws):
    hdr = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    off = 1 if hdr and str(hdr[0]).strip().upper() == "LINHA" else 0
    # DESCRIÇÃO, TIPO, MÓDULO, EQUIPAMENTO, SIGLA, INDEX  (1-based col numbers)
    return {"desc": 2 + off, "tipo": 3 + off, "mod": 4 + off,
            "equip": 5 + off, "sigla": 6 + off, "idx": 7 + off}


def _shift(idx_str: str, delta: int) -> str:
    parts = [str(int(p) + delta) for p in str(idx_str).split(";")]
    return ";".join(parts)


def main():
    wb = openpyxl.load_workbook(SRC)

    # canonical AL21: (tipo, norm_desc) -> (sigla, equip_kind, rel_index)
    ws21 = wb["AL21"]; c = _cols(ws21)
    canon = {}; canon_desc = {}
    for r in range(2, ws21.max_row + 1):
        desc = ws21.cell(r, c["desc"]).value
        sig = ws21.cell(r, c["sigla"]).value
        idx = ws21.cell(r, c["idx"]).value
        if not desc or not sig or idx in (None, ""):
            continue
        equip = str(ws21.cell(r, c["equip"]).value or "").strip()
        kind = "BRK" if equip.startswith("52-") else equip   # BRK | TC | TP
        tipo = TYPE_MAP.get(str(ws21.cell(r, c["tipo"]).value or "").strip(),
                            str(ws21.cell(r, c["tipo"]).value or "").strip())
        key = (tipo, _norm(desc))
        canon[key] = (str(sig).strip(), kind, _shift(idx, -AL21_BASE))
        canon_desc[key] = str(desc)
    print(f"gabarito AL21: {len(canon)} sinais")

    for mod, base in BASES.items():
        ws = wb[mod]; c = _cols(ws)
        alnum = mod[2:]                       # '12', '13', ...
        brk = f"52-{alnum}"
        filled = 0; reserva = []
        for r in range(2, ws.max_row + 1):
            desc = ws.cell(r, c["desc"]).value
            if not desc:
                continue
            raw_t = str(ws.cell(r, c["tipo"]).value or "").strip()
            tipo = TYPE_MAP.get(raw_t, raw_t)
            key = (tipo, _norm(desc))
            hit = canon.get(key)
            if not hit:
                reserva.append(f"[{raw_t}] {desc}")
                continue
            sig, kind, rel = hit
            ws.cell(r, c["sigla"]).value = sig
            ws.cell(r, c["idx"]).value = _shift(rel, base)
            ws.cell(r, c["tipo"]).value = tipo          # normaliza SP/DP/ME_FL
            ws.cell(r, c["equip"]).value = brk if kind == "BRK" else kind
            filled += 1
        # --- ADICIONA as linhas do gabarito AL21 que a aba não tem (79LO, DR,
        #     FCOM, AJG2...) — o "79LO tava faltando" do chefe é uma delas ---
        present = set()
        for r in range(2, ws.max_row + 1):
            d = ws.cell(r, c["desc"]).value
            if d:
                raw = str(ws.cell(r, c["tipo"]).value or "").strip()
                present.add((TYPE_MAP.get(raw, raw), _norm(d)))
        # topologia da aba (coluna 2) e nº da coluna LINHA
        topo_col = c["desc"] - 1
        topo = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, topo_col).value:
                topo = ws.cell(r, topo_col).value
                break
        added = []
        for (tipo, ndesc), (sig, kind, rel) in canon.items():
            if (tipo, ndesc) in present:
                continue
            # descrição real do gabarito (com nº do disjuntor trocado)
            real_desc = canon_desc[(tipo, ndesc)].replace("52-21", brk)
            row = ws.max_row + 1
            if topo_col >= 1:
                ws.cell(row, topo_col).value = topo
            ws.cell(row, c["desc"]).value = real_desc
            ws.cell(row, c["tipo"]).value = tipo
            ws.cell(row, c["mod"]).value = mod
            ws.cell(row, c["equip"]).value = brk if kind == "BRK" else kind
            ws.cell(row, c["sigla"]).value = sig
            ws.cell(row, c["idx"]).value = _shift(rel, base)
            added.append(f"[{tipo}] {sig} idx={_shift(rel, base)} ({real_desc})")

        print(f"{mod} (base {base}): {filled} preenchidos | {len(added)} ADICIONADOS | {len(reserva)} reserva")
        for x in added:
            print(f"      + {x}")

    wb.save(OUT)
    print(f"\nsalva: {OUT.name}")
    print("intactos (ja prontos): AL11, AL21, AL22, BC1")


if __name__ == "__main__":
    main()
