"""
casca_status.py — Cruza o retorno do ADMS com a TDT e diz, dispositivo por
dispositivo, o que ainda falta no Casca_Obra.

O CSV de erros do ADMS é a fonte mais confiável que existe sobre o modelo: se
um sinal mapeou, o dispositivo EXISTE e é ÚNICO; se falhou, o texto diz se ele
está duplicado ("Found multiple devices") ou ausente ("Could not find any").

Gera CASCA_STATUS_IMPORT.xlsx:
  1-Resumo            quantos mapearam / falharam
  2-Por modulo        idem, quebrado por vao
  3-Por dispositivo   cada Device Mapping: usado por N sinais, status e acao
  4-Sinais que falharam
  5-Dispositivos OK   os que ja estao certos no modelo (nao mexer)

Uso: python casca_status.py [erros.csv]
"""
from __future__ import annotations
import collections
import csv
import io
import re
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

DOWN = Path("C:/Users/egnpo/Downloads")
TDT = DOWN / "TDT_CASCA_UTR_CAS_3.xlsx"
_n = DOWN / "TDT_CASCA_UTR_CAS_3_NOVA.xlsx"
if _n.exists() and (not TDT.exists() or _n.stat().st_mtime > TDT.stat().st_mtime):
    TDT = _n
OUT = DOWN / "CASCA_STATUS_IMPORT.xlsx"
HR = 4


def ler_erros(caminho: Path):
    """nome do sinal -> (classe, device mapping citado, mensagem)."""
    out = {}
    with io.open(caminho, encoding="utf-8-sig", errors="replace") as fh:
        for r in csv.reader(fh, delimiter=";"):
            if len(r) < 4 or not r[0] or r[0].lower().startswith(("severidade", "informa")):
                continue
            d = r[3]
            m = re.search(r"Mapping: (\S+?)\.? Signal", d)
            dm = m.group(1) if m else ""
            if "successfully mapped" in d or "updated with new values" in d:
                classe = "MAPEOU"                        # sucesso!
                m2 = re.search(r"device[:]? (\S+?)[\. ]", d)
                dm = m2.group(1) if m2 else dm
            elif "Found multiple" in d:
                classe = "DUPLICADO no modelo"
            elif "Could not find any" in d:
                classe = ("na UTR (dispositivo nao existe)" if dm == "UTR_CAS_3"
                          else "DISPOSITIVO SEM _NEW")
            elif "already" in d:                          # 2 sinais / 1 device
                classe = "SINAL JA MAPEADO no dispositivo"
                m2 = re.search(r"on device[:]? (\S+)", d)
                dm = m2.group(1).rstrip(".") if m2 else dm
            else:
                classe = "outro"
            out[r[1]] = (classe, dm, d)
    return out


def ler_tdt():
    wb = openpyxl.load_workbook(TDT, read_only=True, data_only=True)
    sinais = []
    for sn in ("DNP3_DiscreteSignals", "DNP3_AnalogSignals", "DNP3_DiscreteAnalog"):
        rows = list(wb[sn].iter_rows(values_only=True))
        if len(rows) <= HR:
            continue
        hdr = [str(c or "").strip() for c in rows[HR - 1]]
        ix = {n: i for i, n in enumerate(hdr) if n}
        for r in rows[HR:]:
            if not r or not r[0]:
                continue
            sinais.append({
                "nome": str(r[0]),
                "tipo": "analogico" if "Analog" in sn else "discreto",
                "dm": str(r[ix["Device Mapping"]] or ""),
                "modulo": str(r[0]).split("_")[1] if str(r[0]).count("_") > 1 else "",
            })
    wb.close()
    return sinais


ACAO = {
    "DUPLICADO no modelo":
        "Dar ID de Mapeamento SCADA UNICO a cada dispositivo. O mesmo ID esta em "
        "2+ equipamentos (a copia Casca_Obra herdou o ID da CASCA original).",
    "NAO EXISTE no modelo":
        "Criar o dispositivo no Casca_Obra com este ID de Mapeamento SCADA.",
    "SINAL JA EXISTE no dispositivo":
        "O dispositivo ja tem um sinal com o mesmo papel vindo da UTR antiga. "
        "Decidir se o ponto novo entra em outro dispositivo ou se o antigo sai.",
    "outro": "Ver a mensagem completa na aba 4.",
}


def main():
    csv_erros = Path(sys.argv[1]) if len(sys.argv) > 1 else DOWN / "eros3.csv"
    erros = ler_erros(csv_erros)
    sinais = ler_tdt()
    for s in sinais:
        e = erros.get(s["nome"])
        s["status"] = e[0] if e else "MAPEOU"
        s["msg"] = e[2] if e else ""
        # o ADMS cita o DM que ele tentou; vale mais que o meu
        if e and e[1]:
            s["dm_adms"] = e[1]

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    bold = Font(bold=True); hdrf = PatternFill("solid", fgColor="DDEBF7")
    vermelho = PatternFill("solid", fgColor="FFC7CE")
    verde = PatternFill("solid", fgColor="C6EFCE")
    amarelo = PatternFill("solid", fgColor="FFEB9C")

    def sheet(t, cols, rows, fill=None, larg=(46, 14, 12, 12, 30, 70)):
        ws = wb.create_sheet(t[:31]); ws.append(cols)
        for c in range(1, len(cols) + 1):
            ws.cell(1, c).font = bold; ws.cell(1, c).fill = hdrf
        for r in rows:
            ws.append(r)
            f = fill(r) if fill else None
            if f:
                for c in range(1, len(cols) + 1):
                    ws.cell(ws.max_row, c).fill = f
        for i, w in enumerate(larg, 1):
            if i <= len(cols):
                ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"
        return ws

    ok = sum(1 for s in sinais if s["status"] == "MAPEOU")
    porstatus = collections.Counter(s["status"] for s in sinais)
    sheet("1-Resumo", ["Item", "Qtde", "%"],
          [["Sinais na TDT", len(sinais), "100%"],
           ["MAPEARAM no modelo", ok, f"{ok*100//len(sinais)}%"],
           ["Falharam", len(sinais) - ok, f"{(len(sinais)-ok)*100//len(sinais)}%"]]
          + [[f"  .. {k}", v, ""] for k, v in porstatus.most_common() if k != "MAPEOU"],
          larg=(40, 10, 8))

    pm = collections.defaultdict(lambda: collections.Counter())
    for s in sinais:
        pm[s["modulo"]]["tot"] += 1
        pm[s["modulo"]]["ok" if s["status"] == "MAPEOU" else "erro"] += 1
    sheet("2-Por modulo", ["Modulo", "Sinais", "Mapearam", "Falharam", "Situacao"],
          [[m, c["tot"], c["ok"], c["erro"],
            "COMPLETO" if not c["erro"] else
            ("nada mapeou" if not c["ok"] else "parcial")]
           for m, c in sorted(pm.items(), key=lambda x: -x[1]["erro"])],
          fill=lambda r: verde if r[3] == 0 else (vermelho if r[2] == 0 else amarelo),
          larg=(14, 10, 12, 12, 16))

    pd = collections.defaultdict(lambda: {"n": 0, "ok": 0, "st": collections.Counter(),
                                          "ex": [], "tipos": set()})
    for s in sinais:
        d = pd[s.get("dm_adms") or s["dm"]]
        d["n"] += 1
        d["tipos"].add(s["tipo"])
        if s["status"] == "MAPEOU":
            d["ok"] += 1
        else:
            d["st"][s["status"]] += 1
            if len(d["ex"]) < 3:
                d["ex"].append(s["nome"])
    linhas = []
    for dm, d in sorted(pd.items(), key=lambda x: (-sum(x[1]["st"].values()), x[0])):
        st = d["st"].most_common(1)[0][0] if d["st"] else "MAPEOU"
        linhas.append([dm, d["n"], d["ok"], d["n"] - d["ok"], st,
                       ACAO.get(st, "") if st != "MAPEOU" else "nada a fazer"])
    sheet("3-Por dispositivo",
          ["Device Mapping", "Sinais", "Mapearam", "Falharam", "Status", "Acao"],
          linhas,
          fill=lambda r: verde if r[3] == 0 else (vermelho if r[2] == 0 else amarelo))

    sheet("4-Sinais que falharam",
          ["Sinal", "Modulo", "Tipo", "Device Mapping", "Status", "Mensagem do ADMS"],
          [[s["nome"], s["modulo"], s["tipo"], s.get("dm_adms") or s["dm"],
            s["status"], s["msg"]] for s in sinais if s["status"] != "MAPEOU"],
          fill=lambda r: vermelho)

    sheet("5-Dispositivos OK",
          ["Device Mapping", "Sinais que mapearam", "Tipos de sinal"],
          [[dm, d["ok"], ", ".join(sorted(d["tipos"]))]
           for dm, d in sorted(pd.items()) if not d["st"]],
          fill=lambda r: verde, larg=(46, 20, 20))

    destino = OUT
    try:
        destino.unlink(missing_ok=True)
        wb.save(destino)
    except PermissionError:                       # aberto no Excel
        destino = OUT.with_name(OUT.stem + "_NOVO.xlsx")
        destino.unlink(missing_ok=True)
        wb.save(destino)
    globals()["OUT"] = destino
    print(f"{destino.name}: {ok}/{len(sinais)} sinais mapearam "
          f"({len(sinais)-ok} falharam)")
    for k, v in porstatus.most_common():
        if k != "MAPEOU":
            print(f"   {v:>4}  {k}")
    disp_ruins = sum(1 for d in pd.values() if d["st"])
    print(f"   dispositivos a resolver no modelo: {disp_ruins}")


if __name__ == "__main__":
    main()
