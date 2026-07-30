"""
futura_motivos.py — Por que cada sinal da TDT FUTURA nao entrou na ATUAL.

Cruza TDT_CASCA_FUTURA.xlsx com a aba "19-Sem dispositivo no modelo" do
relatorio e produz UMA LINHA POR SINAL, com a categoria, o motivo cru do
gerador e o que precisa ser feito no modelo para o sinal entrar.

Uso: python futura_motivos.py
Saida: C:/Users/egnpo/Downloads/CASCA_FUTURA_MOTIVOS.xlsx
"""
from __future__ import annotations
import collections
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import excel_native

DOWN = Path("C:/Users/egnpo/Downloads")
FUT = DOWN / "TDT_CASCA_FUTURA.xlsx"
REL = DOWN / "CASCA_RELATORIO.xlsx"
OUT = DOWN / "CASCA_FUTURA_MOTIVOS.xlsx"
HR = 4

# categoria -> (rotulo curto, o que fazer no modelo)
ACAO = {
    "A": ("Falta o sufixo _NEW",
          "O dispositivo EXISTE no Cas_Obra, so nao foi renomeado com _NEW. "
          "Renomear o ID de Mapeamento SCADA — e o conserto mais barato."),
    "B": ("Vao inteiro nao existe",
          "Desenhar o vao no Cas_Obra (disjuntor, TC, TP, seccionadoras e "
          "reles). E obra, nao e erro de mapeamento."),
    "C": ("Rele especifico nao existe",
          "Criar o rele da funcao no vao. Enquanto ele nao existe, a "
          "TDT_CASCA_ATUAL_COMPLETA poe o sinal no _PROT generico do vao "
          "(ou no disjuntor, se for sinal solto) — ver as duas colunas ao lado."),
    "D": ("Papel fisico ja ocupado",
          "O modelo tem MENOS equipamento que a lista (ex.: TR2AT so tem a "
          "89-16 e a lista traz 89-20/22/24). Criar a chave/medidor que falta."),
    "E": ("Medida de tensao sem TP",
          "O vao tem TC mas nao tem TP. Tensao/frequencia/angulo nao podem "
          "ficar num transformador de corrente."),
    "Z": ("A conferir", "Motivo nao classificado — ver a coluna ao lado."),
}


def categoria(motivo: str | None, dm: str, renomear: set) -> str:
    if motivo is None:
        return "A" if dm in renomear else "Z"
    if "ja tem esse papel" in motivo or "ja tem um sinal" in motivo:
        return "D"
    if "nao tem equivalente" in motivo:
        return "B"
    if "nao carrega sinal" in motivo:
        return "C"
    if "ja tem um sinal" in motivo:
        return "D"
    if "CURRENTTR" in motivo or "nao pode ficar num TC" in motivo:
        return "E"
    if "GENERICO" in motivo:
        return "C"
    return "Z"


def main():
    # 1) os sinais da FUTURA
    wb = openpyxl.load_workbook(FUT, read_only=True, data_only=True)
    sinais = []
    for sn in ("DNP3_DiscreteSignals", "DNP3_AnalogSignals",
               "DNP3_DiscreteAnalog"):
        if sn not in wb.sheetnames:
            continue
        linhas = list(wb[sn].iter_rows(values_only=True))
        hdr = list(linhas[HR - 1])
        ix = {n: i for i, n in enumerate(hdr) if n}
        for r in linhas[HR:]:
            nome = r[ix["Signal Name"]]
            if not nome:
                continue
            sinais.append({
                "nome": str(nome),
                "aba": sn.replace("DNP3_", ""),
                "dm": str(r[ix["Device Mapping"]] or ""),
                "desc": str(r[ix.get("Description", 0)] or "")
                        if "Description" in ix else "",
                "in": r[ix["Input Coordinates"]]
                      if "Input Coordinates" in ix else "",
                "out": r[ix.get("Output Coordinates", 0)]
                       if "Output Coordinates" in ix else "",
            })
    wb.close()

    # 1b) onde a ATUAL_COMPLETA colocou cada um (o _PROT do vao ou o disjuntor)
    completa = {}
    pc = DOWN / "TDT_CASCA_ATUAL_COMPLETA.xlsx"
    if pc.exists():
        wc = openpyxl.load_workbook(pc, read_only=True, data_only=True)
        for sn in ("DNP3_DiscreteSignals", "DNP3_AnalogSignals",
                   "DNP3_DiscreteAnalog"):
            if sn not in wc.sheetnames:
                continue
            ls = list(wc[sn].iter_rows(values_only=True))
            ix = {n: i for i, n in enumerate(ls[HR - 1]) if n}
            for r in ls[HR:]:
                if r[ix["Signal Name"]]:
                    completa[r[ix["Signal Name"]]] = r[ix["Device Mapping"]]
        wc.close()

    # 2) o motivo, do relatorio
    wr = openpyxl.load_workbook(REL, read_only=True, data_only=True)
    motivo = {}
    for r in list(wr["19-Sem dispositivo no modelo"].iter_rows(
            values_only=True))[1:]:
        if r[4]:
            motivo[str(r[4])] = str(r[6] or "")
    renomear = {str(r[0]) for r in list(wr["21-FALTA renomear no modelo"]
                                        .iter_rows(values_only=True))[1:] if r[0]}
    wr.close()

    # Nem todo sinal da FUTURA passa pela aba 19: os barrados por PAPEL ja
    # ocupado (a funcao ANSI daquele rele, a grandeza+fase daquele TP) tem
    # dispositivo valido e nunca entraram na lista de "sem dispositivo".
    # Sem este fallback eles apareciam como "A conferir" — 73 de 267.
    import casca_devmap as _dv
    _ids = _dv.modelo_new()[0]
    for s in sinais:
        m = motivo.get(s["nome"])
        s["cat"] = categoria(m, s["dm"], renomear)
        s["motivo"] = m or ("o dispositivo existe no modelo mas ainda sem _NEW"
                            if s["cat"] == "A" else "")
        if s["cat"] == "Z":
            if s["dm"] in _ids:
                s["cat"] = "D"
                s["motivo"] = (f"{s['dm']} EXISTE, mas o papel deste sinal ja "
                               f"esta ocupado nele (a funcao ANSI, ou a "
                               f"grandeza+fase). Precisa do estagio/dispositivo "
                               f"proprio da funcao.")
            else:
                s["cat"] = "C"
                s["motivo"] = f"{s['dm']} nao existe no modelo — precisa ser criado"

    # 3) planilha
    wb = openpyxl.Workbook()
    hdrf = Font(bold=True, color="FFFFFF")
    hdrfill = PatternFill("solid", fgColor="1F4E78")
    cor = {"A": "C6EFCE", "B": "FFC7CE", "C": "FFEB9C",
           "D": "D9D9D9", "E": "BDD7EE", "Z": "FFFFFF"}

    ws = wb.active
    ws.title = "0-RESUMO"
    c = collections.Counter(s["cat"] for s in sinais)
    vao = collections.Counter(s["nome"].split("_")[1] for s in sinais)
    linhas = [
        [f"POR QUE OS {len(sinais)} SINAIS DA TDT FUTURA NAO ESTAO NA ATUAL"], [],
        ["Nenhum sinal da FUTURA esta ERRADO. Todos tem nome, index DNP3,"],
        ["escala, descricao e comando corretos — sairam da mesma lista e"],
        ["passaram pelas mesmas conferencias dos sinais da TDT ATUAL."], [],
        ["O que falta e o DISPOSITIVO EXATO no Cas_Obra. No ADMS todo sinal"],
        ["precisa apontar para um equipamento que existe; nao ha 'sinal solto'"],
        ["(ja testamos apontar para a propria UTR e o ADMS recusa igual)."], [],
        [f"MAS {sum(1 for s in sinais if completa.get(s['nome']))} DELES JA TEM"
         " DESTINO NA TDT_CASCA_ATUAL_COMPLETA: o que e protecao vai para o"],
        ["_PROT generico do vao e o que e solto vai para o disjuntor. Veja a"],
        ["coluna 'JA RESOLVIDO na ATUAL_COMPLETA?' da aba 1."], [],
        ["Cat", "Sinais", "%", "Situacao", "O que fazer no modelo"],
    ]
    tot = len(sinais)
    for k in "ABCDEZ":
        if c[k]:
            linhas.append([k, c[k], f"{100*c[k]/tot:.1f}%",
                           ACAO[k][0], ACAO[k][1]])
    linhas += [[], ["SINAIS POR VAO"], ["Vao", "Sinais"]]
    linhas += [[v, n] for v, n in vao.most_common()]
    for r in linhas:
        ws.append(r)
    for row in ws.iter_rows():
        for cel in row:
            if cel.value in ("Cat", "Vao") and cel.column == 1:
                for x in row:
                    x.font = hdrf; x.fill = hdrfill
    ws["A1"].font = Font(bold=True, size=13)
    for col, w in zip("ABCDE", (10, 10, 8, 30, 78)):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["E"].width = 78

    ws = wb.create_sheet("1-SINAL A SINAL")
    cabec = ["#", "Sinal", "Vao", "SIGLA", "Aba", "Descricao",
             "Input", "Output", "Device Mapping que falta", "Cat",
             "Situacao", "JA RESOLVIDO na ATUAL_COMPLETA?",
             "Onde entrou na ATUAL_COMPLETA", "Motivo exato do gerador",
             "O que fazer"]
    ws.append(cabec)
    for cel in ws[1]:
        cel.font = hdrf; cel.fill = hdrfill
        cel.alignment = Alignment(vertical="center", wrap_text=True)
    for i, s in enumerate(sorted(sinais, key=lambda s: (s["cat"], s["nome"])), 1):
        pp = s["nome"].split("_")
        alvo = completa.get(s["nome"], "")
        ws.append([i, s["nome"], pp[1] if len(pp) > 1 else "",
                   "_".join(pp[3:]) if len(pp) > 3 else "",
                   s["aba"], s["desc"], s["in"], s["out"], s["dm"], s["cat"],
                   ACAO[s["cat"]][0], "SIM" if alvo else "nao", alvo,
                   s["motivo"], ACAO[s["cat"]][1]])
        f = PatternFill("solid", fgColor=cor[s["cat"]])
        for j in (10, 11):
            ws.cell(ws.max_row, j).fill = f
        ws.cell(ws.max_row, 12).fill = PatternFill(
            "solid", fgColor="C6EFCE" if alvo else "FFC7CE")
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:O{ws.max_row}"
    for col, w in zip("ABCDEFGHIJKLMNO",
                      (5, 30, 9, 10, 16, 42, 8, 8, 34, 5, 26, 12, 32, 80, 70)):
        ws.column_dimensions[col].width = w

    # ── o que sobrou: nem a ATUAL_COMPLETA achou lugar ────────────────────
    # Aqui nao ha jeito por mapeamento: o dispositivo tem que ser CRIADO.
    # O ADMS aceita no maximo 5 sinais 'RelayTrip' por dispositivo (medido em
    # dois imports, ver data/cota_import.json), entao encher o _PROT generico
    # ou o disjuntor tem teto.
    ws = wb.create_sheet("3-CRIAR (o que sobrou)")
    ws.append(["Dispositivo a CRIAR no Cas_Obra", "Sinais que dependem", "Vao",
               "Tipo", "Exemplos de sinal"])
    for cel in ws[1]:
        cel.font = hdrf; cel.fill = hdrfill
    falta = collections.defaultdict(list)
    for s in sinais:
        if not completa.get(s["nome"]):
            falta[s["dm"]].append(s["nome"])
    for dm, ns in sorted(falta.items(), key=lambda x: -len(x[1])):
        pp = dm.split("_")
        t = "Rele de protecao" if "_PROT" in dm else (
            "Disjuntor" if dm.endswith(("_DJ", "_DJ_NEW")) else
            "Seccionadora" if "_SEC" in dm else
            "Transformador de potencial" if "_TP" in dm else
            "Transformador de corrente" if "_TC" in dm else "?")
        ws.append([dm, len(ns), pp[1] if len(pp) > 1 else "", t,
                   ", ".join(ns[:3])])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    for col, w in zip("ABCDE", (38, 9, 10, 26, 60)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("2-POR DISPOSITIVO")
    ws.append(["Device Mapping que falta", "Sinais", "Vao", "Cat", "Situacao"])
    for cel in ws[1]:
        cel.font = hdrf; cel.fill = hdrfill
    por = collections.Counter(s["dm"] for s in sinais)
    catd = {s["dm"]: s["cat"] for s in sinais}
    for dm, n in por.most_common():
        pp = dm.split("_")
        ws.append([dm, n, pp[1] if len(pp) > 1 else "", catd[dm],
                   ACAO[catd[dm]][0]])
        ws.cell(ws.max_row, 4).fill = PatternFill("solid",
                                                  fgColor=cor[catd[dm]])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    for col, w in zip("ABCDE", (36, 9, 10, 6, 28)):
        ws.column_dimensions[col].width = w

    # se o arquivo estiver aberto no Excel, grava ao lado como _NOVA
    alvo = OUT
    try:
        wb.save(alvo)
    except PermissionError:
        alvo = OUT.with_name(OUT.stem + "_NOVA.xlsx")
        wb.save(alvo)
        print(f"ATENCAO: {OUT.name} estava aberto no Excel — gravado como "
              f"{alvo.name}")
    # resave_native trabalha com BYTES, nao com caminho
    alvo.write_bytes(excel_native.resave_native(alvo.read_bytes()))
    print(f"{alvo.name}: {len(sinais)} sinais, {len(por)} dispositivos")
    for k in "ABCDEZ":
        if c[k]:
            print(f"  {k}  {c[k]:4} sinais  {ACAO[k][0]}")


if __name__ == "__main__":
    main()
